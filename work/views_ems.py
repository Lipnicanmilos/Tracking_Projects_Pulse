from django.shortcuts import render, redirect
from django.http import HttpResponse   # added
from django.contrib.auth.decorators import login_required
from project.settings import cx_Oracle
from .models import BILLDB, EMS_DWHDB
import mimetypes

import paramiko
from paramiko import SSHClient, SSHConfig
from zipfile import ZipFile
import os, shutil

import datetime
from datetime import date, time
import pandas as pd
from xlsxwriter import Workbook
import zipfile
from django.http import FileResponse
from io import BytesIO

from django.contrib import messages
from dateutil.relativedelta import relativedelta

from django.core.mail import EmailMessage

#::  PDF
from datetime import datetime
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm

from django.http import HttpResponseRedirect
from django.urls import reverse

########### lego report code
import codecs
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.units import mm

from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from dateutil.relativedelta import relativedelta
from django.core.files import File

# Define SFTP server's hostname, port, username, and password
hostname = 'logarch.etc.example'
port = 22
username = 'app_user'
# password = 'Lip25nican06!'
password = 'Lip27nican34'


import json

from .forms import BILLDBUpdateForm, EMSDWHDBUpdateForm


def dictfetchall(cursor):
    "Return all rows from a cursor as a dict"
    columns = [col[0] for col in cursor.description]
    return [
        dict(zip(columns, row))
        for row in cursor.fetchall()
    ]

@login_required
def home_ems(request):
    bill_db=BILLDB.objects.get(user=request.user)
    form_bill = BILLDBUpdateForm(request.POST or None, instance=bill_db)
    
    if form_bill.is_valid():
        form_bill.save()
        messages.success(request,'Your Profile has been updated!')
        form_bill = BILLDBUpdateForm(instance=bill_db)
    else:
        form_bill = BILLDBUpdateForm(instance=bill_db)

    ems_dwh_db=EMS_DWHDB.objects.get(user=request.user)
    form_ems_dwh = EMSDWHDBUpdateForm(request.POST or None, instance=ems_dwh_db)
    if form_ems_dwh.is_valid():
        form_ems_dwh.save()
        messages.success(request,'Your Profile has been updated!')
        form_ems_dwh = EMSDWHDBUpdateForm(instance=ems_dwh_db)
    else:
        form_ems_dwh = EMSDWHDBUpdateForm(instance=ems_dwh_db)

    context = { 'form_bill':form_bill, 'form_ems_dwh':form_ems_dwh}

    return render(request, 'base_ems.html', context)
# def home_ems(request):
#     bill_db=BILLDB.objects.get(user=request.user)
#     form_bill = BILLDBUpdateForm(request.POST, instance=bill_db)
    

#     ems_dwh_db=EMS_DWHDB.objects.get(user=request.user)
#     form_ems_dwh = EMSDWHDBUpdateForm(request.POST, instance=ems_dwh_db)
#     if form_bill.is_valid():
#         form_bill.save()
#         messages.success(request,'Your Profile has been updated!')
#         form_bill = BILLDBUpdateForm(instance=bill_db)
#     elif form_ems_dwh.is_valid():
#         form_ems_dwh.save()
#         messages.success(request,'Your Profile has been updated!')
#         form_ems_dwh = EMSDWHDBUpdateForm(instance=ems_dwh_db)
#     else:
#         form_bill = BILLDBUpdateForm(instance=bill_db)
#         form_ems_dwh = EMSDWHDBUpdateForm(instance=ems_dwh_db)   
     


#     context = { 'form_bill':form_bill, 'form_ems_dwh':form_ems_dwh}   

#     return render(request, 'base_ems.html', context)

def reports_View(request):
    return render(request, 'ems/reports_View.html')

@login_required
def validacia(request):

    context = {}
    return render(request, 'ems/validacia_EMS.html', context)

@login_required
def jobs_ems(request):

 

    billdb=BILLDB.objects.get(user=request.user)
    decrypted = billdb.BILLDB_password
    print('b')
    print(decrypted)
    
    con_BILLDB_STDBY = cx_Oracle.connect(billdb.BILLDB_username+"/"+billdb.BILLDB_password+"@"+billdb.BILLDB_hostname+":"+billdb.BILLDB_port+"/"+billdb.BILLDB_servicename)
    # cursor = con_CISDB_STDBY.cursor()

    job_s = ''
    # print('vysledok: '+job_s)\
    q = None

    sql_job = """SELECT * FROM(
                    SELECT CAST(FROM_TZ(CAST(ACTION_DATETIME AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET' AS DATE) as ACTION_DATETIME, ji.SCHEDULED_JOB_ID, JOB_NAME,
                    EXECUTION_STATE_NAME, EXPECTED_NUMBER_OF_ITEMS, SUCCESSFULLY_PROCESSED_ITEMS, UNSUCCESSFULLY_PROCESSED_ITEMS
                    FROM BILLIEN_CO.JOB_INSTANCE ji
                    left join BILLIEN_CO.EXECUTION_STATE_L exs on ji.EXECUTION_STATE_CODE=exs.EXECUTION_STATE_CODE
                    left join BILLIEN_CO.SCHEDULED_JOB sj on ji.SCHEDULED_JOB_ID=sj.SCHEDULED_JOB_ID
                    where
                    exs.LANGUAGE_CODE='SK' and START_ACTION_DATETIME >= trunc(sysdate-3)
                                   --and ji.SCHEDULED_JOB_ID=143
                    and ji.EXECUTION_STATE_CODE!=4
                    --and UNSUCCESSFULLY_PROCESSED_ITEMS>0
                    or EXPECTED_NUMBER_OF_ITEMS !=SUCCESSFULLY_PROCESSED_ITEMS and exs.LANGUAGE_CODE='SK' and START_ACTION_DATETIME >= trunc(sysdate-3)
                    order by START_ACTION_DATETIME desc)
                    --where rownum <=8
                    """
    with con_BILLDB_STDBY.cursor() as cursor:
        cursor.execute(sql_job)
            #rows = cursor.fetchall()
            #rows = namedtuplefetchall(cursor)
        jobs_error = dictfetchall(cursor)

    print(q)

    if 'q' in request.GET:
        q=request.GET['q']

        sql_ROLE_NAME ="""SELECT CAST(FROM_TZ(CAST(ACTION_DATETIME AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET' AS DATE) as ACTION_DATETIME, ji.SCHEDULED_JOB_ID, JOB_NAME, EXECUTION_STATE_NAME, EXPECTED_NUMBER_OF_ITEMS, SUCCESSFULLY_PROCESSED_ITEMS, UNSUCCESSFULLY_PROCESSED_ITEMS
                                FROM BILLIEN_CO.JOB_INSTANCE ji
                                left join BILLIEN_CO.EXECUTION_STATE_L exs on ji.EXECUTION_STATE_CODE=exs.EXECUTION_STATE_CODE
                                left join BILLIEN_CO.SCHEDULED_JOB sj on ji.SCHEDULED_JOB_ID=sj.SCHEDULED_JOB_ID
                                where
                                exs.LANGUAGE_CODE='SK'
                                and START_ACTION_DATETIME >= trunc(sysdate-7)
                                and JOB_NAME like :id
                                order by START_ACTION_DATETIME desc"""
        with con_BILLDB_STDBY.cursor() as cursor:
            c = '%' + q + '%'
            cursor.execute(sql_ROLE_NAME, id=c)
            #rows = cursor.fetchall()
            #rows = namedtuplefetchall(cursor)
            job_s = dictfetchall(cursor)
            # print(job_s)
    # cursor.close()
    # connection.close()

    context = { 'jobe': jobs_error, 'job_s':job_s, 'q':q } #'job_s':job_s,
    return render(request,'ems/jobs_ems.html', context)

@login_required
def ems_logs(request):
    # remove these print statements later
    # if request.method == "GET":
    #     os.startfile(r'\\fileserver.internal.example.com\billien\CIS')

    # if request.method == 'POST':
    #     os.startfile(r'\\fileserver.internal.example.com\billien\CIS')

    return render(request, 'ems/ems_logs.html') #{'fname' : request.user.username }

def EMSAAPP01_realtime_error(request):
    # Create an SSH client object
    ssh = paramiko.SSHClient()

    # Automatically add the server's host key (you can disable this if you want to manually check the host key)
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # Connect to the SFTP server
    ssh.connect(hostname, port, username, password)

    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # delete dir
    dir = filepath = BASE_DIR + '/ems_logs/EMSAAPP01/'
    isExist = os.path.exists(dir)
    if isExist:
        shutil.rmtree(dir)
        directory = "EMSAAPP01"
        parent_dir = BASE_DIR + '/ems_logs/'
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    else:
        # create dir
        directory = "EMSAAPP01"
        parent_dir = BASE_DIR + '/ems_logs/'
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    # Open an SFTP session
    sftp = ssh.open_sftp()
    print('pristup')
    print(sftp)

    # Define the remote directory's path
    remote_directory_path = '/srv/archive/log/upload-by-smbfs/billien-ems/temp/' #'/srv/archive/log/upload-by-smbfs/billien-edz/realtime/EDZISAPP02/'
    print('ulozisko'+remote_directory_path)

    # Define the local directory's path
    local_directory_path = BASE_DIR + '/ems_logs/EMSAAPP01/'
    # local_directory_path = os.listdir(r'\fileserver.internal.example.com\c$\Users\app_user\Downloads')
    print('localne ulozisko'+local_directory_path)

    # Get a list of all files in the remote directory
    files = sftp.listdir(remote_directory_path)
    print('sftp files')
    print(files)

    # Iterate over each file in the remote directory
    for file in files:
        # If the file's name contains "error", copy it from the SFTP server to the local machine
        # if 'EMSISAPP0' in file:
        if 'EMSAAPP01_real' in file:
            print('files')
            print(f'Copying {file}')
            # print(file)

            # Define the remote file's path
            remote_file_path = remote_directory_path+file
            print('remote_file_path')
            print(remote_file_path)
            # Define the local file's path
            local_file_path = local_directory_path+file
            print('local_file_path')
            print(local_file_path)
            # Copy the file from the SFTP server to the local machine
            # ulozi zip subor
            sftp.get(remote_file_path, local_file_path)

            # Ulozim si nazov priecinkov ktore su odzipovane
            unzip_file = local_file_path[:-4]
            print(unzip_file)

            # rozbali zip subor bez poslednych troch pismen
            zf = ZipFile(local_file_path, 'r')
            zf.extractall(local_file_path[:-4])
            zf.close()

            # vymaze zip subor
            print('vymazanie suboru')
            print(local_file_path)
            if os.path.isfile(local_file_path):
                os.remove(local_file_path)

    # Close the SFTP session
    sftp.close()

    # Close the SSH client object
    ssh.close()

    # create txt file
    folder = filepath = BASE_DIR + '/ems_logs/EMSAAPP01/'
    try:
        with open(folder + 'EMSAAPP01_real_ErrorLogs.txt', 'w') as f:
            f.write('---------------------------------------------  ALL ErrorLogs from EMSAAPP01  ----------\n')
            print("Create is_log.txt")
    except FileNotFoundError:
        print("The 'docs' directory does not exist")
    folder_emsaapp01_log = BASE_DIR + '/ems_logs/EMSAAPP01/EMSAAPP01_real_ErrorLogs.txt'
    search_folder = filepath = BASE_DIR + '/ems_logs/EMSAAPP01/EMSAAPP01_realtime/'

    # append log in text
    for filename in os.listdir(search_folder):
        # if filename.endswith('.log'):       # konciaci na .log
        # if filename.startswith('We', 7, 9):   # we je na 7 az 9 mieste
        if 'Error' in str(filename):
    #         shutil.copy( source_folder + filename, destination_folder)
    #         # zapis do suboru web log txt
            with open(( search_folder + filename), 'r', encoding="utf-8") as reader, open(folder_emsaapp01_log,'a', encoding="utf-8") as a_writer:
                dog_breeds = reader.readlines()
                a_writer.write('\n')
                a_writer.write('\n')
                a_writer.write('\nEMSAAPP01//  ')
                a_writer.writelines(filename)
                a_writer.write(' \n')
                a_writer.write('  \n')
                a_writer.writelines((dog_breeds))
                a_writer.close()
                # koniec zapisu
                print('copied /EMSAAPP01_ErrorLogs/', filename )

    # ----download file ------------------
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    filename = 'EMSAAPP01_real_ErrorLogs.txt'
    # Define the full file path
    filepath = BASE_DIR + '/ems_logs/EMSAAPP01/' + filename
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    if isExist:
        shutil.rmtree(search_folder)
    # Return the response value
    return response

def EMSAAPP01_realtime_app(request):
    # Create an SSH client object
    ssh = paramiko.SSHClient()

    # Automatically add the server's host key (you can disable this if you want to manually check the host key)
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # Connect to the SFTP server
    ssh.connect(hostname, port, username, password)

    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # delete dir
    dir = filepath = BASE_DIR + '/ems_logs/EMSAAPP01/'

    isExist = os.path.exists(dir)
    if isExist:
        shutil.rmtree(dir)
        directory = "EMSAAPP01"
        parent_dir = BASE_DIR + '/ems_logs/'
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    else:
        # create dir
        directory = "EMSAAPP01"
        parent_dir = BASE_DIR + '/ems_logs/'
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    # Open an SFTP session
    sftp = ssh.open_sftp()
    print('pristup')
    print(sftp)

    # Define the remote directory's path
    remote_directory_path = '/srv/archive/log/upload-by-smbfs/billien-ems/temp/' #'/srv/archive/log/upload-by-smbfs/billien-edz/realtime/EDZISAPP02/'
    print('ulozisko'+remote_directory_path)

    # Define the local directory's path
    local_directory_path = BASE_DIR + '/ems_logs/EMSAAPP01/'
    # local_directory_path = os.listdir(r'\fileserver.internal.example.com\c$\Users\app_user\Downloads')
    print('localne ulozisko'+local_directory_path)

    # Get a list of all files in the remote directory
    files = sftp.listdir(remote_directory_path)
    print('sftp files')
    print(files)

    # Iterate over each file in the remote directory
    for file in files:
        # If the file's name contains "error", copy it from the SFTP server to the local machine
        # if 'EMSISAPP0' in file:
        if 'EMSAAPP01_real' in file:
            print('files')
            print(f'Copying {file}')
            # print(file)

            # Define the remote file's path
            remote_file_path = remote_directory_path+file
            print('remote_file_path')
            print(remote_file_path)
            # Define the local file's path
            local_file_path = local_directory_path+file
            print('local_file_path')
            print(local_file_path)
            # Copy the file from the SFTP server to the local machine
            # ulozi zip subor
            sftp.get(remote_file_path, local_file_path)

            # Ulozim si nazov priecinkov ktore su odzipovane
            unzip_file = local_file_path[:-4]
            print(unzip_file)

            # rozbali zip subor bez poslednych troch pismen
            zf = ZipFile(local_file_path, 'r')
            zf.extractall(local_file_path[:-4])
            zf.close()

            # vymaze zip subor
            print('vymazanie suboru')
            print(local_file_path)
            if os.path.isfile(local_file_path):
                os.remove(local_file_path)

    # Close the SFTP session
    sftp.close()

    # Close the SSH client object
    ssh.close()

    # create txt file
    folder = filepath = BASE_DIR + '/ems_logs/EMSAAPP01/'
    try:
        with open(folder + 'EMSAAPP01_real_APPLogs.txt', 'w') as f:
            f.write('---------------------------------------------  ALL AppLogs from EMSAAPP01  ----------\n')
            print("Create is_log.txt")
    except FileNotFoundError:
        print("The 'docs' directory does not exist")
    folder_emsaapp01_log = BASE_DIR + '/ems_logs/EMSAAPP01/EMSAAPP01_real_AppLogs.txt'
    search_folder = filepath = BASE_DIR + '/ems_logs/EMSAAPP01/EMSAAPP01_realtime/'

    # append log in text
    for filename in os.listdir(search_folder):
        # if filename.endswith('.log'):       # konciaci na .log
        # if filename.startswith('We', 7, 9):   # we je na 7 az 9 mieste
        if 'App' in str(filename):
    #         shutil.copy( source_folder + filename, destination_folder)
    #         # zapis do suboru web log txt
            with open(( search_folder + filename), 'r', encoding="utf-8") as reader, open(folder_emsaapp01_log,'a', encoding="utf-8") as a_writer:
                dog_breeds = reader.readlines()
                a_writer.write('\n')
                a_writer.write('\n')
                a_writer.write('\nEMSAAPP01//  ')
                a_writer.writelines(filename)
                a_writer.write(' \n')
                a_writer.write('  \n')
                a_writer.writelines((dog_breeds))
                a_writer.close()
                # koniec zapisu
                print('copied /EMSAAPP01_AppLogs/', filename )

    # ----download file ------------------
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    filename = 'EMSAAPP01_real_AppLogs.txt'
    # Define the full file path
    filepath = BASE_DIR + '/ems_logs/EMSAAPP01/' + filename
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    if isExist:
        shutil.rmtree(search_folder)
    # Return the response value
    return response

def EMSAAPP02_realtime_error(request):

    # Create an SSH client object
    ssh = paramiko.SSHClient()

    # Automatically add the server's host key (you can disable this if you want to manually check the host key)
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # Connect to the SFTP server
    ssh.connect(hostname, port, username, password)

    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # delete dir

    top_storage = '/ems_logs/'
    storage = 'EMSAAPP02/'
    search_file = 'EMSAAPP02_realtime'

    dir = filepath = BASE_DIR + top_storage + storage
    isExist = os.path.exists(dir)
    if isExist:
        shutil.rmtree(dir)
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    else:
        # create dir
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    # Open an SFTP session
    sftp = ssh.open_sftp()
    print('pristup')
    print(sftp)

    # Define the remote directory's path
    remote_directory_path = '/srv/archive/log/upload-by-smbfs/billien-ems/temp/' #'/srv/archive/log/upload-by-smbfs/billien-edz/realtime/EDZISAPP02/'
    print('ulozisko'+remote_directory_path)

    # Define the local directory's path
    local_directory_path = BASE_DIR + top_storage + storage
    # local_directory_path = os.listdir(r'\fileserver.internal.example.com\c$\Users\app_user\Downloads')
    print('localne ulozisko'+local_directory_path)

    # Get a list of all files in the remote directory
    files = sftp.listdir(remote_directory_path)
    print('sftp files')
    print(files)

    # Iterate over each file in the remote directory
    for file in files:
        # If the file's name contains "error", copy it from the SFTP server to the local machine
        # if 'EMSISAPP0' in file:
        if search_file in file:
            print('files')
            print(f'Copying {file}')
            # print(file)

            # Define the remote file's path
            remote_file_path = remote_directory_path+file
            print('remote_file_path')
            print(remote_file_path)
            # Define the local file's path
            local_file_path = local_directory_path+file
            print('local_file_path')
            print(local_file_path)
            # Copy the file from the SFTP server to the local machine
            # ulozi zip subor
            sftp.get(remote_file_path, local_file_path)

            # Ulozim si nazov priecinkov ktore su odzipovane
            unzip_file = local_file_path[:-4]
            print(unzip_file)

            # rozbali zip subor bez poslednych troch pismen
            zf = ZipFile(local_file_path, 'r')
            zf.extractall(local_file_path[:-4])
            zf.close()

            # vymaze zip subor
            print('vymazanie suboru')
            print(local_file_path)
            if os.path.isfile(local_file_path):
                os.remove(local_file_path)

    # Close the SFTP session
    sftp.close()

    # Close the SSH client object
    ssh.close()

    # create txt file
    folder = filepath = BASE_DIR + top_storage + storage
    create_files = search_file+'.txt'
    print (create_files)
    try:
        with open(folder + create_files, 'w') as f:
            f.write('---------------------------------------------  ALL ErrorLogs from '+storage+' ----------\n')
            print("Create is_log.txt")
    except FileNotFoundError:
        print("The 'docs' directory does not exist")
    folder_emsaapp01_log = BASE_DIR + top_storage + storage +create_files
    search_folder = filepath = BASE_DIR + top_storage + storage + search_file +'/'

    # append log in text
    for filename in os.listdir(search_folder):
        # if filename.endswith('.log'):       # konciaci na .log
        # if filename.startswith('We', 7, 9):   # we je na 7 az 9 mieste
        if 'Error' in str(filename):
    #         shutil.copy( source_folder + filename, destination_folder)
    #         # zapis do suboru web log txt
            with open(( search_folder + filename), 'r', encoding="utf-8") as reader, open(folder_emsaapp01_log,'a', encoding="utf-8") as a_writer:
                dog_breeds = reader.readlines()
                a_writer.write('\n')
                a_writer.write('\n')
                a_writer.write('\n '+storage+'/  ')
                a_writer.writelines(filename)
                a_writer.write(' \n')
                a_writer.write('  \n')
                a_writer.writelines((dog_breeds))
                a_writer.close()
                # koniec zapisu
                # print('copied /EMSAAPP02_ErrorLogs/', filename )

    # ----download file ------------------
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    filename = create_files
    # Define the full file path
    filepath = BASE_DIR + top_storage + storage + filename
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    if isExist:
        shutil.rmtree(search_folder)
    # Return the response value
    return response

def EMSAAPP02_realtime_app(request):
    # Create an SSH client object
    ssh = paramiko.SSHClient()

    # Automatically add the server's host key (you can disable this if you want to manually check the host key)
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # Connect to the SFTP server
    ssh.connect(hostname, port, username, password)

    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # delete dir

    top_storage = '/ems_logs/'
    storage = 'EMSAAPP02/'
    search_file = 'EMSAAPP02_realtime'

    dir = filepath = BASE_DIR + top_storage + storage
    isExist = os.path.exists(dir)
    if isExist:
        shutil.rmtree(dir)
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    else:
        # create dir
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    # Open an SFTP session
    sftp = ssh.open_sftp()
    print('pristup')
    print(sftp)

    # Define the remote directory's path
    remote_directory_path = '/srv/archive/log/upload-by-smbfs/billien-ems/temp/' #'/srv/archive/log/upload-by-smbfs/billien-edz/realtime/EDZISAPP02/'
    print('ulozisko'+remote_directory_path)

    # Define the local directory's path
    local_directory_path = BASE_DIR + top_storage + storage
    # local_directory_path = os.listdir(r'\fileserver.internal.example.com\c$\Users\app_user\Downloads')
    print('localne ulozisko'+local_directory_path)

    # Get a list of all files in the remote directory
    files = sftp.listdir(remote_directory_path)
    print('sftp files')
    print(files)

    # Iterate over each file in the remote directory
    for file in files:
        # If the file's name contains "error", copy it from the SFTP server to the local machine
        # if 'EMSISAPP0' in file:
        if search_file in file:
            print('files')
            print(f'Copying {file}')
            # print(file)

            # Define the remote file's path
            remote_file_path = remote_directory_path+file
            print('remote_file_path')
            print(remote_file_path)
            # Define the local file's path
            local_file_path = local_directory_path+file
            print('local_file_path')
            print(local_file_path)
            # Copy the file from the SFTP server to the local machine
            # ulozi zip subor
            sftp.get(remote_file_path, local_file_path)

            # Ulozim si nazov priecinkov ktore su odzipovane
            unzip_file = local_file_path[:-4]
            print(unzip_file)

            # rozbali zip subor bez poslednych troch pismen
            zf = ZipFile(local_file_path, 'r')
            zf.extractall(local_file_path[:-4])
            zf.close()

            # vymaze zip subor
            print('vymazanie suboru')
            print(local_file_path)
            if os.path.isfile(local_file_path):
                os.remove(local_file_path)

    # Close the SFTP session
    sftp.close()

    # Close the SSH client object
    ssh.close()

    # create txt file
    folder = filepath = BASE_DIR + top_storage + storage
    create_files = search_file+'.txt'
    print (create_files)
    try:
        with open(folder + create_files, 'w') as f:
            f.write('---------------------------------------------  ALL APPLogs from '+storage+' ----------\n')
            print("Create is_log.txt")
    except FileNotFoundError:
        print("The 'docs' directory does not exist")
    folder_emsaapp01_log = BASE_DIR + top_storage + storage +create_files
    search_folder = filepath = BASE_DIR + top_storage + storage + search_file +'/'

    # append log in text
    for filename in os.listdir(search_folder):
        # if filename.endswith('.log'):       # konciaci na .log
        # if filename.startswith('We', 7, 9):   # we je na 7 az 9 mieste
        if 'App' in str(filename):
    #         shutil.copy( source_folder + filename, destination_folder)
    #         # zapis do suboru web log txt
            with open(( search_folder + filename), 'r', encoding="utf-8") as reader, open(folder_emsaapp01_log,'a', encoding="utf-8") as a_writer:
                dog_breeds = reader.readlines()
                a_writer.write('\n')
                a_writer.write('\n')
                a_writer.write('\n '+storage+'/  ')
                a_writer.writelines(filename)
                a_writer.write(' \n')
                a_writer.write('  \n')
                a_writer.writelines((dog_breeds))
                a_writer.close()
                # koniec zapisu
                # print('copied /EMSAAPP02_ErrorLogs/', filename )

    # ----download file ------------------
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    filename = create_files
    # Define the full file path
    filepath = BASE_DIR + top_storage + storage + filename
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    if isExist:
        shutil.rmtree(search_folder)
    # Return the response value
    return response

def EMSBAPP01_realtime_error(request):

    # Create an SSH client object
    ssh = paramiko.SSHClient()

    # Automatically add the server's host key (you can disable this if you want to manually check the host key)
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # Connect to the SFTP server
    ssh.connect(hostname, port, username, password)

    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # delete dir

    top_storage = '/ems_logs/'
    storage = 'EMSBAPP01/'
    search_file = 'EMSBAPP01_realtime'
    name = 'Error'

    dir = filepath = BASE_DIR + top_storage + storage
    isExist = os.path.exists(dir)
    if isExist:
        shutil.rmtree(dir)
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    else:
        # create dir
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    # Open an SFTP session
    sftp = ssh.open_sftp()
    print('pristup')
    print(sftp)

    # Define the remote directory's path
    remote_directory_path = '/srv/archive/log/upload-by-smbfs/billien-ems/temp/' #'/srv/archive/log/upload-by-smbfs/billien-edz/realtime/EDZISAPP02/'
    print('ulozisko'+remote_directory_path)

    # Define the local directory's path
    local_directory_path = BASE_DIR + top_storage + storage
    # local_directory_path = os.listdir(r'\fileserver.internal.example.com\c$\Users\app_user\Downloads')
    print('localne ulozisko'+local_directory_path)

    # Get a list of all files in the remote directory
    files = sftp.listdir(remote_directory_path)
    print('sftp files')
    print(files)

    # Iterate over each file in the remote directory
    for file in files:
        # If the file's name contains "error", copy it from the SFTP server to the local machine
        # if 'EMSISAPP0' in file:
        if search_file in file:
            print('files')
            print(f'Copying {file}')
            # print(file)

            # Define the remote file's path
            remote_file_path = remote_directory_path+file
            print('remote_file_path')
            print(remote_file_path)
            # Define the local file's path
            local_file_path = local_directory_path+file
            print('local_file_path')
            print(local_file_path)
            # Copy the file from the SFTP server to the local machine
            # ulozi zip subor
            sftp.get(remote_file_path, local_file_path)

            # Ulozim si nazov priecinkov ktore su odzipovane
            unzip_file = local_file_path[:-4]
            print(unzip_file)

            # rozbali zip subor bez poslednych troch pismen
            zf = ZipFile(local_file_path, 'r')
            zf.extractall(local_file_path[:-4])
            zf.close()

            # vymaze zip subor
            print('vymazanie suboru')
            print(local_file_path)
            if os.path.isfile(local_file_path):
                os.remove(local_file_path)

    # Close the SFTP session
    sftp.close()

    # Close the SSH client object
    ssh.close()

    # create txt file
    folder = filepath = BASE_DIR + top_storage + storage
    create_files = search_file+'_'+name+'.txt'
    print (create_files)
    try:
        with open(folder + create_files, 'w') as f:
            f.write('---------------------------------------------  ALL '+name+'Logs from '+storage+' ----------\n')
            print("Create is_log.txt")
    except FileNotFoundError:
        print("The 'docs' directory does not exist")
    folder_emsaapp01_log = BASE_DIR + top_storage + storage +create_files
    search_folder = filepath = BASE_DIR + top_storage + storage + search_file +'/'

    # append log in text
    for filename in os.listdir(search_folder):
        # if filename.endswith('.log'):       # konciaci na .log
        # if filename.startswith('We', 7, 9):   # we je na 7 az 9 mieste
        if name in str(filename):
    #         shutil.copy( source_folder + filename, destination_folder)
    #         # zapis do suboru web log txt
            with open(( search_folder + filename), 'r', encoding="utf-8") as reader, open(folder_emsaapp01_log,'a', encoding="utf-8") as a_writer:
                dog_breeds = reader.readlines()
                a_writer.write('\n')
                a_writer.write('\n')
                a_writer.write('\n '+storage+'/  ')
                a_writer.writelines(filename)
                a_writer.write(' \n')
                a_writer.write('  \n')
                a_writer.writelines((dog_breeds))
                a_writer.close()
                # koniec zapisu
                # print('copied /EMSAAPP02_ErrorLogs/', filename )

    # ----download file ------------------
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    filename = create_files
    # Define the full file path
    filepath = BASE_DIR + top_storage + storage + filename
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    if isExist:
        shutil.rmtree(search_folder)
    # Return the response value
    return response

def EMSBAPP01_realtime_app(request):

    # Create an SSH client object
    ssh = paramiko.SSHClient()

    # Automatically add the server's host key (you can disable this if you want to manually check the host key)
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # Connect to the SFTP server
    ssh.connect(hostname, port, username, password)

    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # delete dir

    top_storage = '/ems_logs/'
    storage = 'EMSBAPP01/'
    search_file = 'EMSBAPP01_realtime'
    name = 'App'

    dir = filepath = BASE_DIR + top_storage + storage
    isExist = os.path.exists(dir)
    if isExist:
        shutil.rmtree(dir)
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    else:
        # create dir
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    # Open an SFTP session
    sftp = ssh.open_sftp()
    print('pristup')
    print(sftp)

    # Define the remote directory's path
    remote_directory_path = '/srv/archive/log/upload-by-smbfs/billien-ems/temp/' #'/srv/archive/log/upload-by-smbfs/billien-edz/realtime/EDZISAPP02/'
    print('ulozisko'+remote_directory_path)

    # Define the local directory's path
    local_directory_path = BASE_DIR + top_storage + storage
    # local_directory_path = os.listdir(r'\fileserver.internal.example.com\c$\Users\app_user\Downloads')
    print('localne ulozisko'+local_directory_path)

    # Get a list of all files in the remote directory
    files = sftp.listdir(remote_directory_path)
    print('sftp files')
    print(files)

    # Iterate over each file in the remote directory
    for file in files:
        # If the file's name contains "error", copy it from the SFTP server to the local machine
        # if 'EMSISAPP0' in file:
        if search_file in file:
            print('files')
            print(f'Copying {file}')
            # print(file)

            # Define the remote file's path
            remote_file_path = remote_directory_path+file
            print('remote_file_path')
            print(remote_file_path)
            # Define the local file's path
            local_file_path = local_directory_path+file
            print('local_file_path')
            print(local_file_path)
            # Copy the file from the SFTP server to the local machine
            # ulozi zip subor
            sftp.get(remote_file_path, local_file_path)

            # Ulozim si nazov priecinkov ktore su odzipovane
            unzip_file = local_file_path[:-4]
            print(unzip_file)

            # rozbali zip subor bez poslednych troch pismen
            zf = ZipFile(local_file_path, 'r')
            zf.extractall(local_file_path[:-4])
            zf.close()

            # vymaze zip subor
            print('vymazanie suboru')
            print(local_file_path)
            if os.path.isfile(local_file_path):
                os.remove(local_file_path)

    # Close the SFTP session
    sftp.close()

    # Close the SSH client object
    ssh.close()

    # create txt file
    folder = filepath = BASE_DIR + top_storage + storage
    create_files = search_file+'_'+name+'.txt'
    print (create_files)
    try:
        with open(folder + create_files, 'w') as f:
            f.write('---------------------------------------------  ALL '+name+'Logs from '+storage+' ----------\n')
            print("Create is_log.txt")
    except FileNotFoundError:
        print("The 'docs' directory does not exist")
    folder_emsaapp01_log = BASE_DIR + top_storage + storage +create_files
    search_folder = filepath = BASE_DIR + top_storage + storage + search_file +'/'

    # append log in text
    for filename in os.listdir(search_folder):
        # if filename.endswith('.log'):       # konciaci na .log
        # if filename.startswith('We', 7, 9):   # we je na 7 az 9 mieste
        if name in str(filename):
    #         shutil.copy( source_folder + filename, destination_folder)
    #         # zapis do suboru web log txt
            with open(( search_folder + filename), 'r', encoding="utf-8") as reader, open(folder_emsaapp01_log,'a', encoding="utf-8") as a_writer:
                dog_breeds = reader.readlines()
                a_writer.write('\n')
                a_writer.write('\n')
                a_writer.write('\n '+storage+'/  ')
                a_writer.writelines(filename)
                a_writer.write(' \n')
                a_writer.write('  \n')
                a_writer.writelines((dog_breeds))
                a_writer.close()
                # koniec zapisu
                # print('copied /EMSAAPP02_ErrorLogs/', filename )

    # ----download file ------------------
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    filename = create_files
    # Define the full file path
    filepath = BASE_DIR + top_storage + storage + filename
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    if isExist:
        shutil.rmtree(search_folder)
    # Return the response value
    return response

def EMSBAPP02_realtime_error(request):

    # Create an SSH client object
    ssh = paramiko.SSHClient()

    # Automatically add the server's host key (you can disable this if you want to manually check the host key)
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # Connect to the SFTP server
    ssh.connect(hostname, port, username, password)

    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # delete dir

    top_storage = '/ems_logs/'
    storage = 'EMSBAPP02/'
    search_file = 'EMSBAPP02_realtime'
    name = 'Error'

    dir = filepath = BASE_DIR + top_storage + storage
    isExist = os.path.exists(dir)
    if isExist:
        shutil.rmtree(dir)
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    else:
        # create dir
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    # Open an SFTP session
    sftp = ssh.open_sftp()
    print('pristup')
    print(sftp)

    # Define the remote directory's path
    remote_directory_path = '/srv/archive/log/upload-by-smbfs/billien-ems/temp/' #'/srv/archive/log/upload-by-smbfs/billien-edz/realtime/EDZISAPP02/'
    print('ulozisko'+remote_directory_path)

    # Define the local directory's path
    local_directory_path = BASE_DIR + top_storage + storage
    # local_directory_path = os.listdir(r'\fileserver.internal.example.com\c$\Users\app_user\Downloads')
    print('localne ulozisko'+local_directory_path)

    # Get a list of all files in the remote directory
    files = sftp.listdir(remote_directory_path)
    print('sftp files')
    print(files)

    # Iterate over each file in the remote directory
    for file in files:
        # If the file's name contains "error", copy it from the SFTP server to the local machine
        # if 'EMSISAPP0' in file:
        if search_file in file:
            print('files')
            print(f'Copying {file}')
            # print(file)

            # Define the remote file's path
            remote_file_path = remote_directory_path+file
            print('remote_file_path')
            print(remote_file_path)
            # Define the local file's path
            local_file_path = local_directory_path+file
            print('local_file_path')
            print(local_file_path)
            # Copy the file from the SFTP server to the local machine
            # ulozi zip subor
            sftp.get(remote_file_path, local_file_path)

            # Ulozim si nazov priecinkov ktore su odzipovane
            unzip_file = local_file_path[:-4]
            print(unzip_file)

            # rozbali zip subor bez poslednych troch pismen
            zf = ZipFile(local_file_path, 'r')
            zf.extractall(local_file_path[:-4])
            zf.close()

            # vymaze zip subor
            print('vymazanie suboru')
            print(local_file_path)
            if os.path.isfile(local_file_path):
                os.remove(local_file_path)

    # Close the SFTP session
    sftp.close()

    # Close the SSH client object
    ssh.close()

    # create txt file
    folder = filepath = BASE_DIR + top_storage + storage
    create_files = search_file+'_'+name+'.txt'
    print (create_files)
    try:
        with open(folder + create_files, 'w') as f:
            f.write('---------------------------------------------  ALL '+name+'Logs from '+storage+' ----------\n')
            print("Create is_log.txt")
    except FileNotFoundError:
        print("The 'docs' directory does not exist")
    folder_emsaapp01_log = BASE_DIR + top_storage + storage +create_files
    search_folder = filepath = BASE_DIR + top_storage + storage + search_file +'/'

    # append log in text
    for filename in os.listdir(search_folder):
        # if filename.endswith('.log'):       # konciaci na .log
        # if filename.startswith('We', 7, 9):   # we je na 7 az 9 mieste
        if name in str(filename):
    #         shutil.copy( source_folder + filename, destination_folder)
    #         # zapis do suboru web log txt
            with open(( search_folder + filename), 'r', encoding="utf-8") as reader, open(folder_emsaapp01_log,'a', encoding="utf-8") as a_writer:
                dog_breeds = reader.readlines()
                a_writer.write('\n')
                a_writer.write('\n')
                a_writer.write('\n '+storage+'/  ')
                a_writer.writelines(filename)
                a_writer.write(' \n')
                a_writer.write('  \n')
                a_writer.writelines((dog_breeds))
                a_writer.close()
                # koniec zapisu
                # print('copied /EMSAAPP02_ErrorLogs/', filename )

    # ----download file ------------------
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    filename = create_files
    # Define the full file path
    filepath = BASE_DIR + top_storage + storage + filename
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    if isExist:
        shutil.rmtree(search_folder)
    # Return the response value
    return response

def EMSBAPP02_realtime_app(request):

    # Create an SSH client object
    ssh = paramiko.SSHClient()

    # Automatically add the server's host key (you can disable this if you want to manually check the host key)
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # Connect to the SFTP server
    ssh.connect(hostname, port, username, password)

    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # delete dir

    top_storage = '/ems_logs/'
    storage = 'EMSBAPP02/'
    search_file = 'EMSBAPP02_realtime'
    name = 'App'

    dir = filepath = BASE_DIR + top_storage + storage
    isExist = os.path.exists(dir)
    if isExist:
        shutil.rmtree(dir)
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    else:
        # create dir
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    # Open an SFTP session
    sftp = ssh.open_sftp()
    print('pristup')
    print(sftp)

    # Define the remote directory's path
    remote_directory_path = '/srv/archive/log/upload-by-smbfs/billien-ems/temp/' #'/srv/archive/log/upload-by-smbfs/billien-edz/realtime/EDZISAPP02/'
    print('ulozisko'+remote_directory_path)

    # Define the local directory's path
    local_directory_path = BASE_DIR + top_storage + storage
    # local_directory_path = os.listdir(r'\fileserver.internal.example.com\c$\Users\app_user\Downloads')
    print('localne ulozisko'+local_directory_path)

    # Get a list of all files in the remote directory
    files = sftp.listdir(remote_directory_path)
    print('sftp files')
    print(files)

    # Iterate over each file in the remote directory
    for file in files:
        # If the file's name contains "error", copy it from the SFTP server to the local machine
        # if 'EMSISAPP0' in file:
        if search_file in file:
            print('files')
            print(f'Copying {file}')
            # print(file)

            # Define the remote file's path
            remote_file_path = remote_directory_path+file
            print('remote_file_path')
            print(remote_file_path)
            # Define the local file's path
            local_file_path = local_directory_path+file
            print('local_file_path')
            print(local_file_path)
            # Copy the file from the SFTP server to the local machine
            # ulozi zip subor
            sftp.get(remote_file_path, local_file_path)

            # Ulozim si nazov priecinkov ktore su odzipovane
            unzip_file = local_file_path[:-4]
            print(unzip_file)

            # rozbali zip subor bez poslednych troch pismen
            zf = ZipFile(local_file_path, 'r')
            zf.extractall(local_file_path[:-4])
            zf.close()

            # vymaze zip subor
            print('vymazanie suboru')
            print(local_file_path)
            if os.path.isfile(local_file_path):
                os.remove(local_file_path)

    # Close the SFTP session
    sftp.close()

    # Close the SSH client object
    ssh.close()

    # create txt file
    folder = filepath = BASE_DIR + top_storage + storage
    create_files = search_file+'_'+name+'.txt'
    print (create_files)
    try:
        with open(folder + create_files, 'w') as f:
            f.write('---------------------------------------------  ALL '+name+'Logs from '+storage+' ----------\n')
            print("Create is_log.txt")
    except FileNotFoundError:
        print("The 'docs' directory does not exist")
    folder_emsaapp01_log = BASE_DIR + top_storage + storage +create_files
    search_folder = filepath = BASE_DIR + top_storage + storage + search_file +'/'

    # append log in text
    for filename in os.listdir(search_folder):
        # if filename.endswith('.log'):       # konciaci na .log
        # if filename.startswith('We', 7, 9):   # we je na 7 az 9 mieste
        if name in str(filename):
    #         shutil.copy( source_folder + filename, destination_folder)
    #         # zapis do suboru web log txt
            with open(( search_folder + filename), 'r', encoding="utf-8") as reader, open(folder_emsaapp01_log,'a', encoding="utf-8") as a_writer:
                dog_breeds = reader.readlines()
                a_writer.write('\n')
                a_writer.write('\n')
                a_writer.write('\n '+storage+'/  ')
                a_writer.writelines(filename)
                a_writer.write(' \n')
                a_writer.write('  \n')
                a_writer.writelines((dog_breeds))
                a_writer.close()
                # koniec zapisu
                # print('copied /EMSAAPP02_ErrorLogs/', filename )

    # ----download file ------------------
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    filename = create_files
    # Define the full file path
    filepath = BASE_DIR + top_storage + storage + filename
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    if isExist:
        shutil.rmtree(search_folder)
    # Return the response value
    return response

def EMSCAPP01_realtime_error(request):

    # Create an SSH client object
    ssh = paramiko.SSHClient()

    # Automatically add the server's host key (you can disable this if you want to manually check the host key)
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # Connect to the SFTP server
    ssh.connect(hostname, port, username, password)

    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # delete dir

    top_storage = '/ems_logs/'
    storage = 'EMSCAPP01/'
    search_file = 'EMSCAPP01_realtime'
    name = 'Error'

    dir = filepath = BASE_DIR + top_storage + storage
    isExist = os.path.exists(dir)
    if isExist:
        shutil.rmtree(dir)
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    else:
        # create dir
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    # Open an SFTP session
    sftp = ssh.open_sftp()
    print('pristup')
    print(sftp)

    # Define the remote directory's path
    remote_directory_path = '/srv/archive/log/upload-by-smbfs/billien-ems/temp/' #'/srv/archive/log/upload-by-smbfs/billien-edz/realtime/EDZISAPP02/'
    print('ulozisko'+remote_directory_path)

    # Define the local directory's path
    local_directory_path = BASE_DIR + top_storage + storage
    # local_directory_path = os.listdir(r'\fileserver.internal.example.com\c$\Users\app_user\Downloads')
    print('localne ulozisko'+local_directory_path)

    # Get a list of all files in the remote directory
    files = sftp.listdir(remote_directory_path)
    print('sftp files')
    print(files)

    # Iterate over each file in the remote directory
    for file in files:
        # If the file's name contains "error", copy it from the SFTP server to the local machine
        # if 'EMSISAPP0' in file:
        if search_file in file:
            print('files')
            print(f'Copying {file}')
            # print(file)

            # Define the remote file's path
            remote_file_path = remote_directory_path+file
            print('remote_file_path')
            print(remote_file_path)
            # Define the local file's path
            local_file_path = local_directory_path+file
            print('local_file_path')
            print(local_file_path)
            # Copy the file from the SFTP server to the local machine
            # ulozi zip subor
            sftp.get(remote_file_path, local_file_path)

            # Ulozim si nazov priecinkov ktore su odzipovane
            unzip_file = local_file_path[:-4]
            print(unzip_file)

            # rozbali zip subor bez poslednych troch pismen
            zf = ZipFile(local_file_path, 'r')
            zf.extractall(local_file_path[:-4])
            zf.close()

            # vymaze zip subor
            print('vymazanie suboru')
            print(local_file_path)
            if os.path.isfile(local_file_path):
                os.remove(local_file_path)

    # Close the SFTP session
    sftp.close()

    # Close the SSH client object
    ssh.close()

    # create txt file
    folder = filepath = BASE_DIR + top_storage + storage
    create_files = search_file+'_'+name+'.txt'
    print (create_files)
    try:
        with open(folder + create_files, 'w') as f:
            f.write('---------------------------------------------  ALL '+name+'Logs from '+storage+' ----------\n')
            print("Create is_log.txt")
    except FileNotFoundError:
        print("The 'docs' directory does not exist")
    folder_emsaapp01_log = BASE_DIR + top_storage + storage +create_files
    search_folder = filepath = BASE_DIR + top_storage + storage + search_file +'/'

    # append log in text
    for filename in os.listdir(search_folder):
        # if filename.endswith('.log'):       # konciaci na .log
        # if filename.startswith('We', 7, 9):   # we je na 7 az 9 mieste
        if name in str(filename):
    #         shutil.copy( source_folder + filename, destination_folder)
    #         # zapis do suboru web log txt
            with open(( search_folder + filename), 'r', encoding="utf-8") as reader, open(folder_emsaapp01_log,'a', encoding="utf-8") as a_writer:
                dog_breeds = reader.readlines()
                a_writer.write('\n')
                a_writer.write('\n')
                a_writer.write('\n '+storage+'/  ')
                a_writer.writelines(filename)
                a_writer.write(' \n')
                a_writer.write('  \n')
                a_writer.writelines((dog_breeds))
                a_writer.close()
                # koniec zapisu
                # print('copied /EMSAAPP02_ErrorLogs/', filename )

    # ----download file ------------------
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    filename = create_files
    # Define the full file path
    filepath = BASE_DIR + top_storage + storage + filename
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    if isExist:
        shutil.rmtree(search_folder)
    # Return the response value
    return response

def EMSCAPP01_realtime_app(request):

    # Create an SSH client object
    ssh = paramiko.SSHClient()

    # Automatically add the server's host key (you can disable this if you want to manually check the host key)
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # Connect to the SFTP server
    ssh.connect(hostname, port, username, password)

    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # delete dir

    top_storage = '/ems_logs/'
    storage = 'EMSCAPP01/'
    search_file = 'EMSCAPP01_realtime'
    name = 'App'

    dir = filepath = BASE_DIR + top_storage + storage
    isExist = os.path.exists(dir)
    if isExist:
        shutil.rmtree(dir)
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    else:
        # create dir
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    # Open an SFTP session
    sftp = ssh.open_sftp()
    print('pristup')
    print(sftp)

    # Define the remote directory's path
    remote_directory_path = '/srv/archive/log/upload-by-smbfs/billien-ems/temp/' #'/srv/archive/log/upload-by-smbfs/billien-edz/realtime/EDZISAPP02/'
    print('ulozisko'+remote_directory_path)

    # Define the local directory's path
    local_directory_path = BASE_DIR + top_storage + storage
    # local_directory_path = os.listdir(r'\fileserver.internal.example.com\c$\Users\app_user\Downloads')
    print('localne ulozisko'+local_directory_path)

    # Get a list of all files in the remote directory
    files = sftp.listdir(remote_directory_path)
    print('sftp files')
    print(files)

    # Iterate over each file in the remote directory
    for file in files:
        # If the file's name contains "error", copy it from the SFTP server to the local machine
        # if 'EMSISAPP0' in file:
        if search_file in file:
            print('files')
            print(f'Copying {file}')
            # print(file)

            # Define the remote file's path
            remote_file_path = remote_directory_path+file
            print('remote_file_path')
            print(remote_file_path)
            # Define the local file's path
            local_file_path = local_directory_path+file
            print('local_file_path')
            print(local_file_path)
            # Copy the file from the SFTP server to the local machine
            # ulozi zip subor
            sftp.get(remote_file_path, local_file_path)

            # Ulozim si nazov priecinkov ktore su odzipovane
            unzip_file = local_file_path[:-4]
            print(unzip_file)

            # rozbali zip subor bez poslednych troch pismen
            zf = ZipFile(local_file_path, 'r')
            zf.extractall(local_file_path[:-4])
            zf.close()

            # vymaze zip subor
            print('vymazanie suboru')
            print(local_file_path)
            if os.path.isfile(local_file_path):
                os.remove(local_file_path)

    # Close the SFTP session
    sftp.close()

    # Close the SSH client object
    ssh.close()

    # create txt file
    folder = filepath = BASE_DIR + top_storage + storage
    create_files = search_file+'_'+name+'.txt'
    print (create_files)
    try:
        with open(folder + create_files, 'w') as f:
            f.write('---------------------------------------------  ALL '+name+'Logs from '+storage+' ----------\n')
            print("Create is_log.txt")
    except FileNotFoundError:
        print("The 'docs' directory does not exist")
    folder_emsaapp01_log = BASE_DIR + top_storage + storage +create_files
    search_folder = filepath = BASE_DIR + top_storage + storage + search_file +'/'

    # append log in text
    for filename in os.listdir(search_folder):
        # if filename.endswith('.log'):       # konciaci na .log
        # if filename.startswith('We', 7, 9):   # we je na 7 az 9 mieste
        if name in str(filename):
    #         shutil.copy( source_folder + filename, destination_folder)
    #         # zapis do suboru web log txt
            with open(( search_folder + filename), 'r', encoding="utf-8") as reader, open(folder_emsaapp01_log,'a', encoding="utf-8") as a_writer:
                dog_breeds = reader.readlines()
                a_writer.write('\n')
                a_writer.write('\n')
                a_writer.write('\n'+storage+'/  ')
                a_writer.writelines(filename)
                a_writer.write(' \n')
                a_writer.write('  \n')
                a_writer.writelines((dog_breeds))
                a_writer.close()
                # koniec zapisu
                # print('copied /EMSAAPP02_ErrorLogs/', filename )

    # ----download file ------------------
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    filename = create_files
    # Define the full file path
    filepath = BASE_DIR + top_storage + storage + filename
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    if isExist:
        shutil.rmtree(search_folder)
    # Return the response value
    return response

def EMSCAPP02_realtime_error(request):

    # Create an SSH client object
    ssh = paramiko.SSHClient()

    # Automatically add the server's host key (you can disable this if you want to manually check the host key)
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # Connect to the SFTP server
    ssh.connect(hostname, port, username, password)

    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # delete dir

    top_storage = '/ems_logs/'
    storage = 'EMSCAPP02/'
    search_file = 'EMSCAPP02_realtime'
    name = 'Error'

    dir = filepath = BASE_DIR + top_storage + storage
    isExist = os.path.exists(dir)
    if isExist:
        shutil.rmtree(dir)
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    else:
        # create dir
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    # Open an SFTP session
    sftp = ssh.open_sftp()
    print('pristup')
    print(sftp)

    # Define the remote directory's path
    remote_directory_path = '/srv/archive/log/upload-by-smbfs/billien-ems/temp/' #'/srv/archive/log/upload-by-smbfs/billien-edz/realtime/EDZISAPP02/'
    print('ulozisko'+remote_directory_path)

    # Define the local directory's path
    local_directory_path = BASE_DIR + top_storage + storage
    # local_directory_path = os.listdir(r'\fileserver.internal.example.com\c$\Users\app_user\Downloads')
    print('localne ulozisko'+local_directory_path)

    # Get a list of all files in the remote directory
    files = sftp.listdir(remote_directory_path)
    print('sftp files')
    print(files)

    # Iterate over each file in the remote directory
    for file in files:
        # If the file's name contains "error", copy it from the SFTP server to the local machine
        # if 'EMSISAPP0' in file:
        if search_file in file:
            print('files')
            print(f'Copying {file}')
            # print(file)

            # Define the remote file's path
            remote_file_path = remote_directory_path+file
            print('remote_file_path')
            print(remote_file_path)
            # Define the local file's path
            local_file_path = local_directory_path+file
            print('local_file_path')
            print(local_file_path)
            # Copy the file from the SFTP server to the local machine
            # ulozi zip subor
            sftp.get(remote_file_path, local_file_path)

            # Ulozim si nazov priecinkov ktore su odzipovane
            unzip_file = local_file_path[:-4]
            print(unzip_file)

            # rozbali zip subor bez poslednych troch pismen
            zf = ZipFile(local_file_path, 'r')
            zf.extractall(local_file_path[:-4])
            zf.close()

            # vymaze zip subor
            print('vymazanie suboru')
            print(local_file_path)
            if os.path.isfile(local_file_path):
                os.remove(local_file_path)

    # Close the SFTP session
    sftp.close()

    # Close the SSH client object
    ssh.close()

    # create txt file
    folder = filepath = BASE_DIR + top_storage + storage
    create_files = search_file+'_'+name+'.txt'
    print (create_files)
    try:
        with open(folder + create_files, 'w') as f:
            f.write('---------------------------------------------  ALL '+name+'Logs from '+storage+' ----------\n')
            print("Create is_log.txt")
    except FileNotFoundError:
        print("The 'docs' directory does not exist")
    folder_emsaapp01_log = BASE_DIR + top_storage + storage +create_files
    search_folder = filepath = BASE_DIR + top_storage + storage + search_file +'/'

    # append log in text
    for filename in os.listdir(search_folder):
        # if filename.endswith('.log'):       # konciaci na .log
        # if filename.startswith('We', 7, 9):   # we je na 7 az 9 mieste
        if name in str(filename):
    #         shutil.copy( source_folder + filename, destination_folder)
    #         # zapis do suboru web log txt
            with open(( search_folder + filename), 'r', encoding="utf-8") as reader, open(folder_emsaapp01_log,'a', encoding="utf-8") as a_writer:
                dog_breeds = reader.readlines()
                a_writer.write('\n')
                a_writer.write('\n')
                a_writer.write('\n '+storage+'/  ')
                a_writer.writelines(filename)
                a_writer.write(' \n')
                a_writer.write('  \n')
                a_writer.writelines((dog_breeds))
                a_writer.close()
                # koniec zapisu
                # print('copied /EMSAAPP02_ErrorLogs/', filename )

    # ----download file ------------------
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    filename = create_files
    # Define the full file path
    filepath = BASE_DIR + top_storage + storage + filename
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    if isExist:
        shutil.rmtree(search_folder)
    # Return the response value
    return response

def EMSCAPP02_realtime_app(request):

    # Create an SSH client object
    ssh = paramiko.SSHClient()

    # Automatically add the server's host key (you can disable this if you want to manually check the host key)
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # Connect to the SFTP server
    ssh.connect(hostname, port, username, password)

    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # delete dir

    top_storage = '/ems_logs/'
    storage = 'EMSCAPP02/'
    search_file = 'EMSCAPP02_realtime'
    name = 'App'

    dir = filepath = BASE_DIR + top_storage + storage
    isExist = os.path.exists(dir)
    if isExist:
        shutil.rmtree(dir)
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    else:
        # create dir
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    # Open an SFTP session
    sftp = ssh.open_sftp()
    print('pristup')
    print(sftp)

    # Define the remote directory's path
    remote_directory_path = '/srv/archive/log/upload-by-smbfs/billien-ems/temp/' #'/srv/archive/log/upload-by-smbfs/billien-edz/realtime/EDZISAPP02/'
    print('ulozisko'+remote_directory_path)

    # Define the local directory's path
    local_directory_path = BASE_DIR + top_storage + storage
    # local_directory_path = os.listdir(r'\fileserver.internal.example.com\c$\Users\app_user\Downloads')
    print('localne ulozisko'+local_directory_path)

    # Get a list of all files in the remote directory
    files = sftp.listdir(remote_directory_path)
    print('sftp files')
    print(files)

    # Iterate over each file in the remote directory
    for file in files:
        # If the file's name contains "error", copy it from the SFTP server to the local machine
        # if 'EMSISAPP0' in file:
        if search_file in file:
            print('files')
            print(f'Copying {file}')
            # print(file)

            # Define the remote file's path
            remote_file_path = remote_directory_path+file
            print('remote_file_path')
            print(remote_file_path)
            # Define the local file's path
            local_file_path = local_directory_path+file
            print('local_file_path')
            print(local_file_path)
            # Copy the file from the SFTP server to the local machine
            # ulozi zip subor
            sftp.get(remote_file_path, local_file_path)

            # Ulozim si nazov priecinkov ktore su odzipovane
            unzip_file = local_file_path[:-4]
            print(unzip_file)

            # rozbali zip subor bez poslednych troch pismen
            zf = ZipFile(local_file_path, 'r')
            zf.extractall(local_file_path[:-4])
            zf.close()

            # vymaze zip subor
            print('vymazanie suboru')
            print(local_file_path)
            if os.path.isfile(local_file_path):
                os.remove(local_file_path)

    # Close the SFTP session
    sftp.close()

    # Close the SSH client object
    ssh.close()

    # create txt file
    folder = filepath = BASE_DIR + top_storage + storage
    create_files = search_file+'_'+name+'.txt'
    print (create_files)
    try:
        with open(folder + create_files, 'w') as f:
            f.write('---------------------------------------------  ALL '+name+'Logs from '+storage+' ----------\n')
            print("Create is_log.txt")
    except FileNotFoundError:
        print("The 'docs' directory does not exist")
    folder_emsaapp01_log = BASE_DIR + top_storage + storage +create_files
    search_folder = filepath = BASE_DIR + top_storage + storage + search_file +'/'

    # append log in text
    for filename in os.listdir(search_folder):
        # if filename.endswith('.log'):       # konciaci na .log
        # if filename.startswith('We', 7, 9):   # we je na 7 az 9 mieste
        if name in str(filename):
    #         shutil.copy( source_folder + filename, destination_folder)
    #         # zapis do suboru web log txt
            with open(( search_folder + filename), 'r', encoding="utf-8") as reader, open(folder_emsaapp01_log,'a', encoding="utf-8") as a_writer:
                dog_breeds = reader.readlines()
                a_writer.write('\n')
                a_writer.write('\n')
                a_writer.write('\n'+storage+'/  ')
                a_writer.writelines(filename)
                a_writer.write(' \n')
                a_writer.write('  \n')
                a_writer.writelines((dog_breeds))
                a_writer.close()
                # koniec zapisu
                # print('copied /EMSAAPP02_ErrorLogs/', filename )

    # ----download file ------------------
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    filename = create_files
    # Define the full file path
    filepath = BASE_DIR + top_storage + storage + filename
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    if isExist:
        shutil.rmtree(search_folder)
    # Return the response value
    return response

def EMSDAPP01_realtime_error(request):

    # Create an SSH client object
    ssh = paramiko.SSHClient()

    # Automatically add the server's host key (you can disable this if you want to manually check the host key)
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # Connect to the SFTP server
    ssh.connect(hostname, port, username, password)

    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # delete dir

    top_storage = '/ems_logs/'
    storage = 'EMSDAPP01/'
    search_file = 'EMSDAPP01_realtime'
    name = 'Error'

    dir = filepath = BASE_DIR + top_storage + storage
    isExist = os.path.exists(dir)
    if isExist:
        shutil.rmtree(dir)
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    else:
        # create dir
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    # Open an SFTP session
    sftp = ssh.open_sftp()
    print('pristup')
    print(sftp)

    # Define the remote directory's path
    remote_directory_path = '/srv/archive/log/upload-by-smbfs/billien-ems/temp/' #'/srv/archive/log/upload-by-smbfs/billien-edz/realtime/EDZISAPP02/'
    print('ulozisko'+remote_directory_path)

    # Define the local directory's path
    local_directory_path = BASE_DIR + top_storage + storage
    # local_directory_path = os.listdir(r'\fileserver.internal.example.com\c$\Users\app_user\Downloads')
    print('localne ulozisko'+local_directory_path)

    # Get a list of all files in the remote directory
    files = sftp.listdir(remote_directory_path)
    print('sftp files')
    print(files)

    # Iterate over each file in the remote directory
    for file in files:
        # If the file's name contains "error", copy it from the SFTP server to the local machine
        # if 'EMSISAPP0' in file:
        if search_file in file:
            print('files')
            print(f'Copying {file}')
            # print(file)

            # Define the remote file's path
            remote_file_path = remote_directory_path+file
            print('remote_file_path')
            print(remote_file_path)
            # Define the local file's path
            local_file_path = local_directory_path+file
            print('local_file_path')
            print(local_file_path)
            # Copy the file from the SFTP server to the local machine
            # ulozi zip subor
            sftp.get(remote_file_path, local_file_path)

            # Ulozim si nazov priecinkov ktore su odzipovane
            unzip_file = local_file_path[:-4]
            print(unzip_file)

            # rozbali zip subor bez poslednych troch pismen
            zf = ZipFile(local_file_path, 'r')
            zf.extractall(local_file_path[:-4])
            zf.close()

            # vymaze zip subor
            print('vymazanie suboru')
            print(local_file_path)
            if os.path.isfile(local_file_path):
                os.remove(local_file_path)

    # Close the SFTP session
    sftp.close()

    # Close the SSH client object
    ssh.close()

    # create txt file
    folder = filepath = BASE_DIR + top_storage + storage
    create_files = search_file+'_'+name+'.txt'
    print (create_files)
    try:
        with open(folder + create_files, 'w') as f:
            f.write('---------------------------------------------  ALL '+name+'Logs from '+storage+' ----------\n')
            print("Create is_log.txt")
    except FileNotFoundError:
        print("The 'docs' directory does not exist")
    folder_emsaapp01_log = BASE_DIR + top_storage + storage +create_files
    search_folder = filepath = BASE_DIR + top_storage + storage + search_file +'/'

    # append log in text
    for filename in os.listdir(search_folder):
        # if filename.endswith('.log'):       # konciaci na .log
        # if filename.startswith('We', 7, 9):   # we je na 7 az 9 mieste
        if name in str(filename):
    #         shutil.copy( source_folder + filename, destination_folder)
    #         # zapis do suboru web log txt
            with open(( search_folder + filename), 'r', encoding="utf-8") as reader, open(folder_emsaapp01_log,'a', encoding="utf-8") as a_writer:
                dog_breeds = reader.readlines()
                a_writer.write('\n')
                a_writer.write('\n')
                a_writer.write('\n '+storage+'/  ')
                a_writer.writelines(filename)
                a_writer.write(' \n')
                a_writer.write('  \n')
                a_writer.writelines((dog_breeds))
                a_writer.close()
                # koniec zapisu
                # print('copied /EMSAAPP02_ErrorLogs/', filename )

    # ----download file ------------------
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    filename = create_files
    # Define the full file path
    filepath = BASE_DIR + top_storage + storage + filename
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    if isExist:
        shutil.rmtree(search_folder)
    # Return the response value
    return response

def EMSDAPP01_realtime_app(request):

    # Create an SSH client object
    ssh = paramiko.SSHClient()

    # Automatically add the server's host key (you can disable this if you want to manually check the host key)
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # Connect to the SFTP server
    ssh.connect(hostname, port, username, password)

    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # delete dir

    top_storage = '/ems_logs/'
    storage = 'EMSDAPP01/'
    search_file = 'EMSDAPP01_realtime'
    name = 'App'

    dir = filepath = BASE_DIR + top_storage + storage
    isExist = os.path.exists(dir)
    if isExist:
        shutil.rmtree(dir)
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    else:
        # create dir
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    # Open an SFTP session
    sftp = ssh.open_sftp()
    print('pristup')
    print(sftp)

    # Define the remote directory's path
    remote_directory_path = '/srv/archive/log/upload-by-smbfs/billien-ems/temp/' #'/srv/archive/log/upload-by-smbfs/billien-edz/realtime/EDZISAPP02/'
    print('ulozisko'+remote_directory_path)

    # Define the local directory's path
    local_directory_path = BASE_DIR + top_storage + storage
    # local_directory_path = os.listdir(r'\fileserver.internal.example.com\c$\Users\app_user\Downloads')
    print('localne ulozisko'+local_directory_path)

    # Get a list of all files in the remote directory
    files = sftp.listdir(remote_directory_path)
    print('sftp files')
    print(files)

    # Iterate over each file in the remote directory
    for file in files:
        # If the file's name contains "error", copy it from the SFTP server to the local machine
        # if 'EMSISAPP0' in file:
        if search_file in file:
            print('files')
            print(f'Copying {file}')
            # print(file)

            # Define the remote file's path
            remote_file_path = remote_directory_path+file
            print('remote_file_path')
            print(remote_file_path)
            # Define the local file's path
            local_file_path = local_directory_path+file
            print('local_file_path')
            print(local_file_path)
            # Copy the file from the SFTP server to the local machine
            # ulozi zip subor
            sftp.get(remote_file_path, local_file_path)

            # Ulozim si nazov priecinkov ktore su odzipovane
            unzip_file = local_file_path[:-4]
            print(unzip_file)

            # rozbali zip subor bez poslednych troch pismen
            zf = ZipFile(local_file_path, 'r')
            zf.extractall(local_file_path[:-4])
            zf.close()

            # vymaze zip subor
            print('vymazanie suboru')
            print(local_file_path)
            if os.path.isfile(local_file_path):
                os.remove(local_file_path)

    # Close the SFTP session
    sftp.close()

    # Close the SSH client object
    ssh.close()

    # create txt file
    folder = filepath = BASE_DIR + top_storage + storage
    create_files = search_file+'_'+name+'.txt'
    print (create_files)
    try:
        with open(folder + create_files, 'w') as f:
            f.write('---------------------------------------------  ALL '+name+'Logs from '+storage+' ----------\n')
            print("Create is_log.txt")
    except FileNotFoundError:
        print("The 'docs' directory does not exist")
    folder_emsaapp01_log = BASE_DIR + top_storage + storage +create_files
    search_folder = filepath = BASE_DIR + top_storage + storage + search_file +'/'

    # append log in text
    for filename in os.listdir(search_folder):
        # if filename.endswith('.log'):       # konciaci na .log
        # if filename.startswith('We', 7, 9):   # we je na 7 az 9 mieste
        if name in str(filename):
    #         shutil.copy( source_folder + filename, destination_folder)
    #         # zapis do suboru web log txt
            with open(( search_folder + filename), 'r', encoding="utf-8") as reader, open(folder_emsaapp01_log,'a', encoding="utf-8") as a_writer:
                dog_breeds = reader.readlines()
                a_writer.write('\n')
                a_writer.write('\n')
                a_writer.write('\n'+storage+'/  ')
                a_writer.writelines(filename)
                a_writer.write(' \n')
                a_writer.write('  \n')
                a_writer.writelines((dog_breeds))
                a_writer.close()
                # koniec zapisu
                # print('copied /EMSAAPP02_ErrorLogs/', filename )

    # ----download file ------------------
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    filename = create_files
    # Define the full file path
    filepath = BASE_DIR + top_storage + storage + filename
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    if isExist:
        shutil.rmtree(search_folder)
    # Return the response value
    return response

def EMSDAPP02_realtime_error(request):

    # Create an SSH client object
    ssh = paramiko.SSHClient()

    # Automatically add the server's host key (you can disable this if you want to manually check the host key)
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # Connect to the SFTP server
    ssh.connect(hostname, port, username, password)

    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # delete dir

    top_storage = '/ems_logs/'
    storage = 'EMSDAPP02/'
    search_file = 'EMSDAPP02_realtime'
    name = 'Error'

    dir = filepath = BASE_DIR + top_storage + storage
    isExist = os.path.exists(dir)
    if isExist:
        shutil.rmtree(dir)
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    else:
        # create dir
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    # Open an SFTP session
    sftp = ssh.open_sftp()
    print('pristup')
    print(sftp)

    # Define the remote directory's path
    remote_directory_path = '/srv/archive/log/upload-by-smbfs/billien-ems/temp/' #'/srv/archive/log/upload-by-smbfs/billien-edz/realtime/EDZISAPP02/'
    print('ulozisko'+remote_directory_path)

    # Define the local directory's path
    local_directory_path = BASE_DIR + top_storage + storage
    # local_directory_path = os.listdir(r'\fileserver.internal.example.com\c$\Users\app_user\Downloads')
    print('localne ulozisko'+local_directory_path)

    # Get a list of all files in the remote directory
    files = sftp.listdir(remote_directory_path)
    print('sftp files')
    print(files)

    # Iterate over each file in the remote directory
    for file in files:
        # If the file's name contains "error", copy it from the SFTP server to the local machine
        # if 'EMSISAPP0' in file:
        if search_file in file:
            print('files')
            print(f'Copying {file}')
            # print(file)

            # Define the remote file's path
            remote_file_path = remote_directory_path+file
            print('remote_file_path')
            print(remote_file_path)
            # Define the local file's path
            local_file_path = local_directory_path+file
            print('local_file_path')
            print(local_file_path)
            # Copy the file from the SFTP server to the local machine
            # ulozi zip subor
            sftp.get(remote_file_path, local_file_path)

            # Ulozim si nazov priecinkov ktore su odzipovane
            unzip_file = local_file_path[:-4]
            print(unzip_file)

            # rozbali zip subor bez poslednych troch pismen
            zf = ZipFile(local_file_path, 'r')
            zf.extractall(local_file_path[:-4])
            zf.close()

            # vymaze zip subor
            print('vymazanie suboru')
            print(local_file_path)
            if os.path.isfile(local_file_path):
                os.remove(local_file_path)

    # Close the SFTP session
    sftp.close()

    # Close the SSH client object
    ssh.close()

    # create txt file
    folder = filepath = BASE_DIR + top_storage + storage
    create_files = search_file+'_'+name+'.txt'
    print (create_files)
    try:
        with open(folder + create_files, 'w') as f:
            f.write('---------------------------------------------  ALL '+name+'Logs from '+storage+' ----------\n')
            print("Create is_log.txt")
    except FileNotFoundError:
        print("The 'docs' directory does not exist")
    folder_emsaapp01_log = BASE_DIR + top_storage + storage +create_files
    search_folder = filepath = BASE_DIR + top_storage + storage + search_file +'/'

    # append log in text
    for filename in os.listdir(search_folder):
        # if filename.endswith('.log'):       # konciaci na .log
        # if filename.startswith('We', 7, 9):   # we je na 7 az 9 mieste
        if name in str(filename):
    #         shutil.copy( source_folder + filename, destination_folder)
    #         # zapis do suboru web log txt
            with open(( search_folder + filename), 'r', encoding="utf-8") as reader, open(folder_emsaapp01_log,'a', encoding="utf-8") as a_writer:
                dog_breeds = reader.readlines()
                a_writer.write('\n')
                a_writer.write('\n')
                a_writer.write('\n'+storage+'/  ')
                a_writer.writelines(filename)
                a_writer.write(' \n')
                a_writer.write('  \n')
                a_writer.writelines((dog_breeds))
                a_writer.close()
                # koniec zapisu
                # print('copied /EMSAAPP02_ErrorLogs/', filename )

    # ----download file ------------------
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    filename = create_files
    # Define the full file path
    filepath = BASE_DIR + top_storage + storage + filename
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    if isExist:
        shutil.rmtree(search_folder)
    # Return the response value
    return response

def EMSDAPP02_realtime_app(request):

    # Create an SSH client object
    ssh = paramiko.SSHClient()

    # Automatically add the server's host key (you can disable this if you want to manually check the host key)
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # Connect to the SFTP server
    ssh.connect(hostname, port, username, password)

    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # delete dir

    top_storage = '/ems_logs/'
    storage = 'EMSDAPP02/'
    search_file = 'EMSDAPP02_realtime'
    name = 'App'

    dir = filepath = BASE_DIR + top_storage + storage
    isExist = os.path.exists(dir)
    if isExist:
        shutil.rmtree(dir)
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    else:
        # create dir
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    # Open an SFTP session
    sftp = ssh.open_sftp()
    print('pristup')
    print(sftp)

    # Define the remote directory's path
    remote_directory_path = '/srv/archive/log/upload-by-smbfs/billien-ems/temp/' #'/srv/archive/log/upload-by-smbfs/billien-edz/realtime/EDZISAPP02/'
    print('ulozisko'+remote_directory_path)

    # Define the local directory's path
    local_directory_path = BASE_DIR + top_storage + storage
    # local_directory_path = os.listdir(r'\fileserver.internal.example.com\c$\Users\app_user\Downloads')
    print('localne ulozisko'+local_directory_path)

    # Get a list of all files in the remote directory
    files = sftp.listdir(remote_directory_path)
    print('sftp files')
    print(files)

    # Iterate over each file in the remote directory
    for file in files:
        # If the file's name contains "error", copy it from the SFTP server to the local machine
        # if 'EMSISAPP0' in file:
        if search_file in file:
            print('files')
            print(f'Copying {file}')
            # print(file)

            # Define the remote file's path
            remote_file_path = remote_directory_path+file
            print('remote_file_path')
            print(remote_file_path)
            # Define the local file's path
            local_file_path = local_directory_path+file
            print('local_file_path')
            print(local_file_path)
            # Copy the file from the SFTP server to the local machine
            # ulozi zip subor
            sftp.get(remote_file_path, local_file_path)

            # Ulozim si nazov priecinkov ktore su odzipovane
            unzip_file = local_file_path[:-4]
            print(unzip_file)

            # rozbali zip subor bez poslednych troch pismen
            zf = ZipFile(local_file_path, 'r')
            zf.extractall(local_file_path[:-4])
            zf.close()

            # vymaze zip subor
            print('vymazanie suboru')
            print(local_file_path)
            if os.path.isfile(local_file_path):
                os.remove(local_file_path)

    # Close the SFTP session
    sftp.close()

    # Close the SSH client object
    ssh.close()

    # create txt file
    folder = filepath = BASE_DIR + top_storage + storage
    create_files = search_file+'_'+name+'.txt'
    print (create_files)
    try:
        with open(folder + create_files, 'w') as f:
            f.write('---------------------------------------------  ALL '+name+'Logs from '+storage+' ----------\n')
            print("Create is_log.txt")
    except FileNotFoundError:
        print("The 'docs' directory does not exist")
    folder_emsaapp01_log = BASE_DIR + top_storage + storage +create_files
    search_folder = filepath = BASE_DIR + top_storage + storage + search_file +'/'

    # append log in text
    for filename in os.listdir(search_folder):
        # if filename.endswith('.log'):       # konciaci na .log
        # if filename.startswith('We', 7, 9):   # we je na 7 az 9 mieste
        if name in str(filename):
    #         shutil.copy( source_folder + filename, destination_folder)
    #         # zapis do suboru web log txt
            with open(( search_folder + filename), 'r', encoding="utf-8") as reader, open(folder_emsaapp01_log,'a', encoding="utf-8") as a_writer:
                dog_breeds = reader.readlines()
                a_writer.write('\n')
                a_writer.write('\n')
                a_writer.write('\n'+storage+'/  ')
                a_writer.writelines(filename)
                a_writer.write(' \n')
                a_writer.write('  \n')
                a_writer.writelines((dog_breeds))
                a_writer.close()
                # koniec zapisu
                # print('copied /EMSAAPP02_ErrorLogs/', filename )

    # ----download file ------------------
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    filename = create_files
    # Define the full file path
    filepath = BASE_DIR + top_storage + storage + filename
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    if isExist:
        shutil.rmtree(search_folder)
    # Return the response value
    return response

def EMSISAPP01_realtime_error(request):

    # Create an SSH client object
    ssh = paramiko.SSHClient()

    # Automatically add the server's host key (you can disable this if you want to manually check the host key)
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # Connect to the SFTP server
    ssh.connect(hostname, port, username, password)

    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # delete dir

    top_storage = '/ems_logs/'
    storage = 'EMSISAPP01/'
    search_file = 'EMSISAPP01_realtime'
    name = 'Error'

    dir = filepath = BASE_DIR + top_storage + storage
    isExist = os.path.exists(dir)
    if isExist:
        shutil.rmtree(dir)
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    else:
        # create dir
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    # Open an SFTP session
    sftp = ssh.open_sftp()
    print('pristup')
    print(sftp)

    # Define the remote directory's path
    remote_directory_path = '/srv/archive/log/upload-by-smbfs/billien-ems/temp/' #'/srv/archive/log/upload-by-smbfs/billien-edz/realtime/EDZISAPP02/'
    print('ulozisko'+remote_directory_path)

    # Define the local directory's path
    local_directory_path = BASE_DIR + top_storage + storage
    # local_directory_path = os.listdir(r'\fileserver.internal.example.com\c$\Users\app_user\Downloads')
    print('localne ulozisko'+local_directory_path)

    # Get a list of all files in the remote directory
    files = sftp.listdir(remote_directory_path)
    print('sftp files')
    print(files)

    # Iterate over each file in the remote directory
    for file in files:
        # If the file's name contains "error", copy it from the SFTP server to the local machine
        # if 'EMSISAPP0' in file:
        if search_file in file:
            print('files')
            print(f'Copying {file}')
            # print(file)

            # Define the remote file's path
            remote_file_path = remote_directory_path+file
            print('remote_file_path')
            print(remote_file_path)
            # Define the local file's path
            local_file_path = local_directory_path+file
            print('local_file_path')
            print(local_file_path)
            # Copy the file from the SFTP server to the local machine
            # ulozi zip subor
            sftp.get(remote_file_path, local_file_path)

            # Ulozim si nazov priecinkov ktore su odzipovane
            unzip_file = local_file_path[:-4]
            print(unzip_file)

            # rozbali zip subor bez poslednych troch pismen
            zf = ZipFile(local_file_path, 'r')
            zf.extractall(local_file_path[:-4])
            zf.close()

            # vymaze zip subor
            print('vymazanie suboru')
            print(local_file_path)
            if os.path.isfile(local_file_path):
                os.remove(local_file_path)

    # Close the SFTP session
    sftp.close()

    # Close the SSH client object
    ssh.close()

    # create txt file
    folder = filepath = BASE_DIR + top_storage + storage
    create_files = search_file+'_'+name+'.txt'
    print (create_files)
    try:
        with open(folder + create_files, 'w') as f:
            f.write('---------------------------------------------  ALL '+name+'Logs from '+storage+' ----------\n')
            print("Create is_log.txt")
    except FileNotFoundError:
        print("The 'docs' directory does not exist")
    folder_emsaapp01_log = BASE_DIR + top_storage + storage +create_files
    search_folder = filepath = BASE_DIR + top_storage + storage + search_file +'/'

    # append log in text
    for filename in os.listdir(search_folder):
        # if filename.endswith('.log'):       # konciaci na .log
        # if filename.startswith('We', 7, 9):   # we je na 7 az 9 mieste
        if name in str(filename):
    #         shutil.copy( source_folder + filename, destination_folder)
    #         # zapis do suboru web log txt
            with open(( search_folder + filename), 'r', encoding="utf-8") as reader, open(folder_emsaapp01_log,'a', encoding="utf-8") as a_writer:
                dog_breeds = reader.readlines()
                a_writer.write('\n')
                a_writer.write('\n')
                a_writer.write('\n'+storage+'/  ')
                a_writer.writelines(filename)
                a_writer.write(' \n')
                a_writer.write('  \n')
                a_writer.writelines((dog_breeds))
                a_writer.close()
                # koniec zapisu
                # print('copied /EMSAAPP02_ErrorLogs/', filename )

    # ----download file ------------------
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    filename = create_files
    # Define the full file path
    filepath = BASE_DIR + top_storage + storage + filename
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    if isExist:
        shutil.rmtree(search_folder)
    # Return the response value
    return response

def EMSISAPP01_realtime_app(request):

    # Create an SSH client object
    ssh = paramiko.SSHClient()

    # Automatically add the server's host key (you can disable this if you want to manually check the host key)
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # Connect to the SFTP server
    ssh.connect(hostname, port, username, password)

    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # delete dir

    top_storage = '/ems_logs/'
    storage = 'EMSISAPP01/'
    search_file = 'EMSISAPP01_realtime'
    name = 'App'

    dir = filepath = BASE_DIR + top_storage + storage
    isExist = os.path.exists(dir)
    if isExist:
        shutil.rmtree(dir)
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    else:
        # create dir
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    # Open an SFTP session
    sftp = ssh.open_sftp()
    print('pristup')
    print(sftp)

    # Define the remote directory's path
    remote_directory_path = '/srv/archive/log/upload-by-smbfs/billien-ems/temp/' #'/srv/archive/log/upload-by-smbfs/billien-edz/realtime/EDZISAPP02/'
    print('ulozisko'+remote_directory_path)

    # Define the local directory's path
    local_directory_path = BASE_DIR + top_storage + storage
    # local_directory_path = os.listdir(r'\fileserver.internal.example.com\c$\Users\app_user\Downloads')
    print('localne ulozisko'+local_directory_path)

    # Get a list of all files in the remote directory
    files = sftp.listdir(remote_directory_path)
    print('sftp files')
    print(files)

    # Iterate over each file in the remote directory
    for file in files:
        # If the file's name contains "error", copy it from the SFTP server to the local machine
        # if 'EMSISAPP0' in file:
        if search_file in file:
            print('files')
            print(f'Copying {file}')
            # print(file)

            # Define the remote file's path
            remote_file_path = remote_directory_path+file
            print('remote_file_path')
            print(remote_file_path)
            # Define the local file's path
            local_file_path = local_directory_path+file
            print('local_file_path')
            print(local_file_path)
            # Copy the file from the SFTP server to the local machine
            # ulozi zip subor
            sftp.get(remote_file_path, local_file_path)

            # Ulozim si nazov priecinkov ktore su odzipovane
            unzip_file = local_file_path[:-4]
            print(unzip_file)

            # rozbali zip subor bez poslednych troch pismen
            zf = ZipFile(local_file_path, 'r')
            zf.extractall(local_file_path[:-4])
            zf.close()

            # vymaze zip subor
            print('vymazanie suboru')
            print(local_file_path)
            if os.path.isfile(local_file_path):
                os.remove(local_file_path)

    # Close the SFTP session
    sftp.close()

    # Close the SSH client object
    ssh.close()

    # create txt file
    folder = filepath = BASE_DIR + top_storage + storage
    create_files = search_file+'_'+name+'.txt'
    print (create_files)
    try:
        with open(folder + create_files, 'w') as f:
            f.write('---------------------------------------------  ALL '+name+'Logs from '+storage+' ----------\n')
            print("Create is_log.txt")
    except FileNotFoundError:
        print("The 'docs' directory does not exist")
    folder_emsaapp01_log = BASE_DIR + top_storage + storage +create_files
    search_folder = filepath = BASE_DIR + top_storage + storage + search_file +'/'

    # append log in text
    for filename in os.listdir(search_folder):
        # if filename.endswith('.log'):       # konciaci na .log
        # if filename.startswith('We', 7, 9):   # we je na 7 az 9 mieste
        if name in str(filename):
    #         shutil.copy( source_folder + filename, destination_folder)
    #         # zapis do suboru web log txt
            with open(( search_folder + filename), 'r', encoding="utf-8") as reader, open(folder_emsaapp01_log,'a', encoding="utf-8") as a_writer:
                dog_breeds = reader.readlines()
                a_writer.write('\n')
                a_writer.write('\n')
                a_writer.write('\n'+storage+'/  ')
                a_writer.writelines(filename)
                a_writer.write(' \n')
                a_writer.write('  \n')
                a_writer.writelines((dog_breeds))
                a_writer.close()
                # koniec zapisu
                # print('copied /EMSAAPP02_ErrorLogs/', filename )

    # ----download file ------------------
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    filename = create_files
    # Define the full file path
    filepath = BASE_DIR + top_storage + storage + filename
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    if isExist:
        shutil.rmtree(search_folder)
    # Return the response value
    return response

def EMSISAPP02_realtime_error(request):

    # Create an SSH client object
    ssh = paramiko.SSHClient()

    # Automatically add the server's host key (you can disable this if you want to manually check the host key)
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # Connect to the SFTP server
    ssh.connect(hostname, port, username, password)

    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # delete dir

    top_storage = '/ems_logs/'
    storage = 'EMSISAPP02/'
    search_file = 'EMSISAPP02_realtime'
    name = 'Error'

    dir = filepath = BASE_DIR + top_storage + storage
    isExist = os.path.exists(dir)
    if isExist:
        shutil.rmtree(dir)
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    else:
        # create dir
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    # Open an SFTP session
    sftp = ssh.open_sftp()
    print('pristup')
    print(sftp)

    # Define the remote directory's path
    remote_directory_path = '/srv/archive/log/upload-by-smbfs/billien-ems/temp/' #'/srv/archive/log/upload-by-smbfs/billien-edz/realtime/EDZISAPP02/'
    print('ulozisko'+remote_directory_path)

    # Define the local directory's path
    local_directory_path = BASE_DIR + top_storage + storage
    # local_directory_path = os.listdir(r'\fileserver.internal.example.com\c$\Users\app_user\Downloads')
    print('localne ulozisko'+local_directory_path)

    # Get a list of all files in the remote directory
    files = sftp.listdir(remote_directory_path)
    print('sftp files')
    print(files)

    # Iterate over each file in the remote directory
    for file in files:
        # If the file's name contains "error", copy it from the SFTP server to the local machine
        # if 'EMSISAPP0' in file:
        if search_file in file:
            print('files')
            print(f'Copying {file}')
            # print(file)

            # Define the remote file's path
            remote_file_path = remote_directory_path+file
            print('remote_file_path')
            print(remote_file_path)
            # Define the local file's path
            local_file_path = local_directory_path+file
            print('local_file_path')
            print(local_file_path)
            # Copy the file from the SFTP server to the local machine
            # ulozi zip subor
            sftp.get(remote_file_path, local_file_path)

            # Ulozim si nazov priecinkov ktore su odzipovane
            unzip_file = local_file_path[:-4]
            print(unzip_file)

            # rozbali zip subor bez poslednych troch pismen
            zf = ZipFile(local_file_path, 'r')
            zf.extractall(local_file_path[:-4])
            zf.close()

            # vymaze zip subor
            print('vymazanie suboru')
            print(local_file_path)
            if os.path.isfile(local_file_path):
                os.remove(local_file_path)

    # Close the SFTP session
    sftp.close()

    # Close the SSH client object
    ssh.close()

    # create txt file
    folder = filepath = BASE_DIR + top_storage + storage
    create_files = search_file+'_'+name+'.txt'
    print (create_files)
    try:
        with open(folder + create_files, 'w') as f:
            f.write('---------------------------------------------  ALL '+name+'Logs from '+storage+' ----------\n')
            print("Create is_log.txt")
    except FileNotFoundError:
        print("The 'docs' directory does not exist")
    folder_emsaapp01_log = BASE_DIR + top_storage + storage +create_files
    search_folder = filepath = BASE_DIR + top_storage + storage + search_file +'/'

    # append log in text
    for filename in os.listdir(search_folder):
        # if filename.endswith('.log'):       # konciaci na .log
        # if filename.startswith('We', 7, 9):   # we je na 7 az 9 mieste
        if name in str(filename):
    #         shutil.copy( source_folder + filename, destination_folder)
    #         # zapis do suboru web log txt
            with open(( search_folder + filename), 'r', encoding="utf-8") as reader, open(folder_emsaapp01_log,'a', encoding="utf-8") as a_writer:
                dog_breeds = reader.readlines()
                a_writer.write('\n')
                a_writer.write('\n')
                a_writer.write('\n'+storage+'/  ')
                a_writer.writelines(filename)
                a_writer.write(' \n')
                a_writer.write('  \n')
                a_writer.writelines((dog_breeds))
                a_writer.close()
                # koniec zapisu
                # print('copied /EMSAAPP02_ErrorLogs/', filename )

    # ----download file ------------------
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    filename = create_files
    # Define the full file path
    filepath = BASE_DIR + top_storage + storage + filename
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    if isExist:
        shutil.rmtree(search_folder)
    # Return the response value
    return response

def EMSISAPP02_realtime_app(request):

    # Create an SSH client object
    ssh = paramiko.SSHClient()

    # Automatically add the server's host key (you can disable this if you want to manually check the host key)
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # Connect to the SFTP server
    ssh.connect(hostname, port, username, password)

    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # delete dir

    top_storage = '/ems_logs/'
    storage = 'EMSISAPP02/'
    search_file = 'EMSISAPP02_realtime'
    name = 'App'

    dir = filepath = BASE_DIR + top_storage + storage
    isExist = os.path.exists(dir)
    if isExist:
        shutil.rmtree(dir)
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    else:
        # create dir
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    # Open an SFTP session
    sftp = ssh.open_sftp()
    print('pristup')
    print(sftp)

    # Define the remote directory's path
    remote_directory_path = '/srv/archive/log/upload-by-smbfs/billien-ems/temp/' #'/srv/archive/log/upload-by-smbfs/billien-edz/realtime/EDZISAPP02/'
    print('ulozisko'+remote_directory_path)

    # Define the local directory's path
    local_directory_path = BASE_DIR + top_storage + storage
    # local_directory_path = os.listdir(r'\fileserver.internal.example.com\c$\Users\app_user\Downloads')
    print('localne ulozisko'+local_directory_path)

    # Get a list of all files in the remote directory
    files = sftp.listdir(remote_directory_path)
    print('sftp files')
    print(files)

    # Iterate over each file in the remote directory
    for file in files:
        # If the file's name contains "error", copy it from the SFTP server to the local machine
        # if 'EMSISAPP0' in file:
        if search_file in file:
            print('files')
            print(f'Copying {file}')
            # print(file)

            # Define the remote file's path
            remote_file_path = remote_directory_path+file
            print('remote_file_path')
            print(remote_file_path)
            # Define the local file's path
            local_file_path = local_directory_path+file
            print('local_file_path')
            print(local_file_path)
            # Copy the file from the SFTP server to the local machine
            # ulozi zip subor
            sftp.get(remote_file_path, local_file_path)

            # Ulozim si nazov priecinkov ktore su odzipovane
            unzip_file = local_file_path[:-4]
            print(unzip_file)

            # rozbali zip subor bez poslednych troch pismen
            zf = ZipFile(local_file_path, 'r')
            zf.extractall(local_file_path[:-4])
            zf.close()

            # vymaze zip subor
            print('vymazanie suboru')
            print(local_file_path)
            if os.path.isfile(local_file_path):
                os.remove(local_file_path)

    # Close the SFTP session
    sftp.close()

    # Close the SSH client object
    ssh.close()

    # create txt file
    folder = filepath = BASE_DIR + top_storage + storage
    create_files = search_file+'_'+name+'.txt'
    print (create_files)
    try:
        with open(folder + create_files, 'w') as f:
            f.write('---------------------------------------------  ALL '+name+'Logs from '+storage+' ----------\n')
            print("Create is_log.txt")
    except FileNotFoundError:
        print("The 'docs' directory does not exist")
    folder_emsaapp01_log = BASE_DIR + top_storage + storage +create_files
    search_folder = filepath = BASE_DIR + top_storage + storage + search_file +'/'

    # append log in text
    for filename in os.listdir(search_folder):
        # if filename.endswith('.log'):       # konciaci na .log
        # if filename.startswith('We', 7, 9):   # we je na 7 az 9 mieste
        if name in str(filename):
    #         shutil.copy( source_folder + filename, destination_folder)
    #         # zapis do suboru web log txt
            with open(( search_folder + filename), 'r', encoding="utf-8") as reader, open(folder_emsaapp01_log,'a', encoding="utf-8") as a_writer:
                dog_breeds = reader.readlines()
                a_writer.write('\n')
                a_writer.write('\n')
                a_writer.write('\n'+storage+'/  ')
                a_writer.writelines(filename)
                a_writer.write(' \n')
                a_writer.write('  \n')
                a_writer.writelines((dog_breeds))
                a_writer.close()
                # koniec zapisu
                # print('copied /EMSAAPP02_ErrorLogs/', filename )

    # ----download file ------------------
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    filename = create_files
    # Define the full file path
    filepath = BASE_DIR + top_storage + storage + filename
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    if isExist:
        shutil.rmtree(search_folder)
    # Return the response value
    return response

def EMSAWEB01_bo_realtime_web(request):

    # Create an SSH client object
    ssh = paramiko.SSHClient()

    # Automatically add the server's host key (you can disable this if you want to manually check the host key)
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # Connect to the SFTP server
    ssh.connect(hostname, port, username, password)

    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # delete dir

    top_storage = '/ems_logs/'
    storage = 'EMSAWEB01_bo/'
    search_file = 'EMSAWEB01_bo_realtime'
    name = 'Web'

    dir = filepath = BASE_DIR + top_storage + storage
    isExist = os.path.exists(dir)
    if isExist:
        shutil.rmtree(dir)
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    else:
        # create dir
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    # Open an SFTP session
    sftp = ssh.open_sftp()
    print('pristup')
    print(sftp)

    # Define the remote directory's path
    remote_directory_path = '/srv/archive/log/upload-by-smbfs/billien-ems/temp/' #'/srv/archive/log/upload-by-smbfs/billien-edz/realtime/EDZISAPP02/'
    print('ulozisko'+remote_directory_path)

    # Define the local directory's path
    local_directory_path = BASE_DIR + top_storage + storage
    # local_directory_path = os.listdir(r'\fileserver.internal.example.com\c$\Users\app_user\Downloads')
    print('localne ulozisko'+local_directory_path)

    # Get a list of all files in the remote directory
    files = sftp.listdir(remote_directory_path)
    print('sftp files')
    print(files)

    # Iterate over each file in the remote directory
    for file in files:
        # If the file's name contains "error", copy it from the SFTP server to the local machine
        # if 'EMSISAPP0' in file:
        if search_file in file:
            print('files')
            print(f'Copying {file}')
            # print(file)

            # Define the remote file's path
            remote_file_path = remote_directory_path+file
            print('remote_file_path')
            print(remote_file_path)
            # Define the local file's path
            local_file_path = local_directory_path+file
            print('local_file_path')
            print(local_file_path)
            # Copy the file from the SFTP server to the local machine
            # ulozi zip subor
            sftp.get(remote_file_path, local_file_path)

            # Ulozim si nazov priecinkov ktore su odzipovane
            unzip_file = local_file_path[:-4]
            print(unzip_file)

            # rozbali zip subor bez poslednych troch pismen
            zf = ZipFile(local_file_path, 'r')
            zf.extractall(local_file_path[:-4])
            zf.close()

            # vymaze zip subor
            print('vymazanie suboru')
            print(local_file_path)
            if os.path.isfile(local_file_path):
                os.remove(local_file_path)

    # Close the SFTP session
    sftp.close()

    # Close the SSH client object
    ssh.close()

    # create txt file
    folder = filepath = BASE_DIR + top_storage + storage
    create_files = search_file+'_'+name+'.txt'

    try:
        with open(folder + create_files, 'w') as f:
            f.write('---------------------------------------------  ALL '+name+'Logs from '+storage+' ----------\n')
            print("Create is_log.txt")
    except FileNotFoundError:
        print("The 'docs' directory does not exist")
    folder_emsaapp01_log = BASE_DIR + top_storage + storage +create_files
    search_folder = filepath = BASE_DIR + top_storage + storage + search_file +'/realtime/'
    print ('search_folder')
    print (search_folder)
    # append log in text
    for filename in os.listdir(search_folder):
        # if filename.endswith('.log'):       # konciaci na .log
        # if filename.startswith('We', 7, 9):   # we je na 7 az 9 mieste
        if name in str(filename):
    #         shutil.copy( source_folder + filename, destination_folder)
    #         # zapis do suboru web log txt
            with open(( search_folder + filename), 'r', encoding="utf-8") as reader, open(folder_emsaapp01_log,'a', encoding="utf-8") as a_writer:
                dog_breeds = reader.readlines()
                a_writer.write('\n')
                a_writer.write('\n')
                a_writer.write('\n'+storage+'/  ')
                a_writer.writelines(filename)
                a_writer.write(' \n')
                a_writer.write('  \n')
                a_writer.writelines((dog_breeds))
                a_writer.close()
                # koniec zapisu
                # print('copied /EMSAAPP02_ErrorLogs/', filename )

    # ----download file ------------------
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    filename = create_files
    # Define the full file path
    filepath = BASE_DIR + top_storage + storage + filename
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    # if isExist:
    #     shutil.rmtree(search_folder)
    # Return the response value
    return response

def EMSAWEB02_bo_realtime_web(request):

    # Create an SSH client object
    ssh = paramiko.SSHClient()

    # Automatically add the server's host key (you can disable this if you want to manually check the host key)
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # Connect to the SFTP server
    ssh.connect(hostname, port, username, password)

    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # delete dir

    top_storage = '/ems_logs/'
    storage = 'EMSAWEB02_bo/'
    search_file = 'EMSAWEB02_bo_realtime'
    name = 'Web'

    dir = filepath = BASE_DIR + top_storage + storage
    isExist = os.path.exists(dir)
    if isExist:
        shutil.rmtree(dir)
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    else:
        # create dir
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    # Open an SFTP session
    sftp = ssh.open_sftp()
    print('pristup')
    print(sftp)

    # Define the remote directory's path
    remote_directory_path = '/srv/archive/log/upload-by-smbfs/billien-ems/temp/' #'/srv/archive/log/upload-by-smbfs/billien-edz/realtime/EDZISAPP02/'
    print('ulozisko'+remote_directory_path)

    # Define the local directory's path
    local_directory_path = BASE_DIR + top_storage + storage
    # local_directory_path = os.listdir(r'\fileserver.internal.example.com\c$\Users\app_user\Downloads')
    print('localne ulozisko'+local_directory_path)

    # Get a list of all files in the remote directory
    files = sftp.listdir(remote_directory_path)
    print('sftp files')
    print(files)

    # Iterate over each file in the remote directory
    for file in files:
        # If the file's name contains "error", copy it from the SFTP server to the local machine
        # if 'EMSISAPP0' in file:
        if search_file in file:
            print('files')
            print(f'Copying {file}')
            # print(file)

            # Define the remote file's path
            remote_file_path = remote_directory_path+file
            print('remote_file_path')
            print(remote_file_path)
            # Define the local file's path
            local_file_path = local_directory_path+file
            print('local_file_path')
            print(local_file_path)
            # Copy the file from the SFTP server to the local machine
            # ulozi zip subor
            sftp.get(remote_file_path, local_file_path)

            # Ulozim si nazov priecinkov ktore su odzipovane
            unzip_file = local_file_path[:-4]
            print(unzip_file)

            # rozbali zip subor bez poslednych troch pismen
            zf = ZipFile(local_file_path, 'r')
            zf.extractall(local_file_path[:-4])
            zf.close()

            # vymaze zip subor
            print('vymazanie suboru')
            print(local_file_path)
            if os.path.isfile(local_file_path):
                os.remove(local_file_path)

    # Close the SFTP session
    sftp.close()

    # Close the SSH client object
    ssh.close()

    # create txt file
    folder = filepath = BASE_DIR + top_storage + storage
    create_files = search_file+'_'+name+'.txt'

    try:
        with open(folder + create_files, 'w') as f:
            f.write('---------------------------------------------  ALL '+name+'Logs from '+storage+' ----------\n')
            print("Create is_log.txt")
    except FileNotFoundError:
        print("The 'docs' directory does not exist")
    folder_emsaapp01_log = BASE_DIR + top_storage + storage +create_files
    search_folder = filepath = BASE_DIR + top_storage + storage + search_file +'/realtime/'
    print ('search_folder')
    print (search_folder)
    # append log in text
    for filename in os.listdir(search_folder):
        # if filename.endswith('.log'):       # konciaci na .log
        # if filename.startswith('We', 7, 9):   # we je na 7 az 9 mieste
        if name in str(filename):
    #         shutil.copy( source_folder + filename, destination_folder)
    #         # zapis do suboru web log txt
            with open(( search_folder + filename), 'r', encoding="utf-8") as reader, open(folder_emsaapp01_log,'a', encoding="utf-8") as a_writer:
                dog_breeds = reader.readlines()
                a_writer.write('\n')
                a_writer.write('\n')
                a_writer.write('\n'+storage+'/  ')
                a_writer.writelines(filename)
                a_writer.write(' \n')
                a_writer.write('  \n')
                a_writer.writelines((dog_breeds))
                a_writer.close()
                # koniec zapisu
                # print('copied /EMSAAPP02_ErrorLogs/', filename )

    # ----download file ------------------
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    filename = create_files
    # Define the full file path
    filepath = BASE_DIR + top_storage + storage + filename
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    # if isExist:
    #     shutil.rmtree(search_folder)
    # Return the response value
    return response

def EMSAWEB01_pos_realtime_web(request):

    # Create an SSH client object
    ssh = paramiko.SSHClient()

    # Automatically add the server's host key (you can disable this if you want to manually check the host key)
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # Connect to the SFTP server
    ssh.connect(hostname, port, username, password)

    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # delete dir

    top_storage = '/ems_logs/'
    storage = 'EMSAWEB01_pos/'
    search_file = 'EMSAWEB01_pos_realtime'
    name = 'Web'

    dir = filepath = BASE_DIR + top_storage + storage
    isExist = os.path.exists(dir)
    if isExist:
        shutil.rmtree(dir)
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    else:
        # create dir
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    # Open an SFTP session
    sftp = ssh.open_sftp()
    print('pristup')
    print(sftp)

    # Define the remote directory's path
    remote_directory_path = '/srv/archive/log/upload-by-smbfs/billien-ems/temp/' #'/srv/archive/log/upload-by-smbfs/billien-edz/realtime/EDZISAPP02/'
    print('ulozisko'+remote_directory_path)

    # Define the local directory's path
    local_directory_path = BASE_DIR + top_storage + storage
    # local_directory_path = os.listdir(r'\fileserver.internal.example.com\c$\Users\app_user\Downloads')
    print('localne ulozisko'+local_directory_path)

    # Get a list of all files in the remote directory
    files = sftp.listdir(remote_directory_path)
    print('sftp files')
    print(files)

    # Iterate over each file in the remote directory
    for file in files:
        # If the file's name contains "error", copy it from the SFTP server to the local machine
        # if 'EMSISAPP0' in file:
        if search_file in file:
            print('files')
            print(f'Copying {file}')
            # print(file)

            # Define the remote file's path
            remote_file_path = remote_directory_path+file
            print('remote_file_path')
            print(remote_file_path)
            # Define the local file's path
            local_file_path = local_directory_path+file
            print('local_file_path')
            print(local_file_path)
            # Copy the file from the SFTP server to the local machine
            # ulozi zip subor
            sftp.get(remote_file_path, local_file_path)

            # Ulozim si nazov priecinkov ktore su odzipovane
            unzip_file = local_file_path[:-4]
            print(unzip_file)

            # rozbali zip subor bez poslednych troch pismen
            zf = ZipFile(local_file_path, 'r')
            zf.extractall(local_file_path[:-4])
            zf.close()

            # vymaze zip subor
            print('vymazanie suboru')
            print(local_file_path)
            if os.path.isfile(local_file_path):
                os.remove(local_file_path)

    # Close the SFTP session
    sftp.close()

    # Close the SSH client object
    ssh.close()

    # create txt file
    folder = filepath = BASE_DIR + top_storage + storage
    create_files = search_file+'_'+name+'.txt'

    try:
        with open(folder + create_files, 'w') as f:
            f.write('---------------------------------------------  ALL '+name+'Logs from '+storage+' ----------\n')
            print("Create is_log.txt")
    except FileNotFoundError:
        print("The 'docs' directory does not exist")
    folder_emsaapp01_log = BASE_DIR + top_storage + storage +create_files
    search_folder = filepath = BASE_DIR + top_storage + storage + search_file +'/realtime/'
    print ('search_folder')
    print (search_folder)
    # append log in text
    for filename in os.listdir(search_folder):
        # if filename.endswith('.log'):       # konciaci na .log
        # if filename.startswith('We', 7, 9):   # we je na 7 az 9 mieste
        if name in str(filename):
    #         shutil.copy( source_folder + filename, destination_folder)
    #         # zapis do suboru web log txt
            with open(( search_folder + filename), 'r', encoding="utf-8") as reader, open(folder_emsaapp01_log,'a', encoding="utf-8") as a_writer:
                dog_breeds = reader.readlines()
                a_writer.write('\n')
                a_writer.write('\n')
                a_writer.write('\n'+storage+'/  ')
                a_writer.writelines(filename)
                a_writer.write(' \n')
                a_writer.write('  \n')
                a_writer.writelines((dog_breeds))
                a_writer.close()
                # koniec zapisu
                # print('copied /EMSAAPP02_ErrorLogs/', filename )

    # ----download file ------------------
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    filename = create_files
    # Define the full file path
    filepath = BASE_DIR + top_storage + storage + filename
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    # if isExist:
    #     shutil.rmtree(search_folder)
    # Return the response value
    return response

def EMSAWEB02_pos_realtime_web(request):

    # Create an SSH client object
    ssh = paramiko.SSHClient()

    # Automatically add the server's host key (you can disable this if you want to manually check the host key)
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # Connect to the SFTP server
    ssh.connect(hostname, port, username, password)

    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # delete dir

    top_storage = '/ems_logs/'
    storage = 'EMSAWEB02_pos/'
    search_file = 'EMSAWEB02_pos_realtime'
    name = 'Web'

    dir = filepath = BASE_DIR + top_storage + storage
    isExist = os.path.exists(dir)
    if isExist:
        shutil.rmtree(dir)
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    else:
        # create dir
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    # Open an SFTP session
    sftp = ssh.open_sftp()
    print('pristup')
    print(sftp)

    # Define the remote directory's path
    remote_directory_path = '/srv/archive/log/upload-by-smbfs/billien-ems/temp/' #'/srv/archive/log/upload-by-smbfs/billien-edz/realtime/EDZISAPP02/'
    print('ulozisko'+remote_directory_path)

    # Define the local directory's path
    local_directory_path = BASE_DIR + top_storage + storage
    # local_directory_path = os.listdir(r'\fileserver.internal.example.com\c$\Users\app_user\Downloads')
    print('localne ulozisko'+local_directory_path)

    # Get a list of all files in the remote directory
    files = sftp.listdir(remote_directory_path)
    print('sftp files')
    print(files)

    # Iterate over each file in the remote directory
    for file in files:
        # If the file's name contains "error", copy it from the SFTP server to the local machine
        # if 'EMSISAPP0' in file:
        if search_file in file:
            print('files')
            print(f'Copying {file}')
            # print(file)

            # Define the remote file's path
            remote_file_path = remote_directory_path+file
            print('remote_file_path')
            print(remote_file_path)
            # Define the local file's path
            local_file_path = local_directory_path+file
            print('local_file_path')
            print(local_file_path)
            # Copy the file from the SFTP server to the local machine
            # ulozi zip subor
            sftp.get(remote_file_path, local_file_path)

            # Ulozim si nazov priecinkov ktore su odzipovane
            unzip_file = local_file_path[:-4]
            print(unzip_file)

            # rozbali zip subor bez poslednych troch pismen
            zf = ZipFile(local_file_path, 'r')
            zf.extractall(local_file_path[:-4])
            zf.close()

            # vymaze zip subor
            print('vymazanie suboru')
            print(local_file_path)
            if os.path.isfile(local_file_path):
                os.remove(local_file_path)

    # Close the SFTP session
    sftp.close()

    # Close the SSH client object
    ssh.close()

    # create txt file
    folder = filepath = BASE_DIR + top_storage + storage
    create_files = search_file+'_'+name+'.txt'

    try:
        with open(folder + create_files, 'w') as f:
            f.write('---------------------------------------------  ALL '+name+'Logs from '+storage+' ----------\n')
            print("Create is_log.txt")
    except FileNotFoundError:
        print("The 'docs' directory does not exist")
    folder_emsaapp01_log = BASE_DIR + top_storage + storage +create_files
    search_folder = filepath = BASE_DIR + top_storage + storage + search_file +'/realtime/'
    print ('search_folder')
    print (search_folder)
    # append log in text
    for filename in os.listdir(search_folder):
        # if filename.endswith('.log'):       # konciaci na .log
        # if filename.startswith('We', 7, 9):   # we je na 7 az 9 mieste
        if name in str(filename):
    #         shutil.copy( source_folder + filename, destination_folder)
    #         # zapis do suboru web log txt
            with open(( search_folder + filename), 'r', encoding="utf-8") as reader, open(folder_emsaapp01_log,'a', encoding="utf-8") as a_writer:
                dog_breeds = reader.readlines()
                a_writer.write('\n')
                a_writer.write('\n')
                a_writer.write('\n'+storage+'/  ')
                a_writer.writelines(filename)
                a_writer.write(' \n')
                a_writer.write('  \n')
                a_writer.writelines((dog_breeds))
                a_writer.close()
                # koniec zapisu
                # print('copied /EMSAAPP02_ErrorLogs/', filename )

    # ----download file ------------------
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    filename = create_files
    # Define the full file path
    filepath = BASE_DIR + top_storage + storage + filename
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    # if isExist:
    #     shutil.rmtree(search_folder)
    # Return the response value
    return response

def EMSAWEB01_cr_realtime_web(request):

    # Create an SSH client object
    ssh = paramiko.SSHClient()

    # Automatically add the server's host key (you can disable this if you want to manually check the host key)
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # Connect to the SFTP server
    ssh.connect(hostname, port, username, password)

    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # delete dir

    top_storage = '/ems_logs/'
    storage = 'EMSAWEB01_cr/'
    search_file = 'EMSAWEB01_cr_realtime'
    name = 'Web'

    dir = filepath = BASE_DIR + top_storage + storage
    isExist = os.path.exists(dir)
    if isExist:
        shutil.rmtree(dir)
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    else:
        # create dir
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    # Open an SFTP session
    sftp = ssh.open_sftp()
    print('pristup')
    print(sftp)

    # Define the remote directory's path
    remote_directory_path = '/srv/archive/log/upload-by-smbfs/billien-ems/temp/' #'/srv/archive/log/upload-by-smbfs/billien-edz/realtime/EDZISAPP02/'
    print('ulozisko'+remote_directory_path)

    # Define the local directory's path
    local_directory_path = BASE_DIR + top_storage + storage
    # local_directory_path = os.listdir(r'\fileserver.internal.example.com\c$\Users\app_user\Downloads')
    print('localne ulozisko'+local_directory_path)

    # Get a list of all files in the remote directory
    files = sftp.listdir(remote_directory_path)
    print('sftp files')
    print(files)

    # Iterate over each file in the remote directory
    for file in files:
        # If the file's name contains "error", copy it from the SFTP server to the local machine
        # if 'EMSISAPP0' in file:
        if search_file in file:
            print('files')
            print(f'Copying {file}')
            # print(file)

            # Define the remote file's path
            remote_file_path = remote_directory_path+file
            print('remote_file_path')
            print(remote_file_path)
            # Define the local file's path
            local_file_path = local_directory_path+file
            print('local_file_path')
            print(local_file_path)
            # Copy the file from the SFTP server to the local machine
            # ulozi zip subor
            sftp.get(remote_file_path, local_file_path)

            # Ulozim si nazov priecinkov ktore su odzipovane
            unzip_file = local_file_path[:-4]
            print(unzip_file)

            # rozbali zip subor bez poslednych troch pismen
            zf = ZipFile(local_file_path, 'r')
            zf.extractall(local_file_path[:-4])
            zf.close()

            # vymaze zip subor
            print('vymazanie suboru')
            print(local_file_path)
            if os.path.isfile(local_file_path):
                os.remove(local_file_path)

    # Close the SFTP session
    sftp.close()

    # Close the SSH client object
    ssh.close()

    # create txt file
    folder = filepath = BASE_DIR + top_storage + storage
    create_files = search_file+'_'+name+'.txt'

    try:
        with open(folder + create_files, 'w') as f:
            f.write('---------------------------------------------  ALL '+name+'Logs from '+storage+' ----------\n')
            print("Create is_log.txt")
    except FileNotFoundError:
        print("The 'docs' directory does not exist")
    folder_emsaapp01_log = BASE_DIR + top_storage + storage +create_files
    search_folder = filepath = BASE_DIR + top_storage + storage + search_file +'/realtime/'
    print ('search_folder')
    print (search_folder)
    # append log in text
    for filename in os.listdir(search_folder):
        # if filename.endswith('.log'):       # konciaci na .log
        # if filename.startswith('We', 7, 9):   # we je na 7 az 9 mieste
        if name in str(filename):
    #         shutil.copy( source_folder + filename, destination_folder)
    #         # zapis do suboru web log txt
            with open(( search_folder + filename), 'r', encoding="utf-8") as reader, open(folder_emsaapp01_log,'a', encoding="utf-8") as a_writer:
                dog_breeds = reader.readlines()
                a_writer.write('\n')
                a_writer.write('\n')
                a_writer.write('\n'+storage+'/  ')
                a_writer.writelines(filename)
                a_writer.write(' \n')
                a_writer.write('  \n')
                a_writer.writelines((dog_breeds))
                a_writer.close()
                # koniec zapisu
                # print('copied /EMSAAPP02_ErrorLogs/', filename )

    # ----download file ------------------
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    filename = create_files
    # Define the full file path
    filepath = BASE_DIR + top_storage + storage + filename
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    # if isExist:
    #     shutil.rmtree(search_folder)
    # Return the response value
    return response

def EMSAWEB02_cr_realtime_web(request):

    # Create an SSH client object
    ssh = paramiko.SSHClient()

    # Automatically add the server's host key (you can disable this if you want to manually check the host key)
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # Connect to the SFTP server
    ssh.connect(hostname, port, username, password)

    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # delete dir

    top_storage = '/ems_logs/'
    storage = 'EMSAWEB02_cr/'
    search_file = 'EMSAWEB02_cr_realtime'
    name = 'Web'

    dir = filepath = BASE_DIR + top_storage + storage
    isExist = os.path.exists(dir)
    if isExist:
        shutil.rmtree(dir)
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    else:
        # create dir
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    # Open an SFTP session
    sftp = ssh.open_sftp()
    print('pristup')
    print(sftp)

    # Define the remote directory's path
    remote_directory_path = '/srv/archive/log/upload-by-smbfs/billien-ems/temp/' #'/srv/archive/log/upload-by-smbfs/billien-edz/realtime/EDZISAPP02/'
    print('ulozisko'+remote_directory_path)

    # Define the local directory's path
    local_directory_path = BASE_DIR + top_storage + storage
    # local_directory_path = os.listdir(r'\fileserver.internal.example.com\c$\Users\app_user\Downloads')
    print('localne ulozisko'+local_directory_path)

    # Get a list of all files in the remote directory
    files = sftp.listdir(remote_directory_path)
    print('sftp files')
    print(files)

    # Iterate over each file in the remote directory
    for file in files:
        # If the file's name contains "error", copy it from the SFTP server to the local machine
        # if 'EMSISAPP0' in file:
        if search_file in file:
            print('files')
            print(f'Copying {file}')
            # print(file)

            # Define the remote file's path
            remote_file_path = remote_directory_path+file
            print('remote_file_path')
            print(remote_file_path)
            # Define the local file's path
            local_file_path = local_directory_path+file
            print('local_file_path')
            print(local_file_path)
            # Copy the file from the SFTP server to the local machine
            # ulozi zip subor
            sftp.get(remote_file_path, local_file_path)

            # Ulozim si nazov priecinkov ktore su odzipovane
            unzip_file = local_file_path[:-4]
            print(unzip_file)

            # rozbali zip subor bez poslednych troch pismen
            zf = ZipFile(local_file_path, 'r')
            zf.extractall(local_file_path[:-4])
            zf.close()

            # vymaze zip subor
            print('vymazanie suboru')
            print(local_file_path)
            if os.path.isfile(local_file_path):
                os.remove(local_file_path)

    # Close the SFTP session
    sftp.close()

    # Close the SSH client object
    ssh.close()

    # create txt file
    folder = filepath = BASE_DIR + top_storage + storage
    create_files = search_file+'_'+name+'.txt'

    try:
        with open(folder + create_files, 'w') as f:
            f.write('---------------------------------------------  ALL '+name+'Logs from '+storage+' ----------\n')
            print("Create is_log.txt")
    except FileNotFoundError:
        print("The 'docs' directory does not exist")
    folder_emsaapp01_log = BASE_DIR + top_storage + storage +create_files
    search_folder = filepath = BASE_DIR + top_storage + storage + search_file +'/realtime/'
    print ('search_folder')
    print (search_folder)
    # append log in text
    for filename in os.listdir(search_folder):
        # if filename.endswith('.log'):       # konciaci na .log
        # if filename.startswith('We', 7, 9):   # we je na 7 az 9 mieste
        if name in str(filename):
    #         shutil.copy( source_folder + filename, destination_folder)
    #         # zapis do suboru web log txt
            with open(( search_folder + filename), 'r', encoding="utf-8") as reader, open(folder_emsaapp01_log,'a', encoding="utf-8") as a_writer:
                dog_breeds = reader.readlines()
                a_writer.write('\n')
                a_writer.write('\n')
                a_writer.write('\n'+storage+'/  ')
                a_writer.writelines(filename)
                a_writer.write(' \n')
                a_writer.write('  \n')
                a_writer.writelines((dog_breeds))
                a_writer.close()
                # koniec zapisu
                # print('copied /EMSAAPP02_ErrorLogs/', filename )

    # ----download file ------------------
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    filename = create_files
    # Define the full file path
    filepath = BASE_DIR + top_storage + storage + filename
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    # if isExist:
    #     shutil.rmtree(search_folder)
    # Return the response value
    return response

def EMSAWEB01_fci_realtime_web(request):

    # Create an SSH client object
    ssh = paramiko.SSHClient()

    # Automatically add the server's host key (you can disable this if you want to manually check the host key)
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # Connect to the SFTP server
    ssh.connect(hostname, port, username, password)

    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # delete dir

    top_storage = '/ems_logs/'
    storage = 'EMSAWEB01_fci/'
    search_file = 'EMSAWEB01_fci_realtime'
    name = 'Web'

    dir = filepath = BASE_DIR + top_storage + storage
    isExist = os.path.exists(dir)
    if isExist:
        shutil.rmtree(dir)
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    else:
        # create dir
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    # Open an SFTP session
    sftp = ssh.open_sftp()
    print('pristup')
    print(sftp)

    # Define the remote directory's path
    remote_directory_path = '/srv/archive/log/upload-by-smbfs/billien-ems/temp/' #'/srv/archive/log/upload-by-smbfs/billien-edz/realtime/EDZISAPP02/'
    print('ulozisko'+remote_directory_path)

    # Define the local directory's path
    local_directory_path = BASE_DIR + top_storage + storage
    # local_directory_path = os.listdir(r'\fileserver.internal.example.com\c$\Users\app_user\Downloads')
    print('localne ulozisko'+local_directory_path)

    # Get a list of all files in the remote directory
    files = sftp.listdir(remote_directory_path)
    print('sftp files')
    print(files)

    # Iterate over each file in the remote directory
    for file in files:
        # If the file's name contains "error", copy it from the SFTP server to the local machine
        # if 'EMSISAPP0' in file:
        if search_file in file:
            print('files')
            print(f'Copying {file}')
            # print(file)

            # Define the remote file's path
            remote_file_path = remote_directory_path+file
            print('remote_file_path')
            print(remote_file_path)
            # Define the local file's path
            local_file_path = local_directory_path+file
            print('local_file_path')
            print(local_file_path)
            # Copy the file from the SFTP server to the local machine
            # ulozi zip subor
            sftp.get(remote_file_path, local_file_path)

            # Ulozim si nazov priecinkov ktore su odzipovane
            unzip_file = local_file_path[:-4]
            print(unzip_file)

            # rozbali zip subor bez poslednych troch pismen
            zf = ZipFile(local_file_path, 'r')
            zf.extractall(local_file_path[:-4])
            zf.close()

            # vymaze zip subor
            print('vymazanie suboru')
            print(local_file_path)
            if os.path.isfile(local_file_path):
                os.remove(local_file_path)

    # Close the SFTP session
    sftp.close()

    # Close the SSH client object
    ssh.close()

    # create txt file
    folder = filepath = BASE_DIR + top_storage + storage
    create_files = search_file+'_'+name+'.txt'

    try:
        with open(folder + create_files, 'w') as f:
            f.write('---------------------------------------------  ALL '+name+'Logs from '+storage+' ----------\n')
            print("Create is_log.txt")
    except FileNotFoundError:
        print("The 'docs' directory does not exist")
    folder_emsaapp01_log = BASE_DIR + top_storage + storage +create_files
    search_folder = filepath = BASE_DIR + top_storage + storage + search_file +'/realtime/'
    print ('search_folder')
    print (search_folder)
    # append log in text
    for filename in os.listdir(search_folder):
        # if filename.endswith('.log'):       # konciaci na .log
        # if filename.startswith('We', 7, 9):   # we je na 7 az 9 mieste
        if name in str(filename):
    #         shutil.copy( source_folder + filename, destination_folder)
    #         # zapis do suboru web log txt
            with open(( search_folder + filename), 'r', encoding="utf-8") as reader, open(folder_emsaapp01_log,'a', encoding="utf-8") as a_writer:
                dog_breeds = reader.readlines()
                a_writer.write('\n')
                a_writer.write('\n')
                a_writer.write('\n'+storage+'/  ')
                a_writer.writelines(filename)
                a_writer.write(' \n')
                a_writer.write('  \n')
                a_writer.writelines((dog_breeds))
                a_writer.close()
                # koniec zapisu
                # print('copied /EMSAAPP02_ErrorLogs/', filename )

    # ----download file ------------------
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    filename = create_files
    # Define the full file path
    filepath = BASE_DIR + top_storage + storage + filename
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    # if isExist:
    #     shutil.rmtree(search_folder)
    # Return the response value
    return response

def EMSAWEB02_fci_realtime_web(request):

    # Create an SSH client object
    ssh = paramiko.SSHClient()

    # Automatically add the server's host key (you can disable this if you want to manually check the host key)
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # Connect to the SFTP server
    ssh.connect(hostname, port, username, password)

    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # delete dir

    top_storage = '/ems_logs/'
    storage = 'EMSAWEB02_fci/'
    search_file = 'EMSAWEB02_fci_realtime'
    name = 'Web'

    dir = filepath = BASE_DIR + top_storage + storage
    isExist = os.path.exists(dir)
    if isExist:
        shutil.rmtree(dir)
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    else:
        # create dir
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    # Open an SFTP session
    sftp = ssh.open_sftp()
    print('pristup')
    print(sftp)

    # Define the remote directory's path
    remote_directory_path = '/srv/archive/log/upload-by-smbfs/billien-ems/temp/' #'/srv/archive/log/upload-by-smbfs/billien-edz/realtime/EDZISAPP02/'
    print('ulozisko'+remote_directory_path)

    # Define the local directory's path
    local_directory_path = BASE_DIR + top_storage + storage
    # local_directory_path = os.listdir(r'\fileserver.internal.example.com\c$\Users\app_user\Downloads')
    print('localne ulozisko'+local_directory_path)

    # Get a list of all files in the remote directory
    files = sftp.listdir(remote_directory_path)
    print('sftp files')
    print(files)

    # Iterate over each file in the remote directory
    for file in files:
        # If the file's name contains "error", copy it from the SFTP server to the local machine
        # if 'EMSISAPP0' in file:
        if search_file in file:
            print('files')
            print(f'Copying {file}')
            # print(file)

            # Define the remote file's path
            remote_file_path = remote_directory_path+file
            print('remote_file_path')
            print(remote_file_path)
            # Define the local file's path
            local_file_path = local_directory_path+file
            print('local_file_path')
            print(local_file_path)
            # Copy the file from the SFTP server to the local machine
            # ulozi zip subor
            sftp.get(remote_file_path, local_file_path)

            # Ulozim si nazov priecinkov ktore su odzipovane
            unzip_file = local_file_path[:-4]
            print(unzip_file)

            # rozbali zip subor bez poslednych troch pismen
            zf = ZipFile(local_file_path, 'r')
            zf.extractall(local_file_path[:-4])
            zf.close()

            # vymaze zip subor
            print('vymazanie suboru')
            print(local_file_path)
            if os.path.isfile(local_file_path):
                os.remove(local_file_path)

    # Close the SFTP session
    sftp.close()

    # Close the SSH client object
    ssh.close()

    # create txt file
    folder = filepath = BASE_DIR + top_storage + storage
    create_files = search_file+'_'+name+'.txt'

    try:
        with open(folder + create_files, 'w') as f:
            f.write('---------------------------------------------  ALL '+name+'Logs from '+storage+' ----------\n')
            print("Create is_log.txt")
    except FileNotFoundError:
        print("The 'docs' directory does not exist")
    folder_emsaapp01_log = BASE_DIR + top_storage + storage +create_files
    search_folder = filepath = BASE_DIR + top_storage + storage + search_file +'/realtime/'
    print ('search_folder')
    print (search_folder)
    # append log in text
    for filename in os.listdir(search_folder):
        # if filename.endswith('.log'):       # konciaci na .log
        # if filename.startswith('We', 7, 9):   # we je na 7 az 9 mieste
        if name in str(filename):
    #         shutil.copy( source_folder + filename, destination_folder)
    #         # zapis do suboru web log txt
            with open(( search_folder + filename), 'r', encoding="utf-8") as reader, open(folder_emsaapp01_log,'a', encoding="utf-8") as a_writer:
                dog_breeds = reader.readlines()
                a_writer.write('\n')
                a_writer.write('\n')
                a_writer.write('\n'+storage+'/  ')
                a_writer.writelines(filename)
                a_writer.write(' \n')
                a_writer.write('  \n')
                a_writer.writelines((dog_breeds))
                a_writer.close()
                # koniec zapisu
                # print('copied /EMSAAPP02_ErrorLogs/', filename )

    # ----download file ------------------
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    filename = create_files
    # Define the full file path
    filepath = BASE_DIR + top_storage + storage + filename
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    # if isExist:
    #     shutil.rmtree(search_folder)
    # Return the response value
    return response

def EMSAWEB01_sc_realtime_web(request):

    # Create an SSH client object
    ssh = paramiko.SSHClient()

    # Automatically add the server's host key (you can disable this if you want to manually check the host key)
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # Connect to the SFTP server
    ssh.connect(hostname, port, username, password)

    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # delete dir

    top_storage = '/ems_logs/'
    storage = 'EMSAWEB01_sc/'
    search_file = 'EMSAWEB01_sc_realtime'
    name = 'Web'

    dir = filepath = BASE_DIR + top_storage + storage
    isExist = os.path.exists(dir)
    if isExist:
        shutil.rmtree(dir)
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    else:
        # create dir
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    # Open an SFTP session
    sftp = ssh.open_sftp()
    print('pristup')
    print(sftp)

    # Define the remote directory's path
    remote_directory_path = '/srv/archive/log/upload-by-smbfs/billien-ems/temp/' #'/srv/archive/log/upload-by-smbfs/billien-edz/realtime/EDZISAPP02/'
    print('ulozisko'+remote_directory_path)

    # Define the local directory's path
    local_directory_path = BASE_DIR + top_storage + storage
    # local_directory_path = os.listdir(r'\fileserver.internal.example.com\c$\Users\app_user\Downloads')
    print('localne ulozisko'+local_directory_path)

    # Get a list of all files in the remote directory
    files = sftp.listdir(remote_directory_path)
    print('sftp files')
    print(files)

    # Iterate over each file in the remote directory
    for file in files:
        # If the file's name contains "error", copy it from the SFTP server to the local machine
        # if 'EMSISAPP0' in file:
        if search_file in file:
            print('files')
            print(f'Copying {file}')
            # print(file)

            # Define the remote file's path
            remote_file_path = remote_directory_path+file
            print('remote_file_path')
            print(remote_file_path)
            # Define the local file's path
            local_file_path = local_directory_path+file
            print('local_file_path')
            print(local_file_path)
            # Copy the file from the SFTP server to the local machine
            # ulozi zip subor
            sftp.get(remote_file_path, local_file_path)

            # Ulozim si nazov priecinkov ktore su odzipovane
            unzip_file = local_file_path[:-4]
            print(unzip_file)

            # rozbali zip subor bez poslednych troch pismen
            zf = ZipFile(local_file_path, 'r')
            zf.extractall(local_file_path[:-4])
            zf.close()

            # vymaze zip subor
            print('vymazanie suboru')
            print(local_file_path)
            if os.path.isfile(local_file_path):
                os.remove(local_file_path)

    # Close the SFTP session
    sftp.close()

    # Close the SSH client object
    ssh.close()

    # create txt file
    folder = filepath = BASE_DIR + top_storage + storage
    create_files = search_file+'_'+name+'.txt'

    try:
        with open(folder + create_files, 'w') as f:
            f.write('---------------------------------------------  ALL '+name+'Logs from '+storage+' ----------\n')
            print("Create is_log.txt")
    except FileNotFoundError:
        print("The 'docs' directory does not exist")
    folder_emsaapp01_log = BASE_DIR + top_storage + storage +create_files
    search_folder = filepath = BASE_DIR + top_storage + storage + search_file +'/realtime/'
    print ('search_folder')
    print (search_folder)
    # append log in text
    for filename in os.listdir(search_folder):
        # if filename.endswith('.log'):       # konciaci na .log
        # if filename.startswith('We', 7, 9):   # we je na 7 az 9 mieste
        if name in str(filename):
    #         shutil.copy( source_folder + filename, destination_folder)
    #         # zapis do suboru web log txt
            with open(( search_folder + filename), 'r', encoding="utf-8") as reader, open(folder_emsaapp01_log,'a', encoding="utf-8") as a_writer:
                dog_breeds = reader.readlines()
                a_writer.write('\n')
                a_writer.write('\n')
                a_writer.write('\n'+storage+'/  ')
                a_writer.writelines(filename)
                a_writer.write(' \n')
                a_writer.write('  \n')
                a_writer.writelines((dog_breeds))
                a_writer.close()
                # koniec zapisu
                # print('copied /EMSAAPP02_ErrorLogs/', filename )

    # ----download file ------------------
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    filename = create_files
    # Define the full file path
    filepath = BASE_DIR + top_storage + storage + filename
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    if isExist:
        shutil.rmtree(search_folder)
    # Return the response value
    return response

def EMSAWEB02_sc_realtime_web(request):

    # Create an SSH client object
    ssh = paramiko.SSHClient()

    # Automatically add the server's host key (you can disable this if you want to manually check the host key)
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # Connect to the SFTP server
    ssh.connect(hostname, port, username, password)

    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # delete dir

    top_storage = '/ems_logs/'
    storage = 'EMSAWEB02_sc/'
    search_file = 'EMSAWEB02_sc_realtime'
    name = 'Web'

    dir = filepath = BASE_DIR + top_storage + storage
    isExist = os.path.exists(dir)
    if isExist:
        shutil.rmtree(dir)
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    else:
        # create dir
        directory = storage
        parent_dir = BASE_DIR + top_storage
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        print("Directory '% s' created" % directory)
    # Open an SFTP session
    sftp = ssh.open_sftp()
    print('pristup')
    print(sftp)

    # Define the remote directory's path
    remote_directory_path = '/srv/archive/log/upload-by-smbfs/billien-ems/temp/' #'/srv/archive/log/upload-by-smbfs/billien-edz/realtime/EDZISAPP02/'
    print('ulozisko'+remote_directory_path)

    # Define the local directory's path
    local_directory_path = BASE_DIR + top_storage + storage
    # local_directory_path = os.listdir(r'\fileserver.internal.example.com\c$\Users\app_user\Downloads')
    print('localne ulozisko'+local_directory_path)

    # Get a list of all files in the remote directory
    files = sftp.listdir(remote_directory_path)
    print('sftp files')
    print(files)

    # Iterate over each file in the remote directory
    for file in files:
        # If the file's name contains "error", copy it from the SFTP server to the local machine
        # if 'EMSISAPP0' in file:
        if search_file in file:
            print('files')
            print(f'Copying {file}')
            # print(file)

            # Define the remote file's path
            remote_file_path = remote_directory_path+file
            print('remote_file_path')
            print(remote_file_path)
            # Define the local file's path
            local_file_path = local_directory_path+file
            print('local_file_path')
            print(local_file_path)
            # Copy the file from the SFTP server to the local machine
            # ulozi zip subor
            sftp.get(remote_file_path, local_file_path)

            # Ulozim si nazov priecinkov ktore su odzipovane
            unzip_file = local_file_path[:-4]
            print(unzip_file)

            # rozbali zip subor bez poslednych troch pismen
            zf = ZipFile(local_file_path, 'r')
            zf.extractall(local_file_path[:-4])
            zf.close()

            # vymaze zip subor
            print('vymazanie suboru')
            print(local_file_path)
            if os.path.isfile(local_file_path):
                os.remove(local_file_path)

    # Close the SFTP session
    sftp.close()

    # Close the SSH client object
    ssh.close()

    # create txt file
    folder = filepath = BASE_DIR + top_storage + storage
    create_files = search_file+'_'+name+'.txt'

    try:
        with open(folder + create_files, 'w') as f:
            f.write('---------------------------------------------  ALL '+name+'Logs from '+storage+' ----------\n')
            print("Create is_log.txt")
    except FileNotFoundError:
        print("The 'docs' directory does not exist")
    folder_emsaapp01_log = BASE_DIR + top_storage + storage +create_files
    search_folder = filepath = BASE_DIR + top_storage + storage + search_file +'/realtime/'
    print ('search_folder')
    print (search_folder)
    # append log in text
    for filename in os.listdir(search_folder):
        # if filename.endswith('.log'):       # konciaci na .log
        # if filename.startswith('We', 7, 9):   # we je na 7 az 9 mieste
        if name in str(filename):
    #         shutil.copy( source_folder + filename, destination_folder)
    #         # zapis do suboru web log txt
            with open(( search_folder + filename), 'r', encoding="utf-8") as reader, open(folder_emsaapp01_log,'a', encoding="utf-8") as a_writer:
                dog_breeds = reader.readlines()
                a_writer.write('\n')
                a_writer.write('\n')
                a_writer.write('\n'+storage+'/  ')
                a_writer.writelines(filename)
                a_writer.write(' \n')
                a_writer.write('  \n')
                a_writer.writelines((dog_breeds))
                a_writer.close()
                # koniec zapisu
                # print('copied /EMSAAPP02_ErrorLogs/', filename )

    # ----download file ------------------
    # Define Django project base directory
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    filename = create_files
    # Define the full file path
    filepath = BASE_DIR + top_storage + storage + filename
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename

    if isExist:
        shutil.rmtree(search_folder)
    # Return the response value
    return response

# ::::::::::::::::::: REPORTS ::::::::::::::::::::::::::::::::::

def sql_Invoices(request):
    sql_storage = '/Scripts/Denne/'
    query = 'TOIS_Invoices_v17.sql'
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    path = BASE_DIR + sql_storage

    filename = query
    filepath = path + query
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename
    # Return the response value
    return response



def sql_INACTIVITY_6M_PRP(request):
    sql_storage = '/Scripts/Tyzdenne/'
    query = 'REP_INACTIVITY_6M_PRP.sql'
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    path = BASE_DIR + sql_storage

    filename = query
    filepath = path + query
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename
    # Return the response value
    return response

def sql_INACTIVITY_4M(request):
    sql_storage = '/Scripts/Tyzdenne/'
    query = 'REP_INACTIVITY_4M.sql'
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    path = BASE_DIR + sql_storage

    filename = query
    filepath = path + query
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename
    # Return the response value
    return response

def sql_Vyrubene_myto_kategorie(request):
    sql_storage = '/Scripts/Denne/'
    query = 'Vyrubene_myto_kategorie.sql'
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    path = BASE_DIR + sql_storage

    filename = query
    filepath = path + query
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename
    # Return the response value
    return response


def sql_Vyrubene_myto_krajiny(request):
    sql_storage = '/Scripts/Denne/'
    query = 'Vyrubene_myto_krajiny.sql'
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    path = BASE_DIR + sql_storage

    filename = query
    filepath = path + query
    # Open the file for reading content
    path = open(filepath, 'rb')
    # Set the mime type
    mime_type, _ = mimetypes.guess_type(filepath)
    # Set the return value of the HttpResponse
    response = HttpResponse(path, content_type=mime_type)
    # Set the HTTP header for sending to browser
    response['Content-Disposition'] = "attachment; filename=%s" % filename
    # Return the response value
    return response


from openpyxl import load_workbook
def report_Mark_stat():
    logger.info("------------ START job marketing statistiky " )  # Zaznamenajte správu do info.log
    EMS_DWH=EMS_DWHDB.objects.get(user=1)
    print(EMS_DWH)
    con_BILLDB_STDBY = cx_Oracle.connect(EMS_DWH.EMS_DWHDB_username+"/"+EMS_DWH.EMS_DWHDB_password+"@"+EMS_DWH.EMS_DWHDB_hostname+":"+EMS_DWH.EMS_DWHDB_port+"/"+EMS_DWH.EMS_DWHDB_servicename)
    print(con_BILLDB_STDBY)
    cursor = con_BILLDB_STDBY.cursor()
    
    # # 1 Drop Table if exists T_TMP_MKT_ROAD_USAGE_1
    cursor.execute('''declare n number(10);
                    begin
                    select count(*) into n from tab where tname='T_TMP_MKT_ROAD_USAGE_1';

                    if (n > 0) then 
                        execute immediate 
                        'DROP TABLE T_TMP_MKT_ROAD_USAGE_1';
                    end if;
                    end;''')
    now = datetime.now()
    time = now.strftime("%H:%M:%S")
    print("Table Deleted: T_TMP_MKT_ROAD_USAGE_1", time, " START")
    logger.info("Table Deleted: T_TMP_MKT_ROAD_USAGE_1")
    
    # 1 Create table T_TMP_MKT_ROAD_USAGE_1
    cursor.execute('''CREATE TABLE T_TMP_MKT_ROAD_USAGE_1 AS
                        SELECT
                        /*+ NO_INDEX(pass)*/
                        pass.VEHICLE_ID,
                        to_char(pass.OBU_LOCAL_TIME,'mm') AS MONTH,
                        s.ROAD_TYPE_CODE,
                        pass.OBU_NUMBER_OF_AXLES,
                        NVL(pass.DISCOUNT_RATE, 0) AS DISCOUNT_RATE,
                        SUM(DECODE(TOLL_EVENT_TYPE_CODE, 1, 1, 2, -1, 3, 1, 0)) AS Transaction_count,
                        SUM(pass.PRICE_AMOUNT) AS Transaction_amount,
                        SUM (i.UNITS_USED) AS Tolled_km,
                        SUM(NVL(pass.DISCOUNT_AMOUNT, 0)) AS Discount_amount
                        FROM BILLIEN_MAA.RATED_TOLL_EVENT pass
                        left join BILLIEN_MAA.RTE_ITEM_ALL i ON pass.RATED_TOLL_EVENT_ID = i.RATED_TOLL_EVENT_ID
                        JOIN BILLIEN_MAA.ROAD_SEGMENTS s ON pass.TOLL_SEGMENT_ID=s.TOLL_SEGMENT_ID
                        WHERE pass.OBU_TIMESTAMP >= CAST(FROM_TZ(CAST(ADD_MONTHS(TRUNC(CAST(FROM_TZ(CAST(SYSDATE AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET' AS DATE), 'MM'), -1/*pri ročnom -12*/) AS TIMESTAMP), 'CET') AT TIME ZONE 'GMT' AS DATE)
                        AND pass.OBU_TIMESTAMP < CAST(FROM_TZ(CAST(TRUNC(CAST(FROM_TZ(CAST(SYSDATE AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET' AS DATE), 'MM') AS TIMESTAMP), 'CET') AT TIME ZONE 'GMT' AS DATE)
                        AND pass.EXEMPT_FLAG=0
                        AND pass.TEST_CUSTOMER_FLAG=0
                        AND i.CHARGE_TYPE_CODE=1
                        GROUP BY pass.VEHICLE_ID, to_char(pass.OBU_LOCAL_TIME,'mm'), s.ROAD_TYPE_CODE, pass.OBU_NUMBER_OF_AXLES, NVL(pass.DISCOUNT_RATE, 0)''')
    now = datetime.now()
    time = now.strftime("%H:%M:%S")
    print("Table Created: T_TMP_MKT_ROAD_USAGE_1", time)
    logger.info("Table Created: T_TMP_MKT_ROAD_USAGE_1")
    
    # 2 Drop Table if exists T_TMP_MKT_ROAD_USAGE_2
    cursor.execute('''declare n number(10);
                        begin
                        select count(*) into n from tab where tname='T_TMP_MKT_ROAD_USAGE_2';

                        if (n > 0) then 
                            execute immediate 
                            'DROP TABLE T_TMP_MKT_ROAD_USAGE_2';
                        end if;
                        end;''')
    print("Table Deleted: T_TMP_MKT_ROAD_USAGE_2", time)
    logger.info("Table Deleted: T_TMP_MKT_ROAD_USAGE_2")


    # 2 Create table T_TMP_MKT_ROAD_USAGE_2
    cursor.execute('''CREATE TABLE T_TMP_MKT_ROAD_USAGE_2 AS
                        SELECT
                        /*+ NO_INDEX(pass)*/
                        pass.VEHICLE_ID,
                        to_char(pass.OBU_LOCAL_TIME,'yyyy.mm.dd') AS PassageDate,
                        SUM(DECODE(TOLL_EVENT_TYPE_CODE, 1, 1, 2, -1, 3, 1, 0)) AS Transaction_count,
                        SUM(pass.PRICE_AMOUNT) AS Transaction_amount,
                        SUM (i.UNITS_USED) AS Tolled_km
                        FROM BILLIEN_MAA.RATED_TOLL_EVENT pass
                        left join BILLIEN_MAA.RTE_ITEM_ALL i ON pass.RATED_TOLL_EVENT_ID = i.RATED_TOLL_EVENT_ID
                        JOIN BILLIEN_MAA.ROAD_SEGMENTS s ON pass.TOLL_SEGMENT_ID=s.TOLL_SEGMENT_ID
                        WHERE pass.OBU_TIMESTAMP >= CAST(FROM_TZ(CAST(ADD_MONTHS(TRUNC(CAST(FROM_TZ(CAST(SYSDATE AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET' AS DATE), 'MM'), -1/*pri ročnom -12*/) AS TIMESTAMP), 'CET') AT TIME ZONE 'GMT' AS DATE)
                        AND pass.OBU_TIMESTAMP < CAST(FROM_TZ(CAST(TRUNC(CAST(FROM_TZ(CAST(SYSDATE AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET' AS DATE), 'MM') AS TIMESTAMP), 'CET') AT TIME ZONE 'GMT' AS DATE)
                        AND pass.EXEMPT_FLAG=0
                        AND pass.TEST_CUSTOMER_FLAG=0
                        AND i.CHARGE_TYPE_CODE=1
                        GROUP BY pass.VEHICLE_ID,to_char(pass.OBU_LOCAL_TIME,'yyyy.mm.dd')''')
    now = datetime.now()
    time = now.strftime("%H:%M:%S")
    print("Table Created: T_TMP_MKT_ROAD_USAGE_2", time)
    logger.info("Table Created: T_TMP_MKT_ROAD_USAGE_2")

    
    # 3 Drop Table if exists T_TMP_MKT_ROAD_USAGE_3
    cursor.execute('''declare n number(10);
                        begin
                        select count(*) into n from tab where tname='T_TMP_MKT_ROAD_USAGE_3';

                        if (n > 0) then 
                            execute immediate 
                            'DROP TABLE T_TMP_MKT_ROAD_USAGE_3';
                        end if;
                        end;''')
    print("Table Deleted: T_TMP_MKT_ROAD_USAGE_3", time)
    logger.info("Table Deleted: T_TMP_MKT_ROAD_USAGE_3")

    # 3 Create table T_TMP_MKT_ROAD_USAGE_3
    cursor.execute('''CREATE TABLE T_TMP_MKT_ROAD_USAGE_3 AS
                        SELECT
                        /*+ NO_INDEX(pass)*/
                        pass.VEHICLE_ID,
                        TRUNC(pass.OBU_LOCAL_TIME) AS Passage_date,
                        pass.TOLL_SEGMENT_ID AS Segment_ID,
                        SUM(DECODE(TOLL_EVENT_TYPE_CODE, 1, 1, 2, -1, 3, 1, 0)) AS Transaction_count,
                        SUM(pass.PRICE_AMOUNT) AS Transaction_amount
                        FROM BILLIEN_MAA.RATED_TOLL_EVENT pass
                        WHERE pass.OBU_TIMESTAMP >= CAST(FROM_TZ(CAST(ADD_MONTHS(TRUNC(CAST(FROM_TZ(CAST(SYSDATE AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET' AS DATE), 'MM'), -1/*pri ročnom -12*/) AS TIMESTAMP), 'CET') AT TIME ZONE 'GMT' AS DATE)
                        AND pass.OBU_TIMESTAMP < CAST(FROM_TZ(CAST(TRUNC(CAST(FROM_TZ(CAST(SYSDATE AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET' AS DATE), 'MM') AS TIMESTAMP), 'CET') AT TIME ZONE 'GMT' AS DATE)
                        AND pass.EXEMPT_FLAG=0
                        AND pass.TEST_CUSTOMER_FLAG=0
                        GROUP BY pass.VEHICLE_ID, TRUNC(pass.OBU_LOCAL_TIME), pass.TOLL_SEGMENT_ID''')
    now = datetime.now()
    time = now.strftime("%H:%M:%S")
    print("Table Created: T_TMP_MKT_ROAD_USAGE_3", time," END created tables")
    logger.info("Table Created: T_TMP_MKT_ROAD_USAGE_3 and END created tables")
    # # ---------- end created table --------------------------------------------------------
    # ---------- start created reports --------------------------------------------------------

    # settings date
    today = datetime.now() + relativedelta(months=-1)
    my = today.strftime("%m%Y")

    # Save Script
    # filename=r"C:\NEW\Scripts\Marketing_stat\1_BILLDB_a.sql"
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    query = '1_BILLDB_a.sql'
    sql_storage = '/Scripts/Marketing_stat/'
    name_report='/Reports/Marketing_stat/Marketingove_statistiky_'+my+'.xlsx'
    path = BASE_DIR + sql_storage
    filename = path + query

    # odmazanie suborov
    directory = BASE_DIR +'/Reports/Marketing_stat/'
    for file_name in os.listdir(directory):
        # construct full file path
        file = os.path.join(directory, file_name)
        if os.path.isfile(file):
            os.remove(file)
    
    # open script filename
    f = open(filename)
    full_sql = f.read()
    
    billdb=BILLDB.objects.get(user=1)
    con_BILLDB_STDBY = cx_Oracle.connect(billdb.BILLDB_username+"/"+billdb.BILLDB_password+"@"+billdb.BILLDB_hostname+":"+billdb.BILLDB_port+"/"+billdb.BILLDB_servicename)
    cur_billdb = con_BILLDB_STDBY.cursor()
    cur_billdb.execute(full_sql)

    columns = [desc[0] for desc in cur_billdb.description]
    data = cur_billdb.fetchall()
    df = pd.DataFrame(list(data), columns=columns)
       

    # Create a new excel
  
    df.to_excel(BASE_DIR+name_report, sheet_name='01a', index=False)
    now = datetime.now()
    time = now.strftime("%H:%M:%S")
    print("create report: 01a_"+my+".xlsx "+time)
    logger.info("create report: 01a_"+my+".xlsx ")

    
    # GENERATE REPORT 01b
    query = '1_BILLDB_b.sql'
    filename = path + query
    f = open(filename)
    full_sql = f.read()
    cur_billdb.execute(full_sql)
    columns = [desc[0] for desc in cur_billdb.description]
    data = cur_billdb.fetchall()
    df = pd.DataFrame(list(data), columns=columns)
    
    with pd.ExcelWriter(BASE_DIR+name_report, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, sheet_name='01b', index=False)
        
    # GENERATE REPORT 02a
    con_EMSDWH_STDBY = cx_Oracle.connect(EMS_DWH.EMS_DWHDB_username+"/"+EMS_DWH.EMS_DWHDB_password+"@"+EMS_DWH.EMS_DWHDB_hostname+":"+EMS_DWH.EMS_DWHDB_port+"/"+EMS_DWH.EMS_DWHDB_servicename)
    cur_EMSDWH = con_EMSDWH_STDBY.cursor()
    cursor = con_EMSDWH_STDBY.cursor()
    query = '2_DWH_a.sql'
    filename = path + query
    f = open(filename)
    full_sql = f.read()
    cur_EMSDWH.execute(full_sql)
    columns = [desc[0] for desc in cur_EMSDWH.description]
    data = cur_EMSDWH.fetchall()
    df = pd.DataFrame(list(data), columns=columns)
    
    with pd.ExcelWriter(BASE_DIR+name_report, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, sheet_name='02a', index=False)

    # GENERATE REPORT 02b
    con_EMSDWH_STDBY = cx_Oracle.connect(EMS_DWH.EMS_DWHDB_username+"/"+EMS_DWH.EMS_DWHDB_password+"@"+EMS_DWH.EMS_DWHDB_hostname+":"+EMS_DWH.EMS_DWHDB_port+"/"+EMS_DWH.EMS_DWHDB_servicename)
    cur_EMSDWH = con_EMSDWH_STDBY.cursor()
    cursor = con_EMSDWH_STDBY.cursor()
    query = '2_DWH_b.sql'
    filename = path + query
    f = open(filename)
    full_sql = f.read()
    cur_EMSDWH.execute(full_sql)
    columns = [desc[0] for desc in cur_EMSDWH.description]
    data = cur_EMSDWH.fetchall()
    df = pd.DataFrame(list(data), columns=columns)
    
    with pd.ExcelWriter(BASE_DIR+name_report, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, sheet_name='02b', index=False)
        
    # GENERATE REPORT 02c
    con_EMSDWH_STDBY = cx_Oracle.connect(EMS_DWH.EMS_DWHDB_username+"/"+EMS_DWH.EMS_DWHDB_password+"@"+EMS_DWH.EMS_DWHDB_hostname+":"+EMS_DWH.EMS_DWHDB_port+"/"+EMS_DWH.EMS_DWHDB_servicename)
    cur_EMSDWH = con_EMSDWH_STDBY.cursor()
    cursor = con_EMSDWH_STDBY.cursor()
    query = '2_DWH_c.sql'
    filename = path + query
    f = open(filename)
    full_sql = f.read()
    cur_EMSDWH.execute(full_sql)
    columns = [desc[0] for desc in cur_EMSDWH.description]
    data = cur_EMSDWH.fetchall()
    df = pd.DataFrame(list(data), columns=columns)
    
    with pd.ExcelWriter(BASE_DIR+name_report, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, sheet_name='02c', index=False)
        
    # GENERATE REPORT 10a
    con_EMSDWH_STDBY = cx_Oracle.connect(EMS_DWH.EMS_DWHDB_username+"/"+EMS_DWH.EMS_DWHDB_password+"@"+EMS_DWH.EMS_DWHDB_hostname+":"+EMS_DWH.EMS_DWHDB_port+"/"+EMS_DWH.EMS_DWHDB_servicename)
    cur_EMSDWH = con_EMSDWH_STDBY.cursor()
    cursor = con_EMSDWH_STDBY.cursor()
    query = '10_DWH_a.sql'
    filename = path + query
    f = open(filename)
    full_sql = f.read()
    cur_EMSDWH.execute(full_sql)
    columns = [desc[0] for desc in cur_EMSDWH.description]
    data = cur_EMSDWH.fetchall()
    df = pd.DataFrame(list(data), columns=columns)
    
    with pd.ExcelWriter(BASE_DIR+name_report, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, sheet_name='10a', index=False)
        
    # GENERATE REPORT 10b
    con_EMSDWH_STDBY = cx_Oracle.connect(EMS_DWH.EMS_DWHDB_username+"/"+EMS_DWH.EMS_DWHDB_password+"@"+EMS_DWH.EMS_DWHDB_hostname+":"+EMS_DWH.EMS_DWHDB_port+"/"+EMS_DWH.EMS_DWHDB_servicename)
    cur_EMSDWH = con_EMSDWH_STDBY.cursor()
    cursor = con_EMSDWH_STDBY.cursor()
    query = '10_DWH_b.sql'
    filename = path + query
    f = open(filename)
    full_sql = f.read()
    cur_EMSDWH.execute(full_sql)
    columns = [desc[0] for desc in cur_EMSDWH.description]
    data = cur_EMSDWH.fetchall()
    df = pd.DataFrame(list(data), columns=columns)
    
    with pd.ExcelWriter(BASE_DIR+name_report, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, sheet_name='10b', index=False)
        
    # GENERATE REPORT 10c
    con_EMSDWH_STDBY = cx_Oracle.connect(EMS_DWH.EMS_DWHDB_username+"/"+EMS_DWH.EMS_DWHDB_password+"@"+EMS_DWH.EMS_DWHDB_hostname+":"+EMS_DWH.EMS_DWHDB_port+"/"+EMS_DWH.EMS_DWHDB_servicename)
    cur_EMSDWH = con_EMSDWH_STDBY.cursor()
    cursor = con_EMSDWH_STDBY.cursor()
    query = '10_DWH_c.sql'
    filename = path + query
    f = open(filename)
    full_sql = f.read()
    cur_EMSDWH.execute(full_sql)
    columns = [desc[0] for desc in cur_EMSDWH.description]
    data = cur_EMSDWH.fetchall()
    df = pd.DataFrame(list(data), columns=columns)
    
    with pd.ExcelWriter(BASE_DIR+name_report, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, sheet_name='10c', index=False)
        
    # GENERATE REPORT 10d
    con_EMSDWH_STDBY = cx_Oracle.connect(EMS_DWH.EMS_DWHDB_username+"/"+EMS_DWH.EMS_DWHDB_password+"@"+EMS_DWH.EMS_DWHDB_hostname+":"+EMS_DWH.EMS_DWHDB_port+"/"+EMS_DWH.EMS_DWHDB_servicename)
    cur_EMSDWH = con_EMSDWH_STDBY.cursor()
    cursor = con_EMSDWH_STDBY.cursor()
    query = '10_DWH_d.sql'
    filename = path + query
    f = open(filename)
    full_sql = f.read()
    cur_EMSDWH.execute(full_sql)
    columns = [desc[0] for desc in cur_EMSDWH.description]
    data = cur_EMSDWH.fetchall()
    df = pd.DataFrame(list(data), columns=columns)
    
    with pd.ExcelWriter(BASE_DIR+name_report, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, sheet_name='10d', index=False)
        
    # GENERATE REPORT 10d
    con_EMSDWH_STDBY = cx_Oracle.connect(EMS_DWH.EMS_DWHDB_username+"/"+EMS_DWH.EMS_DWHDB_password+"@"+EMS_DWH.EMS_DWHDB_hostname+":"+EMS_DWH.EMS_DWHDB_port+"/"+EMS_DWH.EMS_DWHDB_servicename)
    cur_EMSDWH = con_EMSDWH_STDBY.cursor()
    cursor = con_EMSDWH_STDBY.cursor()
    query = '11_DWH_a.sql'
    filename = path + query
    f = open(filename)
    full_sql = f.read()
    cur_EMSDWH.execute(full_sql)
    columns = [desc[0] for desc in cur_EMSDWH.description]
    data = cur_EMSDWH.fetchall()
    df = pd.DataFrame(list(data), columns=columns)
    
    with pd.ExcelWriter(BASE_DIR+name_report, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, sheet_name='11a', index=False)
    logger.info("create last tab in excele 11a")
    
    # zazipovanie
    # Define the source and destination file paths
    source_file = BASE_DIR+'\Reports\Marketing_stat\Marketingove_statistiky_'+my+'.xlsx'
    destination_file = BASE_DIR+'\Reports\Marketing_stat\Marketingove_statistiky_'+my+'.zip'

    # Create a zip file from the source file
    with zipfile.ZipFile(destination_file, 'w') as zipf:
        zipf.write(source_file, os.path.basename(source_file))

    # If you want to unzip the file after zipping it, you can use the following code:
    # Unzip the file
    with zipfile.ZipFile(destination_file, 'r') as zipf:
        zipf.extractall(BASE_DIR+'\Reports\Marketing_stat')
        
    def load_orders(file_name):
        with open(file_name, 'r') as file:
            json_inputs = json.load(file)
        # email_addresses = [email for json_input in json_inputs if json_input['model'] == 'marketingove_statistiky' for email in json_input['fields']['emails']]
        email_addresses = [email for json_input in json_inputs if json_input['model'] == 'marketingove_statistiky' for email in json_input['fields'].get('emails', [])]
        return email_addresses
    def load_head(file_name):
        with open(file_name, 'r') as file:
            json_inputs = json.load(file)
        head = [json_input['fields']['head'] for json_input in json_inputs if json_input['model'] == 'marketingove_statistiky'][0] if json_inputs else None
        return head
    def load_body(file_name):
        with open(file_name, 'r') as file:
            json_inputs = json.load(file)
        body = [json_input['fields']['body'].replace('\\n', '\n') for json_input in json_inputs if json_input['model'] == 'marketingove_statistiky'][0] if json_inputs else None
        return body
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    file_name = BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))+'/work/static/json/jobs.json'
    # file_name = 'path/to/orders.json'
    email_addresses = load_orders(file_name)
    head = load_head(file_name)+" - " +my
    body = load_body(file_name)
    print(email_addresses)
    sender_email = "reports@example.com"
    receiver_email=email_addresses
    # Create an EmailMessage object and set headers
    message = EmailMessage(
        subject=head,
        body=body,
        from_email=sender_email,
        to=receiver_email,  # Pass the list of email addresses directly to the 'to' parameter
    )
    
    # Add attachment to email
    message.attach_file(destination_file)

    # Send email
    message.send()

    print('report poslany')
    logger.info("------------- END Marketing Statistiky report poslany")

# ::::::::::::::::::: REPORTS END ::::::::::::::::::::::::::::::::::::::::::::
@login_required
def EMS_bd_ack(request):
    import cx_Oracle
    billdb=BILLDB.objects.get(user=request.user)
    con_BILLDB_STDBY = cx_Oracle.connect(billdb.BILLDB_username+"/"+billdb.BILLDB_password+"@"+billdb.BILLDB_hostname+":"+billdb.BILLDB_port+"/"+billdb.BILLDB_servicename)
    cursor = con_BILLDB_STDBY.cursor()

    sql_ack_None = """
                SELECT *
                FROM
                (SELECT bd.ede_log_id   AS bd_ede_log_id,
                    bd.integration_log_id AS bd_integration_log_id,
                    bd.ede_log_status_code
                    ||'-'
                    || bds.ede_log_status_name AS bd_ede_log_status,
                    bd.eets_provider_id
                    ||'-'
                    || p.PROVIDER_NUMBER
                    ||'-'
                    || p.provider_abbreviation    AS EETS_PROVIDER,
                    TO_CHAR( bd.apdu_identifier ) AS apdu_identifier,
                    --bd.created_on                 AS bd_created_on,
                    CAST(FROM_TZ(CAST(bd.created_on AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET' AS DATE) as bd_created_on,
                    --bd.exported_on                AS bd_exported_on,
                    CAST(FROM_TZ(CAST(bd.exported_on AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET' AS DATE) as bd_exported_on,
                    --ack.created_on                AS ACK_CREATED_ON,
                    CAST(FROM_TZ(CAST(ack.created_on AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET' AS DATE) as ACK_CREATED_ON,
                    24* (to_date( ack.created_on) - to_date(bd.exported_on )) diff_hours
                    
                FROM BILLIEN_ede.ede_log bd
                JOIN BILLIEN_ede.ede_log_status_l bds
                ON bd.ede_log_status_code = bds.ede_log_status_code
                AND bds.language_code     = 'CS'
                JOIN BILLIEN_ECM.EETS_PROVIDER p
                ON bd.eets_provider_id = p.EETS_PROVIDER_ID
                LEFT JOIN BILLIEN_ede.ede_log ack
                ON bd.ede_log_id               = ack.referred_ede_log_id
                AND ack.ede_message_type_code  = 1
                AND ack.ede_log_status_code    = 1
                WHERE p.provider_abbreviation IN ( 'EW', 'ITIS' )
                AND bd.ede_message_type_code   = 4
                --AND bd.exported_on is not NULL
                --AND bd.created_on              > sysdate - 26 / 24
                AND bd.created_on >= trunc(sysdate-7, 'DD') 
                )
                --WHERE NVL(ack_created_on, bd_created_on+1) - bd_created_on > 2/24
                WHERE ACK_CREATED_ON is NULL
               
                ORDER BY bd_created_on desc
              """
    cursor.execute(sql_ack_None)
    ack_None = dictfetchall(cursor)
    x = None
    y = None
    
    if 'x' in request.GET and request.GET['y'] =='':        
        x = request.GET['x']
        y = request.GET['y']
        sql_x="""
            SELECT *
                FROM
                (SELECT bd.ede_log_id   AS bd_ede_log_id,
                    bd.integration_log_id AS bd_integration_log_id,
                    bd.ede_log_status_code
                    ||'-'
                    || bds.ede_log_status_name AS bd_ede_log_status,
                    bd.eets_provider_id
                    ||'-'
                    || p.PROVIDER_NUMBER
                    ||'-'
                    || p.provider_abbreviation    AS EETS_PROVIDER,
                    TO_CHAR( bd.apdu_identifier ) AS apdu_identifier,
                    --bd.created_on                 AS bd_created_on,
                    CAST(FROM_TZ(CAST(bd.created_on AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET' AS DATE) as bd_created_on,
                    --bd.exported_on                AS bd_exported_on,
                    CAST(FROM_TZ(CAST(bd.exported_on AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET' AS DATE) as bd_exported_on,
                    --ack.created_on                AS ACK_CREATED_ON,
                    CAST(FROM_TZ(CAST(ack.created_on AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET' AS DATE) as ACK_CREATED_ON,
                    --24* (to_date( ack.created_on) - to_date(bd.exported_on )) diff_hours
                    ROUND(24* (to_date( ack.created_on) - to_date(bd.created_on)),3) diff_hours,
                   ROUND((to_date(ack.created_on) - to_date(bd.created_on))*24*60,1) DIFF_MINUTES,
                   ROUND((to_date(ack.created_on) - to_date(bd.exported_on))*24*60*60,3) DIFF_SECONDS

                FROM BILLIEN_ede.ede_log bd
                JOIN BILLIEN_ede.ede_log_status_l bds
                ON bd.ede_log_status_code = bds.ede_log_status_code
                AND bds.language_code     = 'CS'
                JOIN BILLIEN_ECM.EETS_PROVIDER p
                ON bd.eets_provider_id = p.EETS_PROVIDER_ID
                LEFT JOIN BILLIEN_ede.ede_log ack
                ON bd.ede_log_id               = ack.referred_ede_log_id
                AND ack.ede_message_type_code  = 1
                AND ack.ede_log_status_code    = 1
                WHERE p.provider_abbreviation IN ( 'EW', 'ITIS' )
                AND bd.ede_message_type_code   = 4
                --AND bd.exported_on is not NULL
                --AND bd.created_on              > sysdate - 26 / 24
                --AND bd.created_on              > sysdate - 3
                )
                WHERE BD_CREATED_ON > to_date(:datex, 'DD-MM-YYYY HH24:MI:SS')
                ORDER BY bd_created_on desc
        """
 
        with con_BILLDB_STDBY.cursor() as cursor:
            cursor.execute(sql_x, datex=x)
            ack = dictfetchall(cursor)
    elif 'y' in request.GET and request.GET['x'] =='':
        x =None
        y = request.GET['y']
        sql_y="""
            SELECT *
                FROM
                (SELECT bd.ede_log_id   AS bd_ede_log_id,
                    bd.integration_log_id AS bd_integration_log_id,
                    bd.ede_log_status_code
                    ||'-'
                    || bds.ede_log_status_name AS bd_ede_log_status,
                    bd.eets_provider_id
                    ||'-'
                    || p.PROVIDER_NUMBER
                    ||'-'
                    || p.provider_abbreviation    AS EETS_PROVIDER,
                    TO_CHAR( bd.apdu_identifier ) AS apdu_identifier,
                    --bd.created_on                 AS bd_created_on,
                    CAST(FROM_TZ(CAST(bd.created_on AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET' AS DATE) as bd_created_on,
                    --bd.exported_on                AS bd_exported_on,
                    CAST(FROM_TZ(CAST(bd.exported_on AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET' AS DATE) as bd_exported_on,
                    --ack.created_on                AS ACK_CREATED_ON,
                    CAST(FROM_TZ(CAST(ack.created_on AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET' AS DATE) as ACK_CREATED_ON,
                    --24* (to_date( ack.created_on) - to_date(bd.exported_on )) diff_hours
                    ROUND(24* (to_date( ack.created_on) - to_date(bd.created_on)),3) diff_hours,
                   ROUND((to_date(ack.created_on) - to_date(bd.created_on))*24*60,1) DIFF_MINUTES,
                   ROUND((to_date(ack.created_on) - to_date(bd.created_on))*24*60*60,3) DIFF_SECONDS

                FROM billien_ede.ede_log bd
                JOIN billien_ede.ede_log_status_l bds
                ON bd.ede_log_status_code = bds.ede_log_status_code
                AND bds.language_code     = 'CS'
                JOIN billien_ECM.EETS_PROVIDER p
                ON bd.eets_provider_id = p.EETS_PROVIDER_ID
                LEFT JOIN billien_ede.ede_log ack
                ON bd.ede_log_id               = ack.referred_ede_log_id
                AND ack.ede_message_type_code  = 1
                AND ack.ede_log_status_code    = 1
                WHERE p.provider_abbreviation IN ( 'EW', 'ITIS' )
                AND bd.ede_message_type_code   = 4
                --AND bd.exported_on is not NULL
                --AND bd.created_on              > sysdate - 26 / 24
                --AND bd.created_on              > sysdate - 3
                )
                WHERE BD_CREATED_ON < to_date(:datey, 'DD-MM-YYYY HH24:MI:SS')
                ORDER BY bd_created_on desc
        """
        with con_BILLDB_STDBY.cursor() as cursor:
            cursor.execute(sql_y, datey=y)
            ack = dictfetchall(cursor)
    elif 'x' in request.GET and 'y' in request.GET:
        x=request.GET['x']
        y=request.GET['y']
        sql_xy="""
            SELECT *
                FROM
                (SELECT bd.ede_log_id   AS bd_ede_log_id,
                    bd.integration_log_id AS bd_integration_log_id,
                    bd.ede_log_status_code
                    ||'-'
                    || bds.ede_log_status_name AS bd_ede_log_status,
                    bd.eets_provider_id
                    ||'-'
                    || p.PROVIDER_NUMBER
                    ||'-'
                    || p.provider_abbreviation    AS EETS_PROVIDER,
                    TO_CHAR( bd.apdu_identifier ) AS apdu_identifier,
                    --bd.created_on                 AS bd_created_on,
                    CAST(FROM_TZ(CAST(bd.created_on AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET' AS DATE) as bd_created_on,
                    --bd.exported_on                AS bd_exported_on,
                    CAST(FROM_TZ(CAST(bd.exported_on AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET' AS DATE) as bd_exported_on,
                    --ack.created_on                AS ACK_CREATED_ON,
                    CAST(FROM_TZ(CAST(ack.created_on AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET' AS DATE) as ACK_CREATED_ON,
                    --24* (to_date( ack.created_on) - to_date(bd.exported_on )) diff_hours
                    ROUND(24* (to_date( ack.created_on) - to_date(bd.created_on)),3) diff_hours,
                   ROUND((to_date(ack.created_on) - to_date(bd.created_on))*24*60,1) DIFF_MINUTES,
                   ROUND((to_date(ack.created_on) - to_date(bd.created_on))*24*60*60,3) DIFF_SECONDS

                FROM billien_ede.ede_log bd
                JOIN billien_ede.ede_log_status_l bds
                ON bd.ede_log_status_code = bds.ede_log_status_code
                AND bds.language_code     = 'CS'
                JOIN billien_ECM.EETS_PROVIDER p
                ON bd.eets_provider_id = p.EETS_PROVIDER_ID
                LEFT JOIN billien_ede.ede_log ack
                ON bd.ede_log_id               = ack.referred_ede_log_id
                AND ack.ede_message_type_code  = 1
                AND ack.ede_log_status_code    = 1
                WHERE p.provider_abbreviation IN ( 'EW', 'ITIS' )
                AND bd.ede_message_type_code   = 4
                --AND bd.exported_on is not NULL
                --AND bd.created_on              > sysdate - 26 / 24
                --AND bd.created_on              > sysdate - 3
                )
                WHERE BD_CREATED_ON > to_date(:datex, 'DD-MM-YYYY HH24:MI:SS')
                        AND BD_CREATED_ON < to_date(:datey, 'DD-MM-YYYY HH24:MI:SS')
                ORDER BY bd_created_on desc
        """
        with con_BILLDB_STDBY.cursor() as cursor:
            cursor.execute(sql_xy, datex=x, datey=y)
            ack = dictfetchall(cursor)
    else:
        sql_ack = """
                SELECT *
                FROM
                (SELECT bd.ede_log_id   AS bd_ede_log_id,
                    bd.integration_log_id AS bd_integration_log_id,
                    bd.ede_log_status_code
                    ||'-'
                    || bds.ede_log_status_name AS bd_ede_log_status,
                    bd.eets_provider_id
                    ||'-'
                    || p.PROVIDER_NUMBER
                    ||'-'
                    || p.provider_abbreviation    AS EETS_PROVIDER,
                    TO_CHAR( bd.apdu_identifier ) AS apdu_identifier,
                    --bd.created_on                 AS bd_created_on,
                    CAST(FROM_TZ(CAST(bd.created_on AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET' AS DATE) as bd_created_on,
                    --bd.exported_on                AS bd_exported_on,
                    CAST(FROM_TZ(CAST(bd.exported_on AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET' AS DATE) as bd_exported_on,
                    --ack.created_on                AS ACK_CREATED_ON,
                    CAST(FROM_TZ(CAST(ack.created_on AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET' AS DATE) as ACK_CREATED_ON,
                    --24* (to_date( ack.created_on) - to_date(bd.exported_on )) diff_hours
                    --ROUND(24* (to_date( ack.created_on) - to_date(bd.created_on)),3) diff_hours,
                   --ROUND((to_date(ack.created_on) - to_date(bd.created_on))*24*60,1) MINUTES,
                    ROUND((NVL(ack.created_on,bd.created_on)  -bd.created_on)*24*60, 1 ) || ' min' as DIFF_MINUTES
                   --ROUND((to_date(ack.created_on) - to_date(bd.created_on))*24*60*60,3) DIFF_SECONDS

                FROM billien_ede.ede_log bd
                JOIN billien_ede.ede_log_status_l bds
                ON bd.ede_log_status_code = bds.ede_log_status_code
                AND bds.language_code     = 'CS'
                JOIN billien_ECM.EETS_PROVIDER p
                ON bd.eets_provider_id = p.EETS_PROVIDER_ID
                LEFT JOIN billien_ede.ede_log ack
                ON bd.ede_log_id               = ack.referred_ede_log_id
                AND ack.ede_message_type_code  = 1
                AND ack.ede_log_status_code    = 1
                WHERE p.provider_abbreviation IN ( 'EW', 'ITIS' )
                AND bd.ede_message_type_code   = 4
                --AND bd.exported_on is not NULL
                --AND bd.created_on              > sysdate - 26 / 24
                AND bd.created_on              > sysdate - 3
                )
                --WHERE NVL(ack_created_on, bd_created_on+1) - bd_created_on > 2/24
                ORDER BY bd_created_on desc
              """

        with con_BILLDB_STDBY.cursor() as cursor:
            cursor.execute(sql_ack)
            ack = dictfetchall(cursor)
 
    context = {'ack':ack, 'ack_None':ack_None}
    return render(request, 'ems/ems_bd_ack.html', context)

@login_required
def EMS_bad_trans(request):
    # con_CISDB_STDBY = cx_Oracle.connect("DB_USER/DB_PASSWORD@db-host-b.internal.example.com:1521/cissrv_rdo")
    import cx_Oracle
    billdb=BILLDB.objects.get(user=request.user)
    con_BILLDB_STDBY = cx_Oracle.connect(billdb.BILLDB_username+"/"+billdb.BILLDB_password+"@"+billdb.BILLDB_hostname+":"+billdb.BILLDB_port+"/"+billdb.BILLDB_servicename)
    cursor = con_BILLDB_STDBY.cursor()

    sql_REJECTED_TRANSACTION = """Select *
                                from BILLIEN_FA.REJECTED_TRANSACTION
                                where DELETED_ON is NULL
                                ORDER BY TRANSACTION_DATA_ID
                                """
    with con_BILLDB_STDBY.cursor() as cursor:
        cursor.execute(sql_REJECTED_TRANSACTION)
            #rows = cursor.fetchall()        
            #rows = namedtuplefetchall(cursor)
        REJECTED_TRANSACTION = dictfetchall(cursor)

    if os.path.exists("download_trans/Top_5_best_criteria.txt"): 
        os.remove("download_trans/Top_5_best_criteria.txt")
        file_name = r"download_trans/Top_5_best_criteria.txt"
        with open(file_name, 'a') as f:
            f.write(str('Top 5 best-fit criteria'))
        # print('delete file')
    # else:
        # print("The file does not exist")
    # con_CISDB_STDBY = cx_Oracle.connect("DB_USER/DB_PASSWORD@db-host-b.internal.example.com:1521/cissrv_rdo")
    cursor = con_BILLDB_STDBY.cursor()

    # cursor = connections['cis_db'].cursor()

    cursor.callproc("dbms_output.enable")
    
    sqlCode = """
                declare
                ----------------------------------------------------------------------------------------------
                v_transaction_data_id       billien_fa.table_of_number_18_t := billien_fa.table_of_number_18_t(:cd);
                tid_to_display_count        number := 5;
                consider_deleted_definition number := 0;
                ----------------------------------------------------------------------------------------------
                type t is table of varchar2(4000);
                criteria_ok                t := t();
                criteria_bad               t := t();
                transaction_item_def_id    billien_fa.table_of_number_18_t := billien_fa.table_of_number_18_t();
                transaction_item_def_short billien_fa.table_of_varchar2_50_t := billien_fa.table_of_varchar2_50_t();
                oktab                      billien_fa.table_of_number_18_t := billien_fa.table_of_number_18_t();
                badtab                     billien_fa.table_of_number_18_t := billien_fa.table_of_number_18_t();
                val                        number;
                ok                         number;
                bad                        number;
                begin
                dbms_output.put_line('Top ' || tid_to_display_count ||
                                    ' best-fit criteria');
                dbms_output.new_line;
                -- transaction data loop
                for i in (select td.transaction_data_id, ft.accounting_datetime
                            from billien_fa.transaction_data td
                            left join billien_fa.financial_transaction ft
                                on ft.transaction_data_id = td.transaction_data_id
                            where td.transaction_data_id in
                                (select column_value from table(v_transaction_data_id))) loop
                    dbms_output.put_line(lpad('*', 100, '*'));
                    dbms_output.put_line(rpad(' ', 18, ' ') ||
                                        'check transaction_data_id ' ||
                                        i.transaction_data_id ||
                                        ', accounting datetime: ' ||
                                        to_char(billien_co.to_local_datetime(i.accounting_datetime)
                                                ,'dd.mm.yyyy'));
                    dbms_output.put_line(lpad('*', 100, '*'));
                    -- clear variables
                    transaction_item_def_id.delete;
                    criteria_ok.delete;
                    criteria_bad.delete;
                    transaction_item_def_short.delete;
                    oktab.delete;
                    badtab.delete;
                    -- transaction item def
                    for j in (select transaction_item_def_id
                                    ,transaction_item_def_short
                                    ,rownum ord
                                from billien_fa.transaction_item_def td
                                join billien_fa.transaction_def d
                            using (transaction_def_id)
                            where exists
                            (select 1.0
                                        from billien_fa.selection_criteria sc
                                    where sc.transaction_item_def_id =
                                            td.transaction_item_def_id)
                                and (deleted_on is null and consider_deleted_definition = 0 or
                                    consider_deleted_definition = 1)
                            
                            order by transaction_item_def_short) loop
                    -- initialize variables
                    transaction_item_def_id.extend(1);
                    transaction_item_def_id(transaction_item_def_id.count) := j.transaction_item_def_id;
                    transaction_item_def_short.extend(1);
                    transaction_item_def_short(transaction_item_def_short.count) := j.transaction_item_def_short;
                    criteria_ok.extend(1);
                    criteria_ok(criteria_ok.count) := '';
                    criteria_bad.extend(1);
                    criteria_bad(criteria_bad.count) := '';
                    ok := 0;
                    bad := 0;
                    -- criteria
                    for k in (select data_column_system_name
                                    ,'nvl(to_char(' || case upper(data_column_system_name)
                                        when 'OBU_OPERATION_TYPE_CODE' then
                                        'OBU_STOCK_OPERATION_CODE'
                                        else
                                        data_column_system_name
                                    end || '),''null'')' || ' in (' ||
                                    listagg('''' || nvl(to_char(criteria_value), 'null') || ''''
                                            ,', ') within group(order by criteria_value nulls last) || ')' criteria
                                    , data_column_system_name || case count(*)
                                        when 1 then
                                        ' = '
                                        else
                                        ' in ('
                                        end ||
                                        listagg(nvl(to_char(criteria_value), 'null'), ', ') within group(order by criteria_value nulls last) ||case
                                        count(*)
                                        when 1 then
                                        ''
                                        else
                                        ')'
                                    end disp_criteria
                                from billien_fa.selection_criteria sc
                                join billien_fa.data_column
                                using (data_column_code)
                                where sc.transaction_item_def_id =
                                    j.transaction_item_def_id
                                group by data_column_system_name
                                order by data_column_system_name) loop
                        begin
                        execute immediate 'select ' ||
                                            case upper(k.data_column_system_name)
                                            when 'OBU_OPERATION_TYPE_CODE' then
                                            'OBU_STOCK_OPERATION_CODE'
                                            else
                                            k.data_column_system_name
                                            end ||
                                            ' from billien_fa.transaction_data where transaction_data_id = ' ||
                                            i.transaction_data_id || ' and ' || k.criteria
                            into val;
                        criteria_ok(criteria_ok.count) := criteria_ok(criteria_ok.count) || case
                                                            when ok = 0 then
                                                            ''
                                                            else
                                                            chr(10)
                                                            end || '          ' || k.disp_criteria ||
                                                            ' --> ' || nvl(to_char(val), 'null');
                        ok := ok + 1;
                        exception
                        when no_data_found then
                            execute immediate 'select ' ||
                                            case upper(k.data_column_system_name)
                                                when 'OBU_OPERATION_TYPE_CODE' then
                                                'OBU_STOCK_OPERATION_CODE'
                                                else
                                                k.data_column_system_name
                                            end ||
                                            ' from billien_fa.transaction_data where transaction_data_id = ' ||
                                            i.transaction_data_id
                            into val;
                            criteria_bad(criteria_bad.count) := criteria_bad(criteria_bad.count) || case
                                                                when bad = 0 then
                                                                ''
                                                                else
                                                                chr(10)
                                                                end || '        x ' ||
                                                                k.disp_criteria || ' --> ' ||
                                                                nvl(to_char(val), 'null');
                            bad := bad + 1;
                        end;
                    end loop;
                    oktab.extend(1);
                    oktab(oktab.count) := ok;
                    badtab.extend(1);
                    badtab(badtab.count) := bad;
                    end loop;
                    -- display results
                    for i in (with b as
                                (select column_value bad, rownum rn from table(badtab)),
                                o as
                                (select column_value ok, rownum rn from table(oktab)),
                                tid as
                                (select column_value transaction_item_def_id, rownum rn
                                from table(transaction_item_def_id)),
                                tsh as
                                (select column_value transaction_item_def_short, rownum rn
                                from table(transaction_item_def_short))
                                select t.*, rownum ord
                                from (select ok
                                            ,bad
                                            ,rn
                                            ,transaction_item_def_short
                                            ,transaction_item_def_id
                                            ,td.deleted_on
                                            ,td.validity_start
                                            ,td.validity_end
                                        from b
                                        join o
                                        using (rn)
                                        join tid
                                        using (rn)
                                        join tsh
                                        using (rn)
                                        join billien_fa.transaction_item_def tid
                                        using (transaction_item_def_id, transaction_item_def_short)
                                        join billien_fa.transaction_def td
                                            on td.transaction_def_id = tid.transaction_def_id
                                        order by case
                                                    when ok + bad = 0 then
                                                    0
                                                    else
                                                    ok / (ok + bad)
                                                end desc
                                                ,ok + bad desc
                                                ,transaction_item_def_short
                                                ,deleted_on desc nulls first) t
                                where rownum <= tid_to_display_count) loop
                    dbms_output.put_line(case when i.ok + i.bad = 0 then '0%' else
                                        to_char(round(i.ok / (i.ok + i.bad) * 100)) || '%'
                                        end || ' ' || i.transaction_item_def_short ||
                                        ' (item id ' || i.transaction_item_def_id || '),' ||
                                        ' validity: ' ||
                                        to_char(billien_co.to_local_datetime(i.validity_start)
                                                ,'dd.mm.yyyy') || ' - ' ||
                                        nvl(to_char(billien_co.to_local_datetime(i.validity_end)
                                                    ,'dd.mm.yyyy')
                                            ,'...') || case when
                                        i.deleted_on is not null then
                                        ' deleted on: ' ||
                                        to_char(billien_co.to_local_datetime(i.deleted_on)
                                                ,'dd.mm.yyyy') else '' end);
                    if criteria_bad(i.rn) is not null then
                        dbms_output.put_line(criteria_bad(i.rn));
                    end if;
                    if criteria_ok(i.rn) is not null then
                        dbms_output.put_line(criteria_ok(i.rn));
                    end if;
                    end loop;
                end loop;
                end;
        """
    q = None
    q2 = None
    q3 = None
    if 'q' in request.GET:
        q=request.GET['q']
    if 'q2' in request.GET:
        q2=request.GET['q2']
    if 'q3' in request.GET:
        q3=request.GET['q3']
    # if 'q2' in request.GET:
    #     q=request.GET['q2']
    # if 'q3' in request.GET:
    #     q=request.GET['q3']
    
    data =[{'cd' : q}, {'cd': q2}, {'cd': q3}]
    # data = [{'cd' : 71194061}, {'cd': 64675259}]
    # cursor.execute(sqlCode, data)
    cursor.executemany(sqlCode, data)

    statusVar = cursor.var(cx_Oracle.NUMBER)
    lineVar = cursor.var(cx_Oracle.STRING)
    while True:
        cursor.callproc("dbms_output.get_line", (lineVar, statusVar))
        if statusVar.getvalue() != 0:
            break
        # l = str(statusVar.getvalue())
        m = lineVar.getvalue()
        # print (m)
        # print(type(m))

        file_name = r"download_trans/Top_5_best_criteria.txt"
        with open(file_name, 'a') as f:
            f.write(str(m)+ '\n')
        #     # f.write(str(l))
        #     f.writelines(str(m)
    folder_ = r"download_trans/Top_5_best_criteria.txt"

    p = open(folder_, 'r')
    file_content = p.read()
    p.close()

    # cursor.close()
    # connection.close()   
    
    context = { 'file_content':file_content, 'REJECTED_TRAN':REJECTED_TRANSACTION }
    # context = {  'REJECTED_TRAN':REJECTED_TRANSACTION }

    
    return render(request,'ems/transakce_ems.html', context)

import datetime
def my_job():
    print(f"{datetime.now()} - I love python\n")
    
import os
import json
from django import forms
from django.shortcuts import render, redirect
from django.http import HttpResponse
from .forms import DataForm

# Získajte logger
import logging
logger = logging.getLogger('my_app')  # Použite názov loggera, ktorý ste definovali

# Define the filename of the JSON file to load


import json

def marketing_stats_view(request):
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    file_name = os.path.join(BASE_DIR, 'work', 'static', 'json', 'jobs.json')

    with open(file_name, 'r') as file:
        json_inputs = json.load(file)

    # emails = [json_input['fields']['emails'] for json_input in json_inputs if json_input['model'] == 'marketingove_statistiky']
    # print(emails)
    emails = [email for json_input in json_inputs if json_input['model'] == 'marketingove_statistiky' for email in json_input['fields']['emails']]
    head = [json_input['fields']['head'] for json_input in json_inputs if json_input['model'] == 'marketingove_statistiky'][0] if json_inputs else None
    body = [json_input['fields']['body'] for json_input in json_inputs if json_input['model'] == 'marketingove_statistiky'][0] if json_inputs else None
    
    # print(body)
    rune = [json_input['fields']['run'] for json_input in json_inputs if json_input['model'] == 'marketingove_statistiky'][0] if json_inputs else {}

    if request.method == 'POST':
        if 'emails' in request.POST:
            emails = request.POST['emails'].strip().split(',')
        if 'rune' in request.POST:
            rune = request.POST['rune']
        if 'head' in request.POST:
            head = request.POST['head']
        if 'body' in request.POST:
            body = request.POST['body']
            # body = body.replace('\\n', '\n')
        json_inputs = [json_input for json_input in json_inputs if json_input['model'] != 'marketingove_statistiky']
        json_inputs.append({
            'model': 'marketingove_statistiky',
            'fields': {
                'emails': emails,
                'run': rune,
                'head': head,
                'body':body
            }
        })
        with open(file_name, 'w') as file:
            json.dump(json_inputs, file, indent=4)

        from scheduler import reload_job_if_markstat
        # Reload job pre EETS_Lego
        reload_job_if_markstat('marketingove_statistiky', 'work/static/json/jobs.json')

        return render(request, 'ems/marketing_stats.html', {'emails': emails, 'rune': rune, 'head':head, 'body':body})
    return render(request, 'ems/marketing_stats.html', {'emails': emails, 'rune': rune, 'head':head, 'body':body})

#:::::::::::::: Reports PDF ::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::
def generate_pdf(request):
    from reportlab.lib.pagesizes import letter

    from reportlab.lib import colors

    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

    from reportlab.lib.units import mm
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    image_path = os.path.join(BASE_DIR, 'work', 'static', 'img', 'myto.png')

    # Create a new PDF document
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'filename="mypdf.pdf"'

    # Create a canvas object
    c = canvas.Canvas(response, pagesize=letter)

    # Calculate the center position of the page
    page_width, page_height = letter
    image_width, image_height = 25*mm, 17*mm
    x = 50
    y = 720
    # Draw the PNG image at the center of the page
    c.drawImage(image_path, x, y, width=image_width, height=image_height)

    # Draw the centered text
    font_size = 14
    text = "Prerozdelenie mýta medzi vlastníkov ciest"
    x_text = page_width / 2
    y_text = page_height - 30 * mm
    c.setFont('Helvetica-Bold', font_size)
    c.drawCentredString(x_text, y_text, text)

    # Draw the date in the upper right corner
    from datetime import datetime, timedelta
    date_text = f"Dátum vygenerovania: {datetime.now().strftime('%d. %m. %Y    %H:%M:%S')}"
    x_date = page_width - 10*mm
    y_date = page_height - 20*mm
    c.setFont('Helvetica', 8)
    c.drawRightString(x_date, y_date, date_text)

    font_size = 10
    text = "Mýtne transakcie spracované za obdobie:"
    c.setFont('Helvetica-Bold', font_size)
    c.drawCentredString(150, page_height - 50 * mm, text)
    
    # GENERATE 
    EMS_DWH=EMS_DWHDB.objects.get(user=1)
    con_EMSDWH_STDBY = cx_Oracle.connect(EMS_DWH.EMS_DWHDB_username+"/"+EMS_DWH.EMS_DWHDB_password+"@"+EMS_DWH.EMS_DWHDB_hostname+":"+EMS_DWH.EMS_DWHDB_port+"/"+EMS_DWH.EMS_DWHDB_servicename)
    cur_EMSDWH = con_EMSDWH_STDBY.cursor()
    cursor = con_EMSDWH_STDBY.cursor()

    query ="""SELECT * FROM   "REP_CRM"."V_LEGO_REP_01" "V_LEGO_REP_01"
                WHERE "V_LEGO_REP_01"."TRANSACTION_ARRIVAL_MONTH" >= add_months(trunc(sysdate, 'MM'), -1)
                AND "V_LEGO_REP_01"."TRANSACTION_ARRIVAL_MONTH" < trunc(sysdate, 'MM')
                ORDER BY OWNER_NAME, PASSAGEDATE_MONTHLY"""
    with con_EMSDWH_STDBY.cursor() as cursor:
            cursor.execute(query)
            #rows = cursor.fetchall()
            #rows = namedtuplefetchall(cursor)
            data = dictfetchall(cursor)
    print(data)

    filtered_data = [row for row in data if row['OWNER_NAME'] == 'Granvia']
    print(filtered_data)
  

    # print(granvia_data['OWNER_NAME'])
    # Save and return the PDF document
    c.save()
    # print(data)


    return response
import io
from reportlab.lib.units import inch
def generate_report(request):
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter, bottomup=0)
    textob = c.beginText()
    textob.setTextOrigin(inch, inch)
    textob.setFont("Helvetica", 14)

    EMS_DWH=EMS_DWHDB.objects.get(user=1)
    con_EMSDWH_STDBY = cx_Oracle.connect(EMS_DWH.EMS_DWHDB_username+"/"+EMS_DWH.EMS_DWHDB_password+"@"+EMS_DWH.EMS_DWHDB_hostname+":"+EMS_DWH.EMS_DWHDB_port+"/"+EMS_DWH.EMS_DWHDB_servicename)
    cur_EMSDWH = con_EMSDWH_STDBY.cursor()
    cursor = con_EMSDWH_STDBY.cursor()

    query ="""SELECT * FROM   "REP_CRM"."V_LEGO_REP_01" "V_LEGO_REP_01"
                WHERE "V_LEGO_REP_01"."TRANSACTION_ARRIVAL_MONTH" >= add_months(trunc(sysdate, 'MM'), -1)
                AND "V_LEGO_REP_01"."TRANSACTION_ARRIVAL_MONTH" < trunc(sysdate, 'MM')
                ORDER BY OWNER_NAME, PASSAGEDATE_MONTHLY"""
    with con_EMSDWH_STDBY.cursor() as cursor:
            cursor.execute(query)
            #rows = cursor.fetchall()
            #rows = namedtuplefetchall(cursor)
            data = dictfetchall(cursor)

    filtered_data = [row for row in data if row['OWNER_NAME'] == 'Granvia']


    # Create a list of strings, where each string represents a row in the filtered data

    filtered_data_strings = [f"{row['OWNER_NAME']} {row['PASSAGEDATE_MONTHLY']} {row['PASSAGEDATE_MONTHLY']}" for row in filtered_data]


    for line in filtered_data_strings:

        textob.textLine(line)


    c.drawText(textob)

    c.showPage()

    c.save()

    buf.seek(0)

    return FileResponse(buf, as_attachment=True, filename='test.pdf')

from django.template.loader import get_template
from xhtml2pdf import pisa
def render_pdf_view(request):
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    template_path = os.path.join(BASE_DIR, 'work', 'templates', 'ems', 'lego_pdf.html')
    print(template_path)

    EMS_DWH=EMS_DWHDB.objects.get(user=1)
    con_EMSDWH_STDBY = cx_Oracle.connect(EMS_DWH.EMS_DWHDB_username+"/"+EMS_DWH.EMS_DWHDB_password+"@"+EMS_DWH.EMS_DWHDB_hostname+":"+EMS_DWH.EMS_DWHDB_port+"/"+EMS_DWH.EMS_DWHDB_servicename)
    cur_EMSDWH = con_EMSDWH_STDBY.cursor()
    cursor = con_EMSDWH_STDBY.cursor()

    query ="""SELECT * FROM   "REP_CRM"."V_LEGO_REP_01" "V_LEGO_REP_01"
                WHERE "V_LEGO_REP_01"."TRANSACTION_ARRIVAL_MONTH" >= add_months(trunc(sysdate, 'MM'), -1)
                AND "V_LEGO_REP_01"."TRANSACTION_ARRIVAL_MONTH" < trunc(sysdate, 'MM')
                ORDER BY OWNER_NAME, PASSAGEDATE_MONTHLY"""
    with con_EMSDWH_STDBY.cursor() as cursor:
            cursor.execute(query)
            #rows = cursor.fetchall()
            #rows = namedtuplefetchall(cursor)
            data = dictfetchall(cursor)

    # filtered_data = [row for row in data if row['OWNER_NAME'] == 'Granvia']
    # context = {'myvar': 'sadfsdfasdfadsfdsfds'}
    filtered_data = [row for row in data if row['OWNER_NAME'] == 'Granvia']
    context = {'myvar': filtered_data}
    print(context)


    response =HttpResponse(content_type='aplication/pdf')

    # if download
    # response['Content-Disposition'] = attachment; 'filename="report.pdf"'

    response['Content-Disposition'] = 'filename="report.pdf"'

    template = get_template(template_path)
    html = template.render(context)

    pisa_status=pisa.CreatePDF(
        html, dest=response
    )
    if pisa_status.err:
        return HttpResponse('sadfsdafdsafdfsa <pre>'+ html + '</pre>')
    return response


@login_required
def pos_user(request):
    billdb=BILLDB.objects.get(user=request.user)
    con_BILLDB_STDBY = cx_Oracle.connect(billdb.BILLDB_username+"/"+billdb.BILLDB_password+"@"+billdb.BILLDB_hostname+":"+billdb.BILLDB_port+"/"+billdb.BILLDB_servicename)
    cursor = con_BILLDB_STDBY.cursor()

    sql_ = """SELECT POS_ID, POS_NUMBER, POS_NAME, pos.POS_TYPE_CODE, POS_TYPE_NAME, pos.RETAIL_PARTNER_ID, rp.RETAIL_PARTNER_NAME
                                    from billien_co.pos pos
                                    left join billien_co.POS_TYPE_L typ on pos.POS_TYPE_CODE=typ.POS_TYPE_CODE
                                    left join billien_co.retail_partner rp on pos.retail_partner_id = rp.retail_partner_id
                                    where pos.priority != '-1'
                                    and typ.LANGUAGE_CODE='SK'
                                    order by POS_NUMBER
                                """
    with con_BILLDB_STDBY.cursor() as cursor:
        cursor.execute(sql_)
            #rows = cursor.fetchall()        
            #rows = namedtuplefetchall(cursor)
        Poses = dictfetchall(cursor)
    users = ''
    if 'x' in request.GET and request.GET['y'] =='':
        
    # if x is not None and x != '' :
        x = request.GET['x']
        y = request.GET['y']

        sql_ROLE_NAME ="""SELECT
                                pos.pos_id,
                                ad.user_id,
                                ad.user_name,
                                ad.first_name,
                                ad.last_name,
                                ad.user_status_code,
                                USER_STATUS_NAME,
                                ad.selfcare_user_group_id,
                                CAST(from_tz(CAST(ad.creation_datetime AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET' AS DATE) AS creation_datetime,
                                CAST(from_tz(CAST(ad.expiration_date AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET' AS DATE) AS expiration_date,
                                CAST(from_tz(CAST(ad.deleted_on AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET' AS DATE) AS deleted_on,
                                LISTAGG(rl.role_name || '; ') WITHIN GROUP(
                                    ORDER BY
                                        usr.user_id
                                ) as "ROLE"
                            FROM
                                billien_ac.user_detail    ad
                                LEFT JOIN billien_ac.user_in_role   usr ON ad.user_id = usr.user_id
                                LEFT JOIN billien_ac.role           rl ON usr.role_id = rl.role_id
                                LEFT JOIN billien_ac.user_in_pos    pos ON ad.user_id = pos.user_id
                                left join BILLIEN_AC.user_status_l stat on ad.USER_STATUS_CODE=stat.USER_STATUS_CODE
                            WHERE
                                POS_ID=:id
                                and stat.LANGUAGE_CODE='SK'
                            GROUP BY
                                pos.pos_id,
                                ad.user_id,
                                ad.first_name,
                                ad.last_name,
                                ad.user_status_code,
                                stat.USER_STATUS_NAME,
                                ad.selfcare_user_group_id,
                                ad.user_name,
                                ad.selfcare_user_group_id,
                                ad.creation_datetime,
                                ad.expiration_date,
                                ad.deleted_on
                            ORDER BY
                                creation_datetime DESC
                                """
        with con_BILLDB_STDBY.cursor() as cursor:
            # c = '%' + q + '%'
            cursor.execute(sql_ROLE_NAME, id=x)
            #rows = cursor.fetchall()
            #rows = namedtuplefetchall(cursor)
            users = dictfetchall(cursor)
    elif 'y' in request.GET and request.GET['x'] =='':
        x =None
        y = request.GET['y']
        sql_ROLE_NAME ="""SELECT
                                pos.pos_id,
                                ad.user_id,
                                ad.user_name,
                                ad.first_name,
                                ad.last_name,
                                ad.user_status_code,
                                USER_STATUS_NAME,
                                ad.selfcare_user_group_id,
                                CAST(from_tz(CAST(ad.creation_datetime AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET' AS DATE) AS creation_datetime,
                                CAST(from_tz(CAST(ad.expiration_date AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET' AS DATE) AS expiration_date,
                                CAST(from_tz(CAST(ad.deleted_on AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET' AS DATE) AS deleted_on,
                                LISTAGG(rl.role_name || '; ') WITHIN GROUP(
                                    ORDER BY
                                        usr.user_id
                                ) as "ROLE"
                            FROM
                                billien_ac.user_detail    ad
                                LEFT JOIN billien_ac.user_in_role   usr ON ad.user_id = usr.user_id
                                LEFT JOIN billien_ac.role           rl ON usr.role_id = rl.role_id
                                LEFT JOIN billien_ac.user_in_pos    pos ON ad.user_id = pos.user_id
                                left join BILLIEN_AC.user_status_l stat on ad.USER_STATUS_CODE=stat.USER_STATUS_CODE
                            WHERE
                                user_name LIKE :usr
                                and stat.LANGUAGE_CODE='SK'
                            GROUP BY
                                pos.pos_id,
                                ad.user_id,
                                ad.first_name,
                                ad.last_name,
                                ad.user_status_code,
                                stat.USER_STATUS_NAME,
                                ad.selfcare_user_group_id,
                                ad.user_name,
                                ad.selfcare_user_group_id,
                                ad.creation_datetime,
                                ad.expiration_date,
                                ad.deleted_on
                            ORDER BY
                                creation_datetime DESC
                                """
        with con_BILLDB_STDBY.cursor() as cursor:
            c = '%' + y + '%'
            cursor.execute(sql_ROLE_NAME, usr=c)
            #rows = cursor.fetchall()
            #rows = namedtuplefetchall(cursor)
            users = dictfetchall(cursor)

    context = { 'Poses':Poses, 'users':users }
    return render(request,'ems/poses_users.html', context)

@login_required
def dwh_control(request):
    import cx_Oracle
    # cis_db=CISDB.objects.get(user=request.user)
    EMS_DWH=EMS_DWHDB.objects.get(user=request.user)
    con_EMSDWH_STDBY = cx_Oracle.connect(EMS_DWH.EMS_DWHDB_username+"/"+EMS_DWH.EMS_DWHDB_password+"@"+EMS_DWH.EMS_DWHDB_hostname+":"+EMS_DWH.EMS_DWHDB_port+"/"+EMS_DWH.EMS_DWHDB_servicename)
    cur_EMSDWH = con_EMSDWH_STDBY.cursor()

    sql_logs = """select STEP, TO_CHAR(from_tz(CAST(STEP_DATE_TIME AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET', 'DD.MM.YYYY HH24:MI:SS') as STEP_DATE_TIME_CET
                  from BILLIEN_STA.v_dwh_load_status"""

    with con_EMSDWH_STDBY.cursor() as cursor:
        cursor.execute(sql_logs)
        dwh_logs = dictfetchall(cursor)
        
    sql_detail = """select ID_DWH_LOAD_DETAIL, /*SOURCE_SYSTEM_NAME,*/ LOAD_NUMBER, PARTITION_ID, LOAD_STAGE, STATUS,
                    TO_CHAR(from_tz(CAST(DATE_FROM AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET', 'DD.MM.YYYY HH24:MI:SS') as DATE_FROM_CET,
                    TO_CHAR(from_tz(CAST(DATE_TO AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET', 'DD.MM.YYYY HH24:MI:SS') as DATE_TO_CET,
                    TO_CHAR(from_tz(CAST(LOAD_START AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET', 'DD.MM.YYYY HH24:MI:SS') as LOAD_START_CET,
                    TO_CHAR(from_tz(CAST(LOAD_END AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET', 'DD.MM.YYYY HH24:MI:SS') as LOAD_END_CET,
                    round((LOAD_END - LOAD_START) * 24 * 60, 1) AS diff_minutes,
                    TABLE_NAME, COMMAND
                    from BILLIEN_STA.dwh_load_detail
                    where LOAD_START >= TRUNC(sysdate-3)
                    order by LOAD_START desc
                  """
    with con_EMSDWH_STDBY.cursor() as cursor:
        cursor.execute(sql_detail)
        dwh_detail = dictfetchall(cursor)

    sql_error = """SELECT 
                    ID_DWH_LOAD_ERROR,
                    ID_DWH_LOAD_DETAIL,
                    ERROR ,
                    TO_CHAR(from_tz(CAST(CREATED_ON AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET', 'DD.MM.YYYY HH24:MI:SS') as CREATED_ON
                    FROM BILLIEN_STA.dwh_load_error order by ID_DWH_LOAD_ERROR desc
                  """
    with con_EMSDWH_STDBY.cursor() as cursor:
        cursor.execute(sql_error)
        dwh_error = dictfetchall(cursor)

    sql_merge_tables = """
                        select 
                        LOAD_NUMBER, /*SOURCE_SYSTEM_NAME,*/	PARTITION_ID,	IS_MERGED_FLAG,	
                        TO_CHAR(from_tz(CAST(MERGE_FINISHED_DATETIME AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET', 'DD.MM.YYYY HH24:MI:SS') as 
                        MERGE_FINISHED_DATETIME,	IS_PREPARED_TO_MERGE,	
                        IS_EXPORT_RUNNING_FLAG,	LAST_RUN_ERROR_MESSAGE
                        from BILLIEN_STA.dwh_merge_table
                        order by LOAD_NUMBER desc
                  """
    with con_EMSDWH_STDBY.cursor() as cursor:
        cursor.execute(sql_merge_tables)
        dwh_merge = dictfetchall(cursor)

    sql_tables = """SELECT ext.*, CAST(FROM_TZ(CAST(LAST_RUN AS TIMESTAMP), 'GMT') AT TIME ZONE 'CET' AS DATE) as LAST_RUN_CET  
                    FROM BILLIEN_STA.DWH_EXPORT_TABLE ext order by LAST_RUN desc
                  """
    with con_EMSDWH_STDBY.cursor() as cursor:
        cursor.execute(sql_tables)
        dwh_tables = dictfetchall(cursor)

    context={'dwh_logs': dwh_logs, 'dwh_detail':dwh_detail, 'dwh_error':dwh_error, 'dwh_tables':dwh_tables, 'dwh_merge':dwh_merge }
    return render(request,'ems/dwh_control.html', context)
########## LEGO REPORT view and manual PDF generation ##########
from django.http import HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa

@login_required
def lego_rep_CRM(request):
    export = ''
    column_names_list = ''
    q = None

    con_DWH_REP_CRM= cx_Oracle.connect("DB_USER/DB_PASSWORD@db-host.internal.example.com:1521/dwh")
    
    sql_logs = """select * from V_LEGO_REP_01
                    order by TRANSACTION_ARRIVAL_MONTH desc"""

    with con_DWH_REP_CRM.cursor() as cursor:
        cursor.execute(sql_logs)
        dwh_lego_view = dictfetchall(cursor)

    sql_removed_bilds = """select * from KDX_REMOVED_BILL_IDS"""

    with con_DWH_REP_CRM.cursor() as cursor:
        cursor.execute(sql_removed_bilds)
        dwh_lego_removed_bill = dictfetchall(cursor)
        # print(dwh_lego_removed_bill)

    if request.method == 'POST':

        if 'generation_pdf' in request.POST:
            MM = request.POST.get('MM')  # Get the value of MM from the form
            pdfmetrics.registerFont(TTFont('DejaVuSans', './eets/static/fonts/DejaVuSans.ttf'))
            pdfmetrics.registerFont(TTFont('Arial Unicode MS', './eets/static/fonts/arial-unicode-ms.ttf'))
            
            from datetime import datetime
            current_datetime = datetime.now()
            formatted_datetime = current_datetime.strftime("Dátum vygenerovania: %d. %m. %Y %H:%M:%S")
            previous_month = current_datetime - relativedelta(months=1)
            formatted_previous = previous_month.strftime("%m/%Y")
            formatted_previous_month = previous_month.strftime("%Y")

            con_DWH_REP_CRM= cx_Oracle.connect("DB_USER/DB_PASSWORD@db-host.internal.example.com:1521/dwh")

            pdf_sql = """
                    select * from V_LEGO_REP_01
                    where TRANSACTION_ARRIVAL_MONTH = TO_DATE('01-' || :MM || '-2025', 'DD-MM-YYYY')
                    order by PASSAGEDATE_MONTHLY asc
                    """
            with con_DWH_REP_CRM.cursor() as cursor:
                cursor.execute(pdf_sql, MM=MM)
                pdf_data = dictfetchall(cursor)

            filtered_data_granvia = list(filter(lambda x: x['OWNER_NAME'] == 'Granvia', pdf_data))
            sum_spolu_granvia = sum(row['SPOLU'] for row in filtered_data_granvia)
            
            filtered_data_zbl = list(filter(lambda x: x['OWNER_NAME'] == 'Zero Bypass Limited', pdf_data))
            sum_spolu_zbl = sum(row['SPOLU'] for row in filtered_data_zbl)

            filtered_data_stat = list(filter(lambda x: x['OWNER_NAME'] == 'Štát', pdf_data))
            sum_spolu_stat = round(sum(row['SPOLU'] or 0 for row in filtered_data_stat), 2)

            ppp_sql = """
                    select PASSAGEDATE_MONTHLY, SUM(SPOLU) as SPOLU from V_LEGO_REP_01
                    where TRANSACTION_ARRIVAL_MONTH = TO_DATE('01-' || :MM || '-2025', 'DD-MM-YYYY')
                    and OWNER_NAME in ('Granvia','Zero Bypass Limited')
                    group by PASSAGEDATE_MONTHLY
                    order by PASSAGEDATE_MONTHLY asc
                    """
            with con_DWH_REP_CRM.cursor() as cursor:
                cursor.execute(ppp_sql, MM=MM)
                ppp_data = dictfetchall(cursor)
            sum_spolu_ppp = round(sum(row['SPOLU'] for row in ppp_data), 2)

            spolu_sql = """
                    select PASSAGEDATE_MONTHLY, SUM(SPOLU) as SPOLU from V_LEGO_REP_01
                    where TRANSACTION_ARRIVAL_MONTH = TO_DATE('01-' || :MM || '-2025', 'DD-MM-YYYY')
                    group by PASSAGEDATE_MONTHLY
                    order by PASSAGEDATE_MONTHLY asc
                    """
            with con_DWH_REP_CRM.cursor() as cursor:
                cursor.execute(spolu_sql, MM=MM)
                spolu_data = dictfetchall(cursor)
            sum_spolu = round(sum(row['SPOLU'] for row in spolu_data), 2)

            """Generates a PDF file with the provided content."""
            # odmazanie obsahu
            directory = './Reports/Lego/'  # specify the directory path

            for filename in os.listdir(directory):
                file_path = os.path.join(directory, filename)
                try:
                    os.remove(file_path)
                except Exception as e:
                    print(f"Error removing file {file_path}: {e}")

            output_filename = './Reports/Lego/EETS_LEGO_REP_'+MM+'2025.pdf'
            c = canvas.Canvas(output_filename, pagesize=A4)
            c.drawImage("./eets/static/img/logo_myto.png", 35, 780, width=80, height=50)
            c.setFont("Arial Unicode MS", 8)  # Use a font that supports Czech

            # Set up the text for the information
            c.drawString(414, 800, "Číslo zostavy: EETS_LEGO_REP_01")
            c.drawString(503, 785, "PDF formát")
            c.drawString(385, 770, formatted_datetime)

            # Set up the font for the title and text
            c.setFont("Helvetica-Bold", 14)
            c.drawString(165, 750 , "Prerozdelenie mýta medzi vlastníkov ciest")

            c.setFont("Helvetica-Bold", 10)
            c.drawString(35, 720, "Mýtne transakcie spracované za obdobie:")

            c.setFont("Arial Unicode MS", 8)  # Use a font that supports Czech
            c.drawString(250, 720, MM +'/' +formatted_previous_month)

            c.setFont("Helvetica-Bold", 10)
            c.drawString(35, 690, "Vlastník:")

            c.setFont("Arial Unicode MS", 8)  # Use a font that supports Czech
            c.drawString(80, 690, "Granvia")
            #--------------------
            c.setFont("Helvetica-Bold", 10)
            c.drawString(35, 585, "Vlastník:")

            c.setFont("Arial Unicode MS", 8)  # Use a font that supports Czech
            c.drawString(80, 585, "Zero Bypass Limited")
            #-------------------------
            c.setFont("Helvetica-Bold", 10)
            c.drawString(35, 485, "Celkom (PPP úseky) - Granvia + R7/D4")
            #-------------------------------
            c.setFont("Helvetica-Bold", 10)
            c.drawString(35, 410, "Vlastník:")

            c.setFont("Arial Unicode MS", 8)  # Use a font that supports Czech
            c.drawString(80, 410, "Štát")
            #-------------------------------
            c.setFont("Helvetica-Bold", 10)
            c.drawString(35, 295, "Celkom - PPP + Štát")

        # Granvia
            table_data_GRANVIA = [
                ["Mesiac prejazdu mýtneho úseku", "EETS", "SPOLU"],
            ]
            for row in filtered_data_granvia:
                table_data_GRANVIA.append([row['PASSAGEDATE_MONTHLY'], f"{row['SPOLU']:,.2f}".replace(',', ' '), f"{row['SPOLU']:,.2f}".replace(',', ' ')])
            table_data_GRANVIA.append(['Celková suma:', f"{sum_spolu_granvia:,.2f}".replace(',', ' '), f"{sum_spolu_granvia:,.2f}".replace(',', ' ')])

        # ZBL
            table_data_ZBL = [
                ["Mesiac prejazdu mýtneho úseku", "EETS", "SPOLU"],
            ]
            for row in filtered_data_zbl:
                table_data_ZBL.append([row['PASSAGEDATE_MONTHLY'], f"{row['SPOLU']:,.2f}".replace(',', ' '), f"{row['SPOLU']:,.2f}".replace(',', ' ')])
            table_data_ZBL.append(['Celková suma:', f"{sum_spolu_zbl:,.2f}".replace(',', ' '), f"{sum_spolu_zbl:,.2f}".replace(',', ' ')])

        # PPP
            table_data_PPP = [
                ["Mesiac prejazdu mýtneho úseku", "EETS", "SPOLU"],
            ]
            for row in ppp_data:
                table_data_PPP.append([row['PASSAGEDATE_MONTHLY'], f"{row['SPOLU']:,.2f}".replace(',', ' '), f"{row['SPOLU']:,.2f}".replace(',', ' ')])
            table_data_PPP.append(['Celková suma:', f"{sum_spolu_ppp:,.2f}".replace(',', ' '), f"{sum_spolu_ppp:,.2f}".replace(',', ' ')])

        # Stat
            table_data_stat = [
                ["Mesiac prejazdu mýtneho úseku", "EETS", "SPOLU"],
            ]
            for row in filtered_data_stat:
                table_data_stat.append([row['PASSAGEDATE_MONTHLY'], "{:_}".format(row['SPOLU']).replace('_', ' ') if row['SPOLU'] is not None else '', "{:_}".format(row['SPOLU']).replace('_', ' ') if row['SPOLU'] is not None else ''])
            table_data_stat.append(['Celková suma:', "{:_}".format(sum_spolu_stat).replace('_', ' ') if sum_spolu_stat is not None else '', "{:_}".format(sum_spolu_stat).replace('_', ' ') if sum_spolu_stat is not None else ''])
        # Celkom spolu
            table_data_spolu = [
                ["Mesiac prejazdu mýtneho úseku", "EETS", "SPOLU"],
            ]
            for row in spolu_data:
                table_data_spolu.append([row['PASSAGEDATE_MONTHLY'], "{:_}".format(row['SPOLU']).replace('_', ' ') if row['SPOLU'] is not None else '', "{:_}".format(row['SPOLU']).replace('_', ' ') if row['SPOLU'] is not None else ''])
            table_data_spolu.append(['Celková suma:', f"{sum_spolu:_}".replace('_', ' '), f"{sum_spolu:_}".replace('_', ' ')])
            # Define the column widths
            col_widths = [2.5*inch,  1.5*inch + 2.5*inch/2.54, 1.5*inch]  # Add a spacer column

            # Create the table
            table = Table(table_data_ZBL, colWidths=col_widths)
            
            table_style = TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Arial Unicode MS'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Align the entire table to the left
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, -1), (2, -1), 'Helvetica-Bold'),  # Make the last row's cells bold
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),  # Right-align the second column
                ('ALIGN', (2, 0), (2, -1), 'RIGHT'),  # Right-align the third column
                ('LEADING', (0, 0), (-1, -1), 5),  # Reduce the space between rows to 4 points
                # ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Add a grid to the table
            ])

            table.setStyle(table_style)
            # table_gr.setStyle(table_style)
            # table_ppp.setStyle(table_style)
            # table_stat.setStyle(table_style)
            # table_spolu.setStyle(table_style)

            # Get the page dimensions
            table.wrapOn(c, 120, 800)
            # table_gr.wrapOn(c, 120, 800)
            # table_ppp.wrapOn(c, 120, 800)
            # table_stat.wrapOn(c, 120, 800)
            # table_spolu.wrapOn(c, 120, 800)

            # Calculate the table width
            table_width = sum(col_widths)
            # table_gr.drawOn(c, (A4[0] - table_width) / 4.4, 635)
            table.drawOn(c, (A4[0] - table_width) / 4.4, 530)
            # table_ppp.drawOn(c, (A4[0] - table_width) / 4.4, 430)
            # table_stat.drawOn(c, (A4[0] - table_width) / 4.4, 335)
            # table_spolu.drawOn(c, (A4[0] - table_width) / 4.4, 215)

            # Set the line color to red
            c.setStrokeColorRGB(1, 0, 0)

            # Draw a vertical line from (100, 100) to (100, 200)
            c.line(560, 140, 560, 830)
            c.line(35, 140, 560, 140)

            c.setFont("Arial Unicode MS", 7)  # Use a font that supports Czech
            c.drawString(448, 150, "Sumy sú uvedené v EUR bez DPH.")

            c.setFont("Arial Unicode MS", 6)  # Use a font that supports Czech
            text = """
            example, a.s., Westend Plazza, Lamačská cesta 3/B, 841 04 Bratislava, Slovenská republika
            IČO/ID Nr.: 44 500 734, DIČ/Tax ID Nr.: 2022712153, IČ DPH/VAT Nr.: SK2022712153
            Spoločnosť je zapísaná v Obchodnom registri Okresného súdu Bratislava I. v Odd.: Sa, vložka č.: 4646/B
            Company is registered at Companies Register of the District Court Bratislava I. in Section: Sa, Insert No.: 4646/B"""

            lines = [line for line in text.split('\n') if line.strip() != '']
            y = 130
            for line in lines:
                c.drawString(16, y, line)
                y -= 10  # adjust the y-coordinate for each line

            # cislo strany
            c.saveState()
            c.setFont("Arial Unicode MS", 7)
            page_info = "Strana (Page) %d/1" % (c.getPageNumber())
            c.drawRightString(550, 20, page_info)
            c.restoreState()
            c.showPage()
            c.save()

            # Open the PDF file
            with open(output_filename, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/pdf')
                response['Content-Disposition'] = 'inline; filename="EETS_LEGO_REP_'+MM+'2025.pdf"'
            return response           

        if 'delete_procedure' in request.POST:
            MM = request.POST.get('MM')  # Get the value of MM from the form
            YYYY = request.POST.get('YYYY')  # Get the value of YYYY from the form
            procedure_sql = """
                DECLARE
                    o_result_code NUMBER;
                    o_error_message VARCHAR2(4000);
                BEGIN
                    kdx_pkg_lego_ppp_2020.lego_drop_temporary_agg_tables(
                    o_result_code => o_result_code,
                    o_error_message => o_error_message,
                    i_date_from_01_mm_yyyy => TO_DATE('01-' || :MM || '-' || :YYYY, 'DD-MM-YYYY')
                    );
                    kdx_pkg_lego_ppp_2020.lego_del_kpbpm_by_trans_ar_mon(
                    o_result_code => o_result_code,
                    o_error_message => o_error_message,
                    i_date => TO_DATE('01-' || :MM || '-' || :YYYY, 'DD-MM-YYYY')
                    );
                    kdx_pkg_lego_ppp_2020.lego_del_kspoi_by_trans_ar_mon(
                    o_result_code => o_result_code,
                    o_error_message => o_error_message,
                    i_date => TO_DATE('01-' || :MM || '-' || :YYYY, 'DD-MM-YYYY')
                    );
                END;
                """
            with con_DWH_REP_CRM.cursor() as cursor:
                cursor.execute(procedure_sql, MM=MM, YYYY=YYYY)
                con_DWH_REP_CRM.commit()  # Commit the changes
                    # print("Procedura spuštěna")
            messages.info(request, f"Vymazanie dat za: {MM}-{YYYY} dokončené")
            return HttpResponseRedirect(reverse('work:lego_rep_CRM'))

        if 'delete_REMOVED_BILL_ID' in request.POST:
            TRUNCATE_sql = """TRUNCATE TABLE KDX_REMOVED_BILL_IDS
                """
            with con_DWH_REP_CRM.cursor() as cursor:
                cursor.execute(TRUNCATE_sql)
                con_DWH_REP_CRM.commit()  # Commit the changes
                    # print("Procedura spuštěna")
            messages.info(request, f"Vymazanie dat z tabulky KDX_REMOVED_BILL_IDS dokončené")
            return HttpResponseRedirect(reverse('work:lego_rep_CRM'))
        
        if 'insert_REMOVED_BILL_ID' in request.POST:
            insert_sql = """
                        INSERT INTO kdx_removed_bill_ids (transaction_arrival_month, removed_bill_id)
                        SELECT to_date(TO_CHAR(ADD_MONTHS(TRUNC(Passage_month, 'MM'), -1), 'dd.mm.yyyy'), 'dd.mm.yyyy'), 
                            bill_id
                        FROM (
                        SELECT distinct
                            /*+ FULL(rte) PARALLEL(6) */
                            TRUNC(rte.OBU_TIMESTAMP, 'MM') AS Passage_month,
                            b.bill_id
                        FROM eets_maa.EETS_RTE_ALL rte
                            inner join eets_maa.BILL_ITEM bi on bi.bill_item_id = rte.bill_item_id
                            inner join eets_maa.BILL b on b.bill_id = bi.bill_id
                        where (b.ISSUE_DATE_TIME >= TRUNC(ADD_MONTHS(SYSDATE, -1), 'MM') 
                            AND b.ISSUE_DATE_TIME < TRUNC(ADD_MONTHS(SYSDATE, 1), 'MM') 
                            AND b.BILL_SESSION_ID IS NOT NULL AND b.BILL_ISSUER_CODE=20)   
                            and rte.obu_timestamp >= TRUNC(SYSDATE, 'MM')
                        )
                """
            with con_DWH_REP_CRM.cursor() as cursor:
                cursor.execute(insert_sql)
                con_DWH_REP_CRM.commit()  # Commit the changes
                    # print("Procedura spuštěna")
            messages.info(request, f"Naplnenie tabulky KDX_REMOVED_BILL_IDS dokončené")
            return HttpResponseRedirect(reverse('work:lego_rep_CRM'))

        if 'run_procedure' in request.POST:
            MM = request.POST.get('MM')  # Get the value of MM from the form
            mydate = request.POST.get('mydate')  # Get the value of mydate from the form
            procedure_sql = """
                DECLARE
                    o_result_code NUMBER;
                    o_error_message VARCHAR2(4000);
                BEGIN
                    kdx_pkg_lego_ppp_2020.lego_prepare_kpbpm_agg (
                      o_result_code => o_result_code,
                      o_error_message => o_error_message,
                      i_date_from_01_mm_yyyy => TO_DATE('01-' || :MM || '-2025', 'DD-MM-YYYY'),
                      i_date_from_0203_mm_yyyy => TO_DATE('03-' || :MM || '-2025', 'DD-MM-YYYY'),
                      i_parallel_degree => 6,
                      i_discount_calculated_on_from => TO_DATE('01-' || :MM || '-2025', 'DD-MM-YYYY'),
                      i_discount_calculated_on_to => TO_DATE(:mydate, 'DD-MM-YYYY')
                    );
                    kdx_pkg_lego_ppp_2020.lego_insert_kpbpm (
                      o_result_code => o_result_code,
                      o_error_message => o_error_message,
                      i_date_from_01_mm_yyyy => TO_DATE('01-' || :MM || '-2025', 'DD-MM-YYYY')
                    );
                    kdx_pkg_lego_ppp_2020.lego_prepare_kspoi_agg (
                      o_result_code => o_result_code,
                      o_error_message => o_error_message,
                      i_date_from_01_mm_yyyy => TO_DATE('01-' || :MM || '-2025', 'DD-MM-YYYY'),
                      i_date_from_0203_mm_yyyy => TO_DATE('03-' || :MM || '-2025', 'DD-MM-YYYY'),
                      i_discount_calculated_on_from => TO_DATE('01-' || :MM || '-2025', 'DD-MM-YYYY'),
                      i_discount_calculated_on_to => TO_DATE(:mydate, 'DD-MM-YYYY')
                    );
                END;
            """

            with con_DWH_REP_CRM.cursor() as cursor:
                cursor.execute(procedure_sql, MM=MM, mydate=mydate)
                con_DWH_REP_CRM.commit()  # Commit the changes
            # print("Procedura dokoncena")
            messages.info(request, f"Procedura dokoncena i_date_from_01_mm_yyyy => 01-{MM}-2025 a i_discount_calculated_on_to => {mydate}")
            return HttpResponseRedirect(reverse('work:lego_rep_CRM'))

    con_DWH_REP_CRM.close()  # Close the connection

    context = {'dwh_lego_view':dwh_lego_view, 'dwh_lego_removed_bill':dwh_lego_removed_bill }
    return render(request, 'lego_view.html', context)

########## PAY WELL ##########
import os
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime
from django.conf import settings
from django.shortcuts import render
from django.core.files.storage import FileSystemStorage
from django.http import FileResponse, Http404

import json
from office365.runtime.auth.user_credential import UserCredential
from office365.sharepoint.client_context import ClientContext

UPLOAD_DIR = os.path.join(settings.PAYWELL_ROOT, 'uploads')
os.makedirs(UPLOAD_DIR, exist_ok=True)

def paywell(request):
    message = ''
    pswd = '936BhU14'  # Heslo na odzipovanie
    transactions_summary = {
        'T_FP_CARD_TRANSACTIONS': [],
        'T_FP_FCI_CARD_TRANSACTIONS': [],
        'SUM': [],
    }

    # === PRIPOJENIE K ORACLE A NAČÍTANIE PREHĽADU ===
    try:
        EMS_DWH = EMS_DWHDB.objects.get(user=48)
        con = cx_Oracle.connect(EMS_DWH.EMS_DWHDB_username + "/" + EMS_DWH.EMS_DWHDB_password + "@" + EMS_DWH.EMS_DWHDB_hostname + ":" + EMS_DWH.EMS_DWHDB_port +"/" + EMS_DWH.EMS_DWHDB_servicename)
        cur = con.cursor()

        query_1 = """
            SELECT to_char(date_time, 'dd.mm.yyyy') AS DateOfTransaction, count(ID) as pocet
            FROM REP_CRM.T_FP_CARD_TRANSACTIONS 
            WHERE date_time >= trunc(SYSDATE -5) AND date_time < trunc(SYSDATE)
            GROUP BY to_char(date_time, 'dd.mm.yyyy') 
            ORDER BY 1
        """

        query_2 = """
            SELECT to_char(date_time, 'dd.mm.yyyy') AS DateOfTransaction, count(ID) as pocet
            FROM REP_CRM.T_FP_FCI_CARD_TRANSACTIONS 
            WHERE date_time >= trunc(SYSDATE -5) AND date_time < trunc(SYSDATE)
            GROUP BY to_char(date_time, 'dd.mm.yyyy') 
            ORDER BY 1
        """

        cur.execute(query_1)
        transactions_summary['T_FP_CARD_TRANSACTIONS'] = cur.fetchall()

        cur.execute(query_2)
        transactions_summary['T_FP_FCI_CARD_TRANSACTIONS'] = cur.fetchall()

        cur.close()
        con.close()
    except Exception as e:
        message += f'⚠️ Chyba pri načítaní prehľadu transakcií: {e}<br>'
    
    # Spojenie a súčet podľa dátumu
    from collections import defaultdict

    sum_per_date = defaultdict(int)

    for date, count in transactions_summary['T_FP_CARD_TRANSACTIONS']:
        sum_per_date[date] += count

    for date, count in transactions_summary['T_FP_FCI_CARD_TRANSACTIONS']:
        sum_per_date[date] += count

    # zoradenie podla datumu
    summary_sum = sorted(sum_per_date.items())

    transactions_summary['SUM'] = summary_sum

    # === 1. Nahrávanie ZIP súboru cez formulár ===
    if request.method == 'POST' and request.FILES.get('zipfile'):
        import os
        from django.conf import settings
        UPLOAD_DIR = os.path.join(settings.PAYWELL_ROOT, 'uploads')
        os.makedirs(UPLOAD_DIR, exist_ok=True)
        logger.info("------------ START upload suboru " )  # Zaznamenajte správu do info.log
        zipfile_uploaded = request.FILES['zipfile']
        if zipfile_uploaded.name.endswith('.zip'):
            fs = FileSystemStorage(location=UPLOAD_DIR)
            filename = fs.save(zipfile_uploaded.name, zipfile_uploaded)
            message = f'✅ Súbor "{filename}" bol úspešne nahraný.'
        else:
            message = '❌ Prosím, nahrajte platný ZIP súbor.'

    # === 2. Vymazanie súboru cez formulár ===
    if request.method == 'POST' and request.POST.get('delete_file'):
        import os
        from django.conf import settings
        UPLOAD_DIR = os.path.join(settings.PAYWELL_ROOT, 'uploads')
        os.makedirs(UPLOAD_DIR, exist_ok=True)
        logger.info("------------ START vymazanie suboru " )
        file_to_delete = request.POST['delete_file']
        file_path = os.path.join(UPLOAD_DIR, file_to_delete)
        if os.path.isfile(file_path):
            os.remove(file_path)
            message = f'🗑️ Súbor "{file_to_delete}" bol vymazaný.'
        else:
            message = f'❌ Súbor "{file_to_delete}" neexistuje.'
    
    # === 3. Vymazanie VŠETKÝCH súborov ===
    if request.method == 'POST' and request.POST.get('delete_all'):
        import os
        from django.conf import settings
        UPLOAD_DIR = os.path.join(settings.PAYWELL_ROOT, 'uploads')
        os.makedirs(UPLOAD_DIR, exist_ok=True)
        logger.info("------------ START vymazanie VSETKYCH suborov")
        deleted = 0
        for file in os.listdir(UPLOAD_DIR):
            try:
                os.remove(os.path.join(UPLOAD_DIR, file))
                deleted += 1
            except Exception as e:
                logger.warning(f"Nepodarilo sa zmazať {file}: {e}")
        message = f'🧹 Vymazaných {deleted} súborov.'

    # === 3. Odzipovanie všetkých ZIP súborov a spracovanie XML do jedného SQL ===
    if request.method == 'POST' and 'unzip_files' in request.POST:
        logger.info("------------ START odzipovanie " )
        insert_lines = []
        from datetime import datetime
        import os
        from django.conf import settings
        UPLOAD_DIR = os.path.join(settings.PAYWELL_ROOT, 'uploads')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        sql_file_name = f"insert_example_all_{timestamp}.sql"
        sql_file_path = os.path.join(UPLOAD_DIR, sql_file_name)

        # A. Odzipuj všetky ZIP súbory v priečinku
        for filename in os.listdir(UPLOAD_DIR):
            if filename.lower().endswith('.zip'):
                zip_path = os.path.join(UPLOAD_DIR, filename)
                try:
                    with zipfile.ZipFile(zip_path) as zip_file:
                        zip_file.setpassword(pswd.encode('utf-8'))
                        zip_file.extractall(UPLOAD_DIR)
                        message += f'✅ ZIP "{filename}" bol úspešne odzipovaný.<br>'
                except RuntimeError as e:
                    message += f'❌ Chyba pri odzipovaní "{filename}": {e}<br>'
                except Exception as e:
                    message += f'❌ Neočakávaná chyba so súborom "{filename}": {e}<br>'

        # B. Prejdi všetky .xml a vyber INSERTy
        for xml_file in os.listdir(UPLOAD_DIR):
            logger.info("------------ START priprava insertov " )
            if xml_file.lower().endswith('.xml'):
                xml_path = os.path.join(UPLOAD_DIR, xml_file)
                try:
                    tree = ET.parse(xml_path)
                    text = ET.tostring(tree.getroot(), encoding='utf-8', method='text').decode('utf-8')
                    insert_lines += [
                        line.strip() for line in text.splitlines() if line.strip().upper().startswith("INSERT INTO")
                    ]
                except Exception as e:
                    message += f'⚠️ Chyba pri XML "{xml_file}": {e}<br>'

        # C. Vytvor SQL súbor s INSERTmi
        if insert_lines:
            logger.info("------------ START vytvorenie sql " )
            try:
                with open(sql_file_path, 'w', encoding='utf-8') as sql_out:
                    sql_out.write('\n'.join(insert_lines))
                    sql_out.write('\nCOMMIT;\n')
                message += f'📄 INSERTy boli uložené do "{sql_file_name}".<br>'
            except Exception as e:
                message += f'❌ Chyba pri zápise SQL: {e}<br>'
        else:
            message += '⚠️ Neboli nájdené INSERT príkazy v žiadnom XML súbore.<br>'

    # === 6. Import SQL do Oracle DB ===
    if request.method =='POST' and 'import_sql' in request.POST:
        import os
        from django.conf import settings
        UPLOAD_DIR = os.path.join(settings.PAYWELL_ROOT, 'uploads')
        os.makedirs(UPLOAD_DIR, exist_ok=True)
        message += '🔵 Start importu do DB...<br>'
        logger.info("------------ START import sql do DB")
        try:
            sql_files = [f for f in os.listdir(UPLOAD_DIR) if f.endswith('.sql')]
            if sql_files:
                sql_files.sort()
                file_path = os.path.join(UPLOAD_DIR, sql_files[-1])

                with open(file_path, 'r', encoding='utf-8') as f:
                    full_sql = f.read()

                statements = [s.strip() for s in full_sql.split(';') if s.strip()]
                EMS_DWH = EMS_DWHDB.objects.get(user=48)
                con = cx_Oracle.connect(
                    f"{EMS_DWH.EMS_DWHDB_username}/{EMS_DWH.EMS_DWHDB_password}@{EMS_DWH.EMS_DWHDB_hostname}:{EMS_DWH.EMS_DWHDB_port}/{EMS_DWH.EMS_DWHDB_servicename}"
                )
                cur = con.cursor()

                message += f'✅  Pripojenie do DWH.<br>'
                message += f'❓ Spusteny import SQL súboru "{os.path.basename(file_path)}" do DWH.<br>'
                message += f'❓ Počet príkazov: {len(statements)}<br>'
                message += f'❓ Spúšťam príkazy:<br>'
                
                for stmt in statements:
                    try:
                        cur.execute(stmt)
                    except Exception as e:
                        message += f'❌ Chyba: {e}<br>'
                con.commit()
                cur.close()
                con.close()
                message += f'✅ SQL súbor \"{os.path.basename(file_path)}\" bol importovaný do DB.<br>'
            else:
                message += "❌ Nebol nájdený žiadny .sql súbor na import.<br>"
        except Exception as e:
            message += f"❌ Chyba importu do DB: {e}<br>"

    # === Zmazanie z DB ===
    if request.method =='POST' and 'delete_transactions_by_date' in request.POST:
        import os
        from django.conf import settings
        from datetime import datetime
        UPLOAD_DIR = os.path.join(settings.PAYWELL_ROOT, 'uploads')
        os.makedirs(UPLOAD_DIR, exist_ok=True)
        print("------------ START zmazanie transakcii")
        logger.info("------------ START zmazanie transakcii")
        try:
            # Formát: 15.05.2025
            delete_date = request.POST.get('delete_date')
            datetime.strptime(delete_date, '%d.%m.%Y')  # validácia

            EMS_DWH = EMS_DWHDB.objects.get(user=48)
            con = cx_Oracle.connect(
                f"{EMS_DWH.EMS_DWHDB_username}/{EMS_DWH.EMS_DWHDB_password}@{EMS_DWH.EMS_DWHDB_hostname}:{EMS_DWH.EMS_DWHDB_port}/{EMS_DWH.EMS_DWHDB_servicename}"
            )
            cur = con.cursor()

            delete_sql_1 = f"""
            DELETE FROM REP_CRM.T_FP_CARD_TRANSACTIONS
            WHERE date_time >= TO_DATE('{delete_date}', 'DD.MM.YYYY')
                AND date_time < TO_DATE('{delete_date}', 'DD.MM.YYYY') + 1
            """

            delete_sql_2 = f"""
            DELETE FROM REP_CRM.T_FP_FCI_CARD_TRANSACTIONS
            WHERE date_time >= TO_DATE('{delete_date}', 'DD.MM.YYYY')
                AND date_time < TO_DATE('{delete_date}', 'DD.MM.YYYY') + 1
            """


            cur.execute(delete_sql_1)
            cur.execute(delete_sql_2)
            con.commit()
            cur.close()
            con.close()

            message += f'🗑️ Transakcie k dátumu {delete_date} boli vymazané.<br>'
        except ValueError:
            message += f'❌ Nesprávny formát dátumu. Použi DD.MM.YYYY<br>'
        except Exception as e:
            message += f'❌ Chyba pri mazaní transakcií: {e}<br>'

    # === Generovanie XLSX reportu ===
    if request.method == 'POST' and 'generate_report' in request.POST:
        print("------------ START generovanie reportu")
        logger.info("------------ START generovanie reportu")

            # Definícia UPLOAD_DIR
        import os
        from django.conf import settings
        UPLOAD_DIR = os.path.join(settings.PAYWELL_ROOT, 'uploads')
        os.makedirs(UPLOAD_DIR, exist_ok=True)


        try:
            # Formát: 16.05.2024
            from datetime import datetime
            report_date = request.POST.get('report_date')
            datetime.strptime(report_date, '%d.%m.%Y')  # validácia

            EMS_DWH = EMS_DWHDB.objects.get(user=48)
            con = cx_Oracle.connect(
                f"{EMS_DWH.EMS_DWHDB_username}/{EMS_DWH.EMS_DWHDB_password}@{EMS_DWH.EMS_DWHDB_hostname}:{EMS_DWH.EMS_DWHDB_port}/{EMS_DWH.EMS_DWHDB_servicename}"
            )
            cur = con.cursor()

            # Načítajte SQL zo súboru alebo použite priamo reťazec
            sql_query = """
                        SELECT
                        i.VARIABLE_SYMBOL InvoiceNumber,
                        to_char(c.Customer_Number,'FM000000000009') ClientNumber,
                        to_char(a.Account_Number,'FM000000000009') AccountNumber,
                        i.Payment_Type_Name InvoiceType,  
                        i.Payment_Status_Name InvoiceStatus, 
                        i.Payment_Method_Name PaymentType, 
                        i.Payment_Amount AS Amount,  
                        i.DATE_OF_PAYMENT AS DatePaid,
                        CASE WHEN i.POS_OPERATOR_USERNAME IS NOT NULL THEN i.POS_OPERATOR_USERNAME ELSE i.WORKSTATION_TYPE_NAME END AS UserLoginCreate,
                        i.POS_NUMBER,
                        p.RETAIL_PARTNER_ABBR AS Location,
                        fp.VS,
                        to_char(fp.DATE_TIME,'yyyy.mm.dd hh24:mi:ss') AS DateTime,
                        to_char(fp.CARD_NUMBER) AS CARD_NUMBER,
                        fp.Amount AS FP_Amount,
                        to_char(fp.AUTH_CODE) AS AuthCode,
                        fp.Result AS Result,
                        fp.Terminal AS Terminal,
                        fp.Terminal_PABK AS Terminal_PABK
                        FROM REP_CRM.T_FP_CARD_TRANSACTIONS fp  
                        LEFT JOIN BILLIEN_MAA.Payments i ON i.VARIABLE_SYMBOL=fp.VS
                        LEFT JOIN BILLIEN_MAA.Accounts a ON i.Account_ID=a.Account_ID 
                        LEFT JOIN BILLIEN_MAA.Customers c ON a.Customer_ID=c.Customer_ID
                        LEFT JOIN BILLIEN_MAA.POSES p ON i.POS_NUMBER = p.POS_NUMBER
                        WHERE (fp.date_time >= Trunc(to_date(&DATUM, 'dd.mm.yyyy')) And fp.date_time < Trunc(to_date(&DATUM, 'dd.mm.yyyy')+1)) 
                        AND (fp.VS IS NOT NULL OR i.Payment_Method_Code=20) AND fp.TERMINAL != 'POSTPAY'
                        and fp.VS not like to_char(sysdate,'YY%') 
                        and fp.VS not like to_char( add_months( sysdate, -12 )-1, 'YY%')

                        UNION ALL

                        SELECT
                        i.VARIABLE_SYMBOL InvoiceNumber, 
                        to_char(c.Customer_Number,'FM000000000009') ClientNumber,
                        to_char(a.Account_Number,'FM000000000009') AccountNumber,
                        i.Payment_Type_Name InvoiceType,  
                        i.Payment_Status_Name InvoiceStatus, 
                        i.Payment_Method_Name PaymentType, 
                        i.Payment_Amount AS Amount,  
                        i.DATE_OF_PAYMENT AS DatePaid,
                        CASE WHEN i.POS_OPERATOR_USERNAME IS NOT NULL THEN i.POS_OPERATOR_USERNAME ELSE i.WORKSTATION_TYPE_NAME END AS UserLoginCreate,
                        i.POS_NUMBER,
                        p.RETAIL_PARTNER_ABBR AS Location,
                        fp.VS,
                        to_char(fp.DATE_TIME,'yyyy.mm.dd hh24:mi:ss') AS DateTime,
                        to_char(fp.CARD_NUMBER) AS CARD_NUMBER,
                        fp.Amount AS FP_Amount,
                        to_char(fp.AUTH_CODE) AS AuthCode,
                        fp.Result AS Result,
                        fp.Terminal AS Terminal,
                        NULL AS Terminal_PABK
                        FROM REP_CRM.T_FP_FCI_CARD_TRANSACTIONS fp  
                        LEFT JOIN BILLIEN_MAA.Payments i ON i.VARIABLE_SYMBOL=fp.VS
                        LEFT JOIN BILLIEN_MAA.Accounts a ON i.Account_ID=a.Account_ID 
                        LEFT JOIN BILLIEN_MAA.Customers c ON a.Customer_ID=c.Customer_ID
                        LEFT JOIN BILLIEN_MAA.POSES p ON i.POS_NUMBER = p.POS_NUMBER
                        WHERE (fp.date_time >= Trunc(to_date(&DATUM, 'dd.mm.yyyy')) And fp.date_time < Trunc(to_date(&DATUM, 'dd.mm.yyyy')+1)) 
                        AND (fp.VS IS NOT NULL OR i.Payment_Method_Code=30) AND fp.TERMINAL != 'POSTPAY'
                        and fp.VS not like to_char(sysdate,'YY%') 
                        and fp.VS not like to_char( add_months( sysdate, -12 )-1, 'YY%')
                        
                        UNION ALL

                        SELECT
                        i.VARIABLE_SYMBOL InvoiceNumber, 
                        to_char(c.Customer_Number,'FM000000000009') ClientNumber,
                        to_char(a.Account_Number,'FM000000000009') AccountNumber,
                        i.Payment_Type_Name InvoiceType,  
                        i.Payment_Status_Name InvoiceStatus, 
                        i.Payment_Method_Name PaymentType, 
                        i.Payment_Amount AS Amount,  
                        i.DATE_OF_PAYMENT AS DatePaid,
                        CASE WHEN i.POS_OPERATOR_USERNAME IS NOT NULL THEN i.POS_OPERATOR_USERNAME ELSE i.WORKSTATION_TYPE_NAME END AS UserLoginCreate,
                        i.POS_NUMBER,
                        p.RETAIL_PARTNER_ABBR AS Location,
                        fp.VS,
                        to_char(fp.DATE_TIME,'yyyy.mm.dd hh24:mi:ss') AS DateTime,
                        to_char(fp.CARD_NUMBER) AS CARD_NUMBER,
                        fp.Amount AS FP_Amount,
                        to_char(fp.AUTH_CODE) AS AuthCode,
                        fp.Result AS Result,
                        fp.Terminal AS Terminal,
                        fp.Terminal_PABK AS Terminal_PABK
                        FROM BILLIEN_MAA.Payments i   
                        LEFT JOIN REP_CRM.T_FP_CARD_TRANSACTIONS fp ON i.VARIABLE_SYMBOL = fp.VS
                        LEFT JOIN BILLIEN_MAA.Accounts a ON i.Account_ID=a.Account_ID 
                        LEFT JOIN BILLIEN_MAA.Customers c ON a.Customer_ID=c.Customer_ID
                        LEFT JOIN BILLIEN_MAA.POSES p ON i.POS_NUMBER = p.POS_NUMBER
                        WHERE fp.VS IS NULL AND (i.DATE_OF_PAYMENT >= Trunc(to_date(&DATUM, 'dd.mm.yyyy')) And i.DATE_OF_PAYMENT < Trunc(to_date(&DATUM, 'dd.mm.yyyy')+1)) 
                        AND i.Payment_Method_Code = 20
                        and i.VARIABLE_SYMBOL not like to_char(sysdate,'YY%') 
                        and i.VARIABLE_SYMBOL not like to_char( add_months( sysdate, -12 )-1, 'YY%')
                        
                        UNION ALL

                        SELECT
                        i.VARIABLE_SYMBOL InvoiceNumber, 
                        to_char(c.Customer_Number,'FM000000000009') ClientNumber,
                        to_char(a.Account_Number,'FM000000000009') AccountNumber,
                        i.Payment_Type_Name InvoiceType,  
                        i.Payment_Status_Name InvoiceStatus, 
                        i.Payment_Method_Name PaymentType, 
                        i.Payment_Amount AS Amount,  
                        i.DATE_OF_PAYMENT AS DatePaid,
                        CASE WHEN i.POS_OPERATOR_USERNAME IS NOT NULL THEN i.POS_OPERATOR_USERNAME ELSE i.WORKSTATION_TYPE_NAME END AS UserLoginCreate,
                        i.POS_NUMBER,
                        p.RETAIL_PARTNER_ABBR AS Location,
                        fp.VS,
                        to_char(fp.DATE_TIME,'yyyy.mm.dd hh24:mi:ss') AS DateTime,
                        to_char(fp.CARD_NUMBER) AS CARD_NUMBER,
                        fp.Amount AS FP_Amount,
                        to_char(fp.AUTH_CODE) AS AuthCode,
                        fp.Result AS Result,
                        fp.Terminal AS Terminal,
                        NULL AS Terminal_PABK
                        FROM BILLIEN_MAA.Payments i   
                        LEFT JOIN REP_CRM.T_FP_FCI_CARD_TRANSACTIONS fp ON i.VARIABLE_SYMBOL = fp.VS
                        LEFT JOIN BILLIEN_MAA.Accounts a ON i.Account_ID=a.Account_ID 
                        LEFT JOIN BILLIEN_MAA.Customers c ON a.Customer_ID=c.Customer_ID
                        LEFT JOIN BILLIEN_MAA.POSES p ON i.POS_NUMBER = p.POS_NUMBER
                        WHERE fp.VS IS NULL AND (i.DATE_OF_PAYMENT >= Trunc(to_date(&DATUM, 'dd.mm.yyyy')) And i.DATE_OF_PAYMENT < Trunc(to_date(&DATUM, 'dd.mm.yyyy')+1)) 
                        AND i.Payment_Method_Code = 30
                        and i.VARIABLE_SYMBOL not like to_char(sysdate,'YY%') 
                        and i.VARIABLE_SYMBOL not like to_char( add_months( sysdate, -12 )-1, 'YY%')
                        
                        UNION ALL
                        -- bank kartou platene SKT faktury na EMS, ktore maju VS zacinajuece LIKE 'YY%' ako v EDZ (IM309239)
                        SELECT
                        i.VARIABLE_SYMBOL InvoiceNumber,
                        to_char(c.Customer_Number,'FM000000000009') ClientNumber,
                        to_char(a.Account_Number,'FM000000000009') AccountNumber,
                        i.Payment_Type_Name InvoiceType,  
                        i.Payment_Status_Name InvoiceStatus, 
                        i.Payment_Method_Name PaymentType, 
                        i.Payment_Amount AS Amount,  
                        i.DATE_OF_PAYMENT AS DatePaid,
                        CASE WHEN i.POS_OPERATOR_USERNAME IS NOT NULL THEN i.POS_OPERATOR_USERNAME ELSE i.WORKSTATION_TYPE_NAME END AS UserLoginCreate,
                        i.POS_NUMBER,
                        p.RETAIL_PARTNER_ABBR AS Location,
                        fp.VS,
                        to_char(fp.DATE_TIME,'yyyy.mm.dd hh24:mi:ss') AS DateTime,
                        to_char(fp.CARD_NUMBER) AS CARD_NUMBER,
                        fp.Amount AS FP_Amount,
                        to_char(fp.AUTH_CODE) AS AuthCode,
                        fp.Result AS Result,
                        fp.Terminal AS Terminal,
                        fp.Terminal_PABK AS Terminal_PABK
                        FROM REP_CRM.T_FP_CARD_TRANSACTIONS fp  
                        LEFT JOIN BILLIEN_MAA.Payments i ON i.VARIABLE_SYMBOL=fp.VS
                        LEFT JOIN BILLIEN_MAA.Accounts a ON i.Account_ID=a.Account_ID 
                        LEFT JOIN BILLIEN_MAA.Customers c ON a.Customer_ID=c.Customer_ID
                        LEFT JOIN BILLIEN_MAA.POSES p ON i.POS_NUMBER = p.POS_NUMBER
                        WHERE (fp.date_time >= Trunc(to_date(&DATUM, 'dd.mm.yyyy')) And fp.date_time < Trunc(to_date(&DATUM, 'dd.mm.yyyy')+1)) 
                        AND (fp.VS IS NOT NULL OR i.Payment_Method_Code=20) AND fp.TERMINAL != 'POSTPAY'
                        and (fp.VS like to_char(sysdate,'YY%') or fp.VS like to_char( add_months( sysdate, -12 )-1, 'YY%'))


                        UNION ALL
                        -- fleet kartou platene SKT faktury na EMS, ktore maju VS zacinajuece LIKE 'YY%' ako v EDZ (IM309239)
                        SELECT
                        i.VARIABLE_SYMBOL InvoiceNumber, 
                        to_char(c.Customer_Number,'FM000000000009') ClientNumber,
                        to_char(a.Account_Number,'FM000000000009') AccountNumber,
                        i.Payment_Type_Name InvoiceType,  
                        i.Payment_Status_Name InvoiceStatus, 
                        i.Payment_Method_Name PaymentType, 
                        i.Payment_Amount AS Amount,  
                        i.DATE_OF_PAYMENT AS DatePaid,
                        CASE WHEN i.POS_OPERATOR_USERNAME IS NOT NULL THEN i.POS_OPERATOR_USERNAME ELSE i.WORKSTATION_TYPE_NAME END AS UserLoginCreate,
                        i.POS_NUMBER,
                        p.RETAIL_PARTNER_ABBR AS Location,
                        fp.VS,
                        to_char(fp.DATE_TIME,'yyyy.mm.dd hh24:mi:ss') AS DateTime,
                        to_char(fp.CARD_NUMBER) AS CARD_NUMBER,
                        fp.Amount AS FP_Amount,
                        to_char(fp.AUTH_CODE) AS AuthCode,
                        fp.Result AS Result,
                        fp.Terminal AS Terminal,
                        NULL AS Terminal_PABK
                        FROM REP_CRM.T_FP_FCI_CARD_TRANSACTIONS fp  
                        LEFT JOIN BILLIEN_MAA.Payments i ON i.VARIABLE_SYMBOL=fp.VS
                        LEFT JOIN BILLIEN_MAA.Accounts a ON i.Account_ID=a.Account_ID 
                        LEFT JOIN BILLIEN_MAA.Customers c ON a.Customer_ID=c.Customer_ID
                        LEFT JOIN BILLIEN_MAA.POSES p ON i.POS_NUMBER = p.POS_NUMBER
                        WHERE (fp.date_time >= Trunc(to_date(&DATUM, 'dd.mm.yyyy')) And fp.date_time < Trunc(to_date(&DATUM, 'dd.mm.yyyy')+1)) 
                        AND (fp.VS IS NOT NULL OR i.Payment_Method_Code=30) AND fp.TERMINAL != 'POSTPAY'
                        and (fp.VS like to_char(sysdate,'YY%') or fp.VS like to_char( add_months( sysdate, -12 )-1, 'YY%'))

                        UNION ALL
                        
                        SELECT
                        i.VARIABLE_SYMBOL InvoiceNumber,
                        'eDZ' ClientNumber,
                        'eDZ' AccountNumber,
                        i.Payment_Type_Name InvoiceType,  
                        i.Payment_Status_Name InvoiceStatus, 
                        i.Payment_Method_Name PaymentType, 
                        i.Payment_Amount AS Amount,  
                        i.DATE_OF_PAYMENT AS DatePaid,
                        CASE WHEN i.POS_OPERATOR_USERNAME IS NOT NULL THEN i.POS_OPERATOR_USERNAME ELSE i.WORKSTATION_TYPE_NAME END AS UserLoginCreate,
                        i.POS_NUMBER,
                        p.RETAIL_PARTNER_ABBR AS Location,
                        fp.VS,
                        to_char(fp.DATE_TIME,'yyyy.mm.dd hh24:mi:ss') AS DateTime,
                        to_char(fp.CARD_NUMBER) AS CARD_NUMBER,
                        fp.Amount AS FP_Amount,
                        to_char(fp.AUTH_CODE) AS AuthCode,
                        fp.Result AS Result,
                        fp.Terminal AS Terminal,
                        fp.Terminal_PABK AS Terminal_PABK
                        FROM REP_CRM.T_FP_CARD_TRANSACTIONS fp
                        LEFT JOIN EDZ_MAA.Payments i ON i.VARIABLE_SYMBOL=fp.VS
                        LEFT JOIN EDZ_MAA.POSES p ON i.POS_NUMBER = p.POS_NUMBER
                        WHERE (fp.date_time >= Trunc(to_date(&DATUM, 'dd.mm.yyyy')) And fp.date_time < Trunc(to_date(&DATUM, 'dd.mm.yyyy')+1)) 
                        and (fp.VS like to_char(sysdate,'YY%') or fp.VS like to_char( add_months( sysdate, -12 )-1, 'YY%'))
                        and fp.terminal != 'POSTPAY'

                        UNION ALL

                        SELECT
                        i.VARIABLE_SYMBOL InvoiceNumber,
                        'eDZ' ClientNumber,
                        'eDZ' AccountNumber,
                        i.Payment_Type_Name InvoiceType,  
                        i.Payment_Status_Name InvoiceStatus, 
                        i.Payment_Method_Name PaymentType, 
                        i.Payment_Amount AS Amount,  
                        i.DATE_OF_PAYMENT AS DatePaid,
                        CASE WHEN i.POS_OPERATOR_USERNAME IS NOT NULL THEN i.POS_OPERATOR_USERNAME ELSE i.WORKSTATION_TYPE_NAME END AS UserLoginCreate,
                        i.POS_NUMBER,
                        p.RETAIL_PARTNER_ABBR AS Location,
                        fp.VS,
                        to_char(fp.DATE_TIME,'yyyy.mm.dd hh24:mi:ss') AS DateTime,
                        to_char(fp.CARD_NUMBER) AS CARD_NUMBER,
                        fp.Amount AS FP_Amount,
                        to_char(fp.AUTH_CODE) AS AuthCode,
                        fp.Result AS Result,
                        fp.Terminal AS Terminal,
                        NULL AS Terminal_PABK
                        FROM REP_CRM.T_FP_FCI_CARD_TRANSACTIONS fp
                        LEFT JOIN EDZ_MAA.Payments i ON i.VARIABLE_SYMBOL=fp.VS
                        LEFT JOIN EDZ_MAA.POSES p ON i.POS_NUMBER = p.POS_NUMBER
                        WHERE (fp.date_time >= Trunc(to_date(&DATUM, 'dd.mm.yyyy')) And fp.date_time < Trunc(to_date(&DATUM, 'dd.mm.yyyy')+1)) 
                        and (fp.VS like to_char(sysdate,'YY%') or fp.VS like to_char( add_months( sysdate, -12 )-1, 'YY%'))
                        and fp.terminal != 'POSTPAY'

                        UNION ALL

                        SELECT
                        i.VARIABLE_SYMBOL InvoiceNumber,
                        'eDZ' ClientNumber,
                        'eDZ' AccountNumber,
                        i.Payment_Type_Name InvoiceType,  
                        i.Payment_Status_Name InvoiceStatus, 
                        i.Payment_Method_Name PaymentType, 
                        i.Payment_Amount AS Amount,  
                        i.DATE_OF_PAYMENT AS DatePaid,
                        CASE WHEN i.POS_OPERATOR_USERNAME IS NOT NULL THEN i.POS_OPERATOR_USERNAME ELSE i.WORKSTATION_TYPE_NAME END AS UserLoginCreate,
                        i.POS_NUMBER,
                        p.RETAIL_PARTNER_ABBR AS Location,
                        fp.VS,
                        to_char(fp.DATE_TIME,'yyyy.mm.dd hh24:mi:ss') AS DateTime,
                        to_char(fp.CARD_NUMBER) AS CARD_NUMBER,
                        fp.Amount AS FP_Amount,
                        to_char(fp.AUTH_CODE) AS AuthCode,
                        fp.Result AS Result,
                        fp.Terminal AS Terminal,
                        fp.Terminal_PABK AS Terminal_PABK
                        FROM EDZ_MAA.Payments i   
                        LEFT JOIN REP_CRM.T_FP_CARD_TRANSACTIONS fp ON i.VARIABLE_SYMBOL = fp.VS
                        LEFT JOIN EDZ_MAA.POSES p ON i.POS_NUMBER = p.POS_NUMBER
                        WHERE fp.VS IS NULL AND (i.DATE_OF_PAYMENT >= Trunc(to_date(&DATUM, 'dd.mm.yyyy')) And i.DATE_OF_PAYMENT < Trunc(to_date(&DATUM, 'dd.mm.yyyy')+1)) 

                        AND i.Payment_Method_Code = 20
                        and (i.VARIABLE_SYMBOL like to_char(sysdate,'YY%') or i.VARIABLE_SYMBOL like to_char( add_months( sysdate, -12 )-1, 'YY%'))
                        """

            # Nahraďte DEFINE DATUM hodnotou z formulára
            sql_query = sql_query.replace("&DATUM", f"'{report_date}'")
            # print(sql_query)

            # Vykonajte dotaz
            cur.execute(sql_query)
            rows = cur.fetchall()
            columns = [i[0] for i in cur.description]  # Získajte názvy stĺpcov

            # Vytvorte XLSX súbor
            from openpyxl import Workbook
            import os
            from datetime import datetime

            # Vytvorte workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Report"

            # Pridajte hlavičku
            ws.append(columns)

            # Pridajte dáta
            for row in rows:
                ws.append(row)

            # Uložte do UPLOAD_DIR
            report_date_obj = datetime.strptime(report_date, '%d.%m.%Y')  # Konverzia na dátumový objekt
            filename = f"Invoices_card_payments_{report_date_obj.strftime('%Y%m%d')}.xlsx"
            filepath = os.path.join(UPLOAD_DIR, filename)
            
            wb.save(filepath)
            message += f'📊 Report pre dátum {report_date} bol uložený ako {filename}.<br>'
            
        except ValueError:
            message += '❌ Nesprávny formát dátumu. Použi DD.MM.YYYY<br>'
        except Exception as e:
            message += f'❌ Chyba pri generovaní reportu: {str(e)}<br>'

    # === Upload posledného xlsx súboru na SharePoint ===
    if request.method == 'POST' and 'upload_shp' in request.POST:
        import json
        import os
        import re
        from django.conf import settings
        from datetime import datetime
        from office365.runtime.auth.authentication_context import AuthenticationContext
        from office365.sharepoint.client_context import ClientContext

        try:
            # Nastavenie cesty k upload priečinku
            UPLOAD_DIR = os.path.join(settings.PAYWELL_ROOT, 'uploads')
            os.makedirs(UPLOAD_DIR, exist_ok=True)

            # Nájdi najnovší .xlsx súbor
            xlsx_files = [f for f in os.listdir(UPLOAD_DIR) if f.endswith('.xlsx')]
            if not xlsx_files:
                message += '❌ Žiadny .xlsx súbor nebol nájdený v priečinku uploads.<br>'
                raise Exception("XLSX not found")

            xlsx_files.sort()
            filename = xlsx_files[-1]
            filepath = os.path.join(UPLOAD_DIR, filename)

            # Extrahuj dátum zo súboru
            match = re.search(r'Invoices_card_payments_(\d{8})\.xlsx', filename)
            if not match:
                message += '❌ Názov súboru nie je v správnom formáte: Invoices_card_payments_YYYYMMDD.xlsx<br>'
                raise Exception("Invalid filename format")
            date_str = match.group(1)
            report_date_obj = datetime.strptime(date_str, '%Y%m%d')

            # Načítaj SharePoint prihlasovacie údaje zo JSON
            BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            json_path = os.path.join(BASE_DIR, 'edz', 'static', 'json', 'jobs.json')

            def load_cred(field):
                with open(json_path, 'r', encoding='utf-8') as file:
                    data = json.load(file)
                    return [entry['fields'][field] for entry in data if entry['model'] == 'EDZ_Podania'][0]

            username = load_cred('username')
            password = load_cred('password')

            # Nastav prostredie bez proxy
            os.environ['HTTP_PROXY'] = ''
            os.environ['HTTPS_PROXY'] = ''
            os.environ['NO_PROXY'] = 'login.microsoftonline.com,sharepoint.com'

            # Vytvor relatívnu cestu podľa dátumu
            current_year = report_date_obj.strftime('%Y')
            current_month = report_date_obj.strftime('%Y_%m')
            folder_date = report_date_obj.strftime('%Y_%m_%d')

            site_url = 'https://example.sharepoint.com/teams/coo-infra/5230/'
            sharepoint_folder_path = f'/teams/coo-infra/5230//Reporty/denné/{current_year}/{current_month}/{folder_date}'
            # server_relative_url = f'/teams/coo-infra/5230/{sharepoint_folder_path}'

            # Autentifikácia a nahratie súboru
            ctx_auth = AuthenticationContext(site_url)
            if ctx_auth.acquire_token_for_user(username, password):
                ctx = ClientContext(site_url, ctx_auth)
                with open(filepath, 'rb') as file_obj:
                    file_content = file_obj.read()
                    folder = ctx.web.get_folder_by_server_relative_url(sharepoint_folder_path)
                    upload_result = folder.upload_file(filename, file_content).execute_query()
                    message += f'☁️ Súbor <b>{filename}</b> bol nahraný na SharePoint: <code>{upload_result.serverRelativeUrl}</code><br>'
            else:
                message += f'❌ Nepodarilo sa autentifikovať na SharePoint pomocou poskytnutých údajov.<br>'

        except Exception as e:
            message += f'❌ Chyba pri nahrávaní na SharePoint: {str(e)}<br>'

    # === 4. Načítanie súborov ===
    import os
    from django.conf import settings
    
    # === Načítanie posledných riadkov z auto.log ===
    log_lines = []
    log_file_path = os.path.join(settings.PAYWELL_ROOT, 'auto.log')
    try:
        with open(log_file_path, 'r', encoding='utf-8') as f:
            all_lines = f.readlines()
            log_lines = all_lines[-40:]  # Zobrazí posledných 20 riadkov
    except FileNotFoundError:
        log_lines = ["Žiadny log ešte neexistuje."]
    except Exception as e:
        log_lines = [f"Chyba pri čítaní logu: {str(e)}"]
        

    # Definícia UPLOAD_DIR
    UPLOAD_DIR = os.path.join(settings.PAYWELL_ROOT, 'uploads')
    os.makedirs(UPLOAD_DIR, exist_ok=True)


    try:
        files = os.listdir(UPLOAD_DIR)
    except FileNotFoundError:
        files = []

    # === Cesta k jobs.json ===
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    JOBS_JSON_PATH = os.path.join(BASE_DIR, 'work', 'static', 'json', 'jobs.json')
    import os
    import json

    # === Inicializuj hodnoty ===
    current_run = {
        "hour": "*",
        "minute": "*",
        "day": "*",
        "month": "*"
    }

    # === Načítaj JSON do premennej `data` ===
    try:
        with open(JOBS_JSON_PATH, 'r', encoding='utf-8') as f:
            data = json.load(f)
            for entry in data:
                if entry.get('model') == 'paywell':
                    run_raw = entry.get('fields', {}).get('run', '{}')
                    try:
                        # run je serializovaný JSON string → treba ho deserializovať
                        current_run = json.loads(run_raw)
                    except json.JSONDecodeError as e:
                        message += f"❌ Chyba pri dekódovaní schedulera: {e}<br>"
                    break
    except Exception as e:
        message += f"❌ Chyba pri načítaní jobs.json: {e}<br>"
        data = []


    # === Spracovanie zmeny schedulera cez POST ===
    if request.method == 'POST' and request.POST.get('update_scheduler'):
        new_run = {
            "hour": request.POST.get("hour", "*"),
            "minute": request.POST.get("minute", "*"),
            "day": request.POST.get("day", "*"),
            "month": request.POST.get("month", "*"),
        }
        try:
            updated = False
            for entry in data:
                if entry.get("model") == "paywell":
                    entry["fields"]["run"] = json.dumps(new_run)
                    updated = True
                    break
            if updated:
                with open(JOBS_JSON_PATH, 'w', encoding='utf-8') as f:
                    json.dump(data, f, indent=4, ensure_ascii=False)
                message += "✅ Nastavenie schedulera bolo aktualizované.<br>"
                current_run = new_run
                from scheduler import reload_job_if_paywell
                # Reload job pre EETS_Lego
                reload_job_if_paywell('paywell', 'work/static/json/jobs.json')
            else:
                message += "❌ Nenašiel sa záznam pre 'paywell'.<br>"
        except Exception as e:
            message += f"❌ Chyba pri ukladaní nastavení: {e}<br>"

    # === Čítanie auto.log ===
    try:
        UPLOAD_DIR = os.path.join(settings.PAYWELL_ROOT)
        log_path = os.path.join(UPLOAD_DIR, 'auto.log')
        if os.path.exists(log_path):
            with open(log_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
                log_lines = lines[-30:]  # posledných N riadkov
    except Exception as e:
        log_lines = [f"⚠️ Chyba pri čítaní logu: {e}"]

    # === Súbory v uploads/ ===
    uploads_path = os.path.join(settings.PAYWELL_ROOT, 'uploads')
    try:
        files = os.listdir(uploads_path)

    except FileNotFoundError:
        files = []
    
    # === Vykreslenie stránky ===
    return render(request, 'ems/paywell.html', {
        'files': files,
        'message': message,
        'transactions_summary': transactions_summary,
        'log_lines': log_lines,
        'current_run': current_run,
    })



def download_sql(request, filename):
    logger.info("------------ START stiahnutie suboru " )
    file_path = os.path.join(UPLOAD_DIR, filename)
    if os.path.isfile(file_path):
        return FileResponse(open(file_path, 'rb'), as_attachment=True, filename=filename)
    raise Http404("Súbor neexistuje.")

def download_auto_log(request):
    log_file_path = os.path.join(settings.PAYWELL_ROOT, 'auto.log')
    if os.path.isfile(log_file_path):
        return FileResponse(open(log_file_path, 'rb'), as_attachment=True, filename='auto.log')
    raise Http404("Log súbor neexistuje.")

def paywell_scheduler():
    import os
    import json
    import zipfile
    from datetime import datetime, timedelta
    from django.conf import settings
    import re
    from django.conf import settings
    from datetime import datetime
    from office365.runtime.auth.authentication_context import AuthenticationContext
    from office365.sharepoint.client_context import ClientContext

    message = ''
    pswd = '936BhU14'
    transactions_summary = {
        'T_FP_CARD_TRANSACTIONS': [],
        'T_FP_FCI_CARD_TRANSACTIONS': []
    }
    insert_lines = []

    # === Cesty ===
    UPLOAD_DIR0 = os.path.join(settings.PAYWELL_ROOT)
    UPLOAD_DIR = os.path.join(settings.PAYWELL_ROOT, 'uploads')
    os.makedirs(UPLOAD_DIR, exist_ok=True)

    # === LOGOVANIE ===
    log_file_path = os.path.join(UPLOAD_DIR0, 'auto.log')

    def log(msg):
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        with open(log_file_path, 'a', encoding='utf-8') as log_file:
            log_file.write(f"[{timestamp}] {msg}\n")

    log("🔄 paywell_scheduler spustený.")

    # === Vytvorenie názvu SQL súboru ===
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    sql_file_name = f"insert_example_all_{timestamp}.sql"
    sql_file_path = os.path.join(UPLOAD_DIR, sql_file_name)

    # === Odzipovanie ZIP súborov ===
    zip_total = 0
    zip_success = 0
    zip_failed = 0

    for filename in os.listdir(UPLOAD_DIR):
        if filename.lower().endswith('.zip'):
            zip_total += 1
            zip_path = os.path.join(UPLOAD_DIR, filename)
            try:
                with zipfile.ZipFile(zip_path) as zip_file:
                    zip_file.setpassword(pswd.encode('utf-8'))
                    zip_file.extractall(UPLOAD_DIR)
                    msg = f'✅ ZIP "{filename}" bol úspešne odzipovaný.'
                    zip_success += 1
            except RuntimeError as e:
                msg = f'❌ Chyba pri odzipovaní "{filename}": {e}'
                zip_failed += 1
            except Exception as e:
                msg = f'❌ Neočakávaná chyba so súborom "{filename}": {e}'
                zip_failed += 1

            message += msg + '<br>'
            log(msg)

    # === Sumár ===
    summary = f"🔚 Odzipovanie hotové – celkom: {zip_total}, úspešné: {zip_success}, neúspešné: {zip_failed}"
    message += summary + '<br>'
    log(summary)

    # B. Prejdi všetky .xml a vyber INSERTy
    for xml_file in os.listdir(UPLOAD_DIR):
        #log(f"------------ START priprava insertov pre {xml_file} ------------")
        if xml_file.lower().endswith('.xml'):
            xml_path = os.path.join(UPLOAD_DIR, xml_file)
            try:
                tree = ET.parse(xml_path)
                text = ET.tostring(tree.getroot(), encoding='utf-8', method='text').decode('utf-8')
                insert_lines += [
                    line.strip() for line in text.splitlines() if line.strip().upper().startswith("INSERT INTO")
                ]
                log(f"✅ Úspešne spracovaný XML súbor: {xml_file}, nájdené INSERTy: {len(insert_lines)}")
            except Exception as e:
                message += f'⚠️ Chyba pri XML "{xml_file}": {e}<br>'
                log(f'⚠️ Chyba pri XML "{xml_file}": {e}')
    
    # C. Vytvor SQL súbor s INSERTmi
    if insert_lines:
        log("------------ START vytvorenie SQL súboru ------------")
        try:
            with open(sql_file_path, 'w', encoding='utf-8') as sql_out:
                sql_out.write('\n'.join(insert_lines))
                sql_out.write('\nCOMMIT;\n')
            msg = f'📄 INSERTy boli uložené do "{sql_file_name}".'
            message += msg + '<br>'
            log(msg)
        except Exception as e:
            msg = f'❌ Chyba pri zápise SQL: {e}'
            message += msg + '<br>'
            log(msg)
    else:
        msg = '⚠️ Neboli nájdené INSERT príkazy v žiadnom XML súbore.'
        message += msg + '<br>'
        log(msg)

    # D. Import SQL do databázy
    msg = '🔵 Start importu do DB...'
    message += msg + '<br>'
    log(msg)

    try:
        sql_files = [f for f in os.listdir(UPLOAD_DIR) if f.endswith('.sql')]
        if sql_files:
            sql_files.sort()
            file_path = os.path.join(UPLOAD_DIR, sql_files[-1])

            with open(file_path, 'r', encoding='utf-8') as f:
                full_sql = f.read()

            statements = [s.strip() for s in full_sql.split(';') if s.strip()]
            EMS_DWH = EMS_DWHDB.objects.get(user=48)
            con = cx_Oracle.connect(
                f"{EMS_DWH.EMS_DWHDB_username}/{EMS_DWH.EMS_DWHDB_password}@{EMS_DWH.EMS_DWHDB_hostname}:{EMS_DWH.EMS_DWHDB_port}/{EMS_DWH.EMS_DWHDB_servicename}"
            )
            cur = con.cursor()

            msg = f'✅ Pripojenie do DWH.'
            message += msg + '<br>'
            log(msg)

            msg = f'❓ Spustený import SQL súboru "{os.path.basename(file_path)}" do DWH.'
            message += msg + '<br>'
            log(msg)

            msg = f'❓ Počet príkazov: {len(statements)}'
            message += msg + '<br>'
            log(msg)

            for stmt in statements:
                try:
                    cur.execute(stmt)
                except Exception as e:
                    err_msg = f'❌ Chyba pri vykonávaní príkazu: {e}'
                    message += err_msg + '<br>'
                    log(err_msg)

            con.commit()
            cur.close()
            con.close()

            msg = f'✅ SQL súbor "{os.path.basename(file_path)}" bol importovaný do DB.'
            message += msg + '<br>'
            log(msg)
        else:
            msg = "❌ Nebol nájdený žiadny .sql súbor na import."
            message += msg + '<br>'
            log(msg)
    except Exception as e:
        msg = f"❌ Chyba importu do DB: {e}"
        message += msg + '<br>'
        log(msg)

    # E. Generovanie reportu
    msg = "------------ START generovanie reportu ------------"
    log(msg)

    try:
        # Formát: 16.05.2024
        # Automatický dátum: včerajší deň
        report_date_obj = datetime.now() - timedelta(days=1)
        report_date = report_date_obj.strftime('%d.%m.%Y')


        EMS_DWH = EMS_DWHDB.objects.get(user=48)
        con = cx_Oracle.connect(
            f"{EMS_DWH.EMS_DWHDB_username}/{EMS_DWH.EMS_DWHDB_password}@{EMS_DWH.EMS_DWHDB_hostname}:{EMS_DWH.EMS_DWHDB_port}/{EMS_DWH.EMS_DWHDB_servicename}"
        )
        cur = con.cursor()

        # SQL dotaz (prázdny v príklade – doplň podľa potreby)
        sql_query = """
                        SELECT
                        i.VARIABLE_SYMBOL InvoiceNumber,
                        to_char(c.Customer_Number,'FM000000000009') ClientNumber,
                        to_char(a.Account_Number,'FM000000000009') AccountNumber,
                        i.Payment_Type_Name InvoiceType,  
                        i.Payment_Status_Name InvoiceStatus, 
                        i.Payment_Method_Name PaymentType, 
                        i.Payment_Amount AS Amount,  
                        i.DATE_OF_PAYMENT AS DatePaid,
                        CASE WHEN i.POS_OPERATOR_USERNAME IS NOT NULL THEN i.POS_OPERATOR_USERNAME ELSE i.WORKSTATION_TYPE_NAME END AS UserLoginCreate,
                        i.POS_NUMBER,
                        p.RETAIL_PARTNER_ABBR AS Location,
                        fp.VS,
                        to_char(fp.DATE_TIME,'yyyy.mm.dd hh24:mi:ss') AS DateTime,
                        to_char(fp.CARD_NUMBER) AS CARD_NUMBER,
                        fp.Amount AS FP_Amount,
                        to_char(fp.AUTH_CODE) AS AuthCode,
                        fp.Result AS Result,
                        fp.Terminal AS Terminal,
                        fp.Terminal_PABK AS Terminal_PABK
                        FROM REP_CRM.T_FP_CARD_TRANSACTIONS fp  
                        LEFT JOIN BILLIEN_MAA.Payments i ON i.VARIABLE_SYMBOL=fp.VS
                        LEFT JOIN BILLIEN_MAA.Accounts a ON i.Account_ID=a.Account_ID 
                        LEFT JOIN BILLIEN_MAA.Customers c ON a.Customer_ID=c.Customer_ID
                        LEFT JOIN BILLIEN_MAA.POSES p ON i.POS_NUMBER = p.POS_NUMBER
                        WHERE (fp.date_time >= Trunc(to_date(&DATUM, 'dd.mm.yyyy')) And fp.date_time < Trunc(to_date(&DATUM, 'dd.mm.yyyy')+1)) 
                        AND (fp.VS IS NOT NULL OR i.Payment_Method_Code=20) AND fp.TERMINAL != 'POSTPAY'
                        and fp.VS not like to_char(sysdate,'YY%') 
                        and fp.VS not like to_char( add_months( sysdate, -12 )-1, 'YY%')

                        UNION ALL

                        SELECT
                        i.VARIABLE_SYMBOL InvoiceNumber, 
                        to_char(c.Customer_Number,'FM000000000009') ClientNumber,
                        to_char(a.Account_Number,'FM000000000009') AccountNumber,
                        i.Payment_Type_Name InvoiceType,  
                        i.Payment_Status_Name InvoiceStatus, 
                        i.Payment_Method_Name PaymentType, 
                        i.Payment_Amount AS Amount,  
                        i.DATE_OF_PAYMENT AS DatePaid,
                        CASE WHEN i.POS_OPERATOR_USERNAME IS NOT NULL THEN i.POS_OPERATOR_USERNAME ELSE i.WORKSTATION_TYPE_NAME END AS UserLoginCreate,
                        i.POS_NUMBER,
                        p.RETAIL_PARTNER_ABBR AS Location,
                        fp.VS,
                        to_char(fp.DATE_TIME,'yyyy.mm.dd hh24:mi:ss') AS DateTime,
                        to_char(fp.CARD_NUMBER) AS CARD_NUMBER,
                        fp.Amount AS FP_Amount,
                        to_char(fp.AUTH_CODE) AS AuthCode,
                        fp.Result AS Result,
                        fp.Terminal AS Terminal,
                        NULL AS Terminal_PABK
                        FROM REP_CRM.T_FP_FCI_CARD_TRANSACTIONS fp  
                        LEFT JOIN BILLIEN_MAA.Payments i ON i.VARIABLE_SYMBOL=fp.VS
                        LEFT JOIN BILLIEN_MAA.Accounts a ON i.Account_ID=a.Account_ID 
                        LEFT JOIN BILLIEN_MAA.Customers c ON a.Customer_ID=c.Customer_ID
                        LEFT JOIN BILLIEN_MAA.POSES p ON i.POS_NUMBER = p.POS_NUMBER
                        WHERE (fp.date_time >= Trunc(to_date(&DATUM, 'dd.mm.yyyy')) And fp.date_time < Trunc(to_date(&DATUM, 'dd.mm.yyyy')+1)) 
                        AND (fp.VS IS NOT NULL OR i.Payment_Method_Code=30) AND fp.TERMINAL != 'POSTPAY'
                        and fp.VS not like to_char(sysdate,'YY%') 
                        and fp.VS not like to_char( add_months( sysdate, -12 )-1, 'YY%')
                        
                        UNION ALL

                        SELECT
                        i.VARIABLE_SYMBOL InvoiceNumber, 
                        to_char(c.Customer_Number,'FM000000000009') ClientNumber,
                        to_char(a.Account_Number,'FM000000000009') AccountNumber,
                        i.Payment_Type_Name InvoiceType,  
                        i.Payment_Status_Name InvoiceStatus, 
                        i.Payment_Method_Name PaymentType, 
                        i.Payment_Amount AS Amount,  
                        i.DATE_OF_PAYMENT AS DatePaid,
                        CASE WHEN i.POS_OPERATOR_USERNAME IS NOT NULL THEN i.POS_OPERATOR_USERNAME ELSE i.WORKSTATION_TYPE_NAME END AS UserLoginCreate,
                        i.POS_NUMBER,
                        p.RETAIL_PARTNER_ABBR AS Location,
                        fp.VS,
                        to_char(fp.DATE_TIME,'yyyy.mm.dd hh24:mi:ss') AS DateTime,
                        to_char(fp.CARD_NUMBER) AS CARD_NUMBER,
                        fp.Amount AS FP_Amount,
                        to_char(fp.AUTH_CODE) AS AuthCode,
                        fp.Result AS Result,
                        fp.Terminal AS Terminal,
                        fp.Terminal_PABK AS Terminal_PABK
                        FROM BILLIEN_MAA.Payments i   
                        LEFT JOIN REP_CRM.T_FP_CARD_TRANSACTIONS fp ON i.VARIABLE_SYMBOL = fp.VS
                        LEFT JOIN BILLIEN_MAA.Accounts a ON i.Account_ID=a.Account_ID 
                        LEFT JOIN BILLIEN_MAA.Customers c ON a.Customer_ID=c.Customer_ID
                        LEFT JOIN BILLIEN_MAA.POSES p ON i.POS_NUMBER = p.POS_NUMBER
                        WHERE fp.VS IS NULL AND (i.DATE_OF_PAYMENT >= Trunc(to_date(&DATUM, 'dd.mm.yyyy')) And i.DATE_OF_PAYMENT < Trunc(to_date(&DATUM, 'dd.mm.yyyy')+1)) 
                        AND i.Payment_Method_Code = 20
                        and i.VARIABLE_SYMBOL not like to_char(sysdate,'YY%') 
                        and i.VARIABLE_SYMBOL not like to_char( add_months( sysdate, -12 )-1, 'YY%')
                        
                        UNION ALL

                        SELECT
                        i.VARIABLE_SYMBOL InvoiceNumber, 
                        to_char(c.Customer_Number,'FM000000000009') ClientNumber,
                        to_char(a.Account_Number,'FM000000000009') AccountNumber,
                        i.Payment_Type_Name InvoiceType,  
                        i.Payment_Status_Name InvoiceStatus, 
                        i.Payment_Method_Name PaymentType, 
                        i.Payment_Amount AS Amount,  
                        i.DATE_OF_PAYMENT AS DatePaid,
                        CASE WHEN i.POS_OPERATOR_USERNAME IS NOT NULL THEN i.POS_OPERATOR_USERNAME ELSE i.WORKSTATION_TYPE_NAME END AS UserLoginCreate,
                        i.POS_NUMBER,
                        p.RETAIL_PARTNER_ABBR AS Location,
                        fp.VS,
                        to_char(fp.DATE_TIME,'yyyy.mm.dd hh24:mi:ss') AS DateTime,
                        to_char(fp.CARD_NUMBER) AS CARD_NUMBER,
                        fp.Amount AS FP_Amount,
                        to_char(fp.AUTH_CODE) AS AuthCode,
                        fp.Result AS Result,
                        fp.Terminal AS Terminal,
                        NULL AS Terminal_PABK
                        FROM BILLIEN_MAA.Payments i   
                        LEFT JOIN REP_CRM.T_FP_FCI_CARD_TRANSACTIONS fp ON i.VARIABLE_SYMBOL = fp.VS
                        LEFT JOIN BILLIEN_MAA.Accounts a ON i.Account_ID=a.Account_ID 
                        LEFT JOIN BILLIEN_MAA.Customers c ON a.Customer_ID=c.Customer_ID
                        LEFT JOIN BILLIEN_MAA.POSES p ON i.POS_NUMBER = p.POS_NUMBER
                        WHERE fp.VS IS NULL AND (i.DATE_OF_PAYMENT >= Trunc(to_date(&DATUM, 'dd.mm.yyyy')) And i.DATE_OF_PAYMENT < Trunc(to_date(&DATUM, 'dd.mm.yyyy')+1)) 
                        AND i.Payment_Method_Code = 30
                        and i.VARIABLE_SYMBOL not like to_char(sysdate,'YY%') 
                        and i.VARIABLE_SYMBOL not like to_char( add_months( sysdate, -12 )-1, 'YY%')
                        
                        UNION ALL
                        -- bank kartou platene SKT faktury na EMS, ktore maju VS zacinajuece LIKE 'YY%' ako v EDZ (IM309239)
                        SELECT
                        i.VARIABLE_SYMBOL InvoiceNumber,
                        to_char(c.Customer_Number,'FM000000000009') ClientNumber,
                        to_char(a.Account_Number,'FM000000000009') AccountNumber,
                        i.Payment_Type_Name InvoiceType,  
                        i.Payment_Status_Name InvoiceStatus, 
                        i.Payment_Method_Name PaymentType, 
                        i.Payment_Amount AS Amount,  
                        i.DATE_OF_PAYMENT AS DatePaid,
                        CASE WHEN i.POS_OPERATOR_USERNAME IS NOT NULL THEN i.POS_OPERATOR_USERNAME ELSE i.WORKSTATION_TYPE_NAME END AS UserLoginCreate,
                        i.POS_NUMBER,
                        p.RETAIL_PARTNER_ABBR AS Location,
                        fp.VS,
                        to_char(fp.DATE_TIME,'yyyy.mm.dd hh24:mi:ss') AS DateTime,
                        to_char(fp.CARD_NUMBER) AS CARD_NUMBER,
                        fp.Amount AS FP_Amount,
                        to_char(fp.AUTH_CODE) AS AuthCode,
                        fp.Result AS Result,
                        fp.Terminal AS Terminal,
                        fp.Terminal_PABK AS Terminal_PABK
                        FROM REP_CRM.T_FP_CARD_TRANSACTIONS fp  
                        LEFT JOIN BILLIEN_MAA.Payments i ON i.VARIABLE_SYMBOL=fp.VS
                        LEFT JOIN BILLIEN_MAA.Accounts a ON i.Account_ID=a.Account_ID 
                        LEFT JOIN BILLIEN_MAA.Customers c ON a.Customer_ID=c.Customer_ID
                        LEFT JOIN BILLIEN_MAA.POSES p ON i.POS_NUMBER = p.POS_NUMBER
                        WHERE (fp.date_time >= Trunc(to_date(&DATUM, 'dd.mm.yyyy')) And fp.date_time < Trunc(to_date(&DATUM, 'dd.mm.yyyy')+1)) 
                        AND (fp.VS IS NOT NULL OR i.Payment_Method_Code=20) AND fp.TERMINAL != 'POSTPAY'
                        and (fp.VS like to_char(sysdate,'YY%') or fp.VS like to_char( add_months( sysdate, -12 )-1, 'YY%'))


                        UNION ALL
                        -- fleet kartou platene SKT faktury na EMS, ktore maju VS zacinajuece LIKE 'YY%' ako v EDZ (IM309239)
                        SELECT
                        i.VARIABLE_SYMBOL InvoiceNumber, 
                        to_char(c.Customer_Number,'FM000000000009') ClientNumber,
                        to_char(a.Account_Number,'FM000000000009') AccountNumber,
                        i.Payment_Type_Name InvoiceType,  
                        i.Payment_Status_Name InvoiceStatus, 
                        i.Payment_Method_Name PaymentType, 
                        i.Payment_Amount AS Amount,  
                        i.DATE_OF_PAYMENT AS DatePaid,
                        CASE WHEN i.POS_OPERATOR_USERNAME IS NOT NULL THEN i.POS_OPERATOR_USERNAME ELSE i.WORKSTATION_TYPE_NAME END AS UserLoginCreate,
                        i.POS_NUMBER,
                        p.RETAIL_PARTNER_ABBR AS Location,
                        fp.VS,
                        to_char(fp.DATE_TIME,'yyyy.mm.dd hh24:mi:ss') AS DateTime,
                        to_char(fp.CARD_NUMBER) AS CARD_NUMBER,
                        fp.Amount AS FP_Amount,
                        to_char(fp.AUTH_CODE) AS AuthCode,
                        fp.Result AS Result,
                        fp.Terminal AS Terminal,
                        NULL AS Terminal_PABK
                        FROM REP_CRM.T_FP_FCI_CARD_TRANSACTIONS fp  
                        LEFT JOIN BILLIEN_MAA.Payments i ON i.VARIABLE_SYMBOL=fp.VS
                        LEFT JOIN BILLIEN_MAA.Accounts a ON i.Account_ID=a.Account_ID 
                        LEFT JOIN BILLIEN_MAA.Customers c ON a.Customer_ID=c.Customer_ID
                        LEFT JOIN BILLIEN_MAA.POSES p ON i.POS_NUMBER = p.POS_NUMBER
                        WHERE (fp.date_time >= Trunc(to_date(&DATUM, 'dd.mm.yyyy')) And fp.date_time < Trunc(to_date(&DATUM, 'dd.mm.yyyy')+1)) 
                        AND (fp.VS IS NOT NULL OR i.Payment_Method_Code=30) AND fp.TERMINAL != 'POSTPAY'
                        and (fp.VS like to_char(sysdate,'YY%') or fp.VS like to_char( add_months( sysdate, -12 )-1, 'YY%'))

                        UNION ALL
                        
                        SELECT
                        i.VARIABLE_SYMBOL InvoiceNumber,
                        'eDZ' ClientNumber,
                        'eDZ' AccountNumber,
                        i.Payment_Type_Name InvoiceType,  
                        i.Payment_Status_Name InvoiceStatus, 
                        i.Payment_Method_Name PaymentType, 
                        i.Payment_Amount AS Amount,  
                        i.DATE_OF_PAYMENT AS DatePaid,
                        CASE WHEN i.POS_OPERATOR_USERNAME IS NOT NULL THEN i.POS_OPERATOR_USERNAME ELSE i.WORKSTATION_TYPE_NAME END AS UserLoginCreate,
                        i.POS_NUMBER,
                        p.RETAIL_PARTNER_ABBR AS Location,
                        fp.VS,
                        to_char(fp.DATE_TIME,'yyyy.mm.dd hh24:mi:ss') AS DateTime,
                        to_char(fp.CARD_NUMBER) AS CARD_NUMBER,
                        fp.Amount AS FP_Amount,
                        to_char(fp.AUTH_CODE) AS AuthCode,
                        fp.Result AS Result,
                        fp.Terminal AS Terminal,
                        fp.Terminal_PABK AS Terminal_PABK
                        FROM REP_CRM.T_FP_CARD_TRANSACTIONS fp
                        LEFT JOIN EDZ_MAA.Payments i ON i.VARIABLE_SYMBOL=fp.VS
                        LEFT JOIN EDZ_MAA.POSES p ON i.POS_NUMBER = p.POS_NUMBER
                        WHERE (fp.date_time >= Trunc(to_date(&DATUM, 'dd.mm.yyyy')) And fp.date_time < Trunc(to_date(&DATUM, 'dd.mm.yyyy')+1)) 
                        and (fp.VS like to_char(sysdate,'YY%') or fp.VS like to_char( add_months( sysdate, -12 )-1, 'YY%'))
                        and fp.terminal != 'POSTPAY'

                        UNION ALL

                        SELECT
                        i.VARIABLE_SYMBOL InvoiceNumber,
                        'eDZ' ClientNumber,
                        'eDZ' AccountNumber,
                        i.Payment_Type_Name InvoiceType,  
                        i.Payment_Status_Name InvoiceStatus, 
                        i.Payment_Method_Name PaymentType, 
                        i.Payment_Amount AS Amount,  
                        i.DATE_OF_PAYMENT AS DatePaid,
                        CASE WHEN i.POS_OPERATOR_USERNAME IS NOT NULL THEN i.POS_OPERATOR_USERNAME ELSE i.WORKSTATION_TYPE_NAME END AS UserLoginCreate,
                        i.POS_NUMBER,
                        p.RETAIL_PARTNER_ABBR AS Location,
                        fp.VS,
                        to_char(fp.DATE_TIME,'yyyy.mm.dd hh24:mi:ss') AS DateTime,
                        to_char(fp.CARD_NUMBER) AS CARD_NUMBER,
                        fp.Amount AS FP_Amount,
                        to_char(fp.AUTH_CODE) AS AuthCode,
                        fp.Result AS Result,
                        fp.Terminal AS Terminal,
                        NULL AS Terminal_PABK
                        FROM REP_CRM.T_FP_FCI_CARD_TRANSACTIONS fp
                        LEFT JOIN EDZ_MAA.Payments i ON i.VARIABLE_SYMBOL=fp.VS
                        LEFT JOIN EDZ_MAA.POSES p ON i.POS_NUMBER = p.POS_NUMBER
                        WHERE (fp.date_time >= Trunc(to_date(&DATUM, 'dd.mm.yyyy')) And fp.date_time < Trunc(to_date(&DATUM, 'dd.mm.yyyy')+1)) 
                        and (fp.VS like to_char(sysdate,'YY%') or fp.VS like to_char( add_months( sysdate, -12 )-1, 'YY%'))
                        and fp.terminal != 'POSTPAY'

                        UNION ALL

                        SELECT
                        i.VARIABLE_SYMBOL InvoiceNumber,
                        'eDZ' ClientNumber,
                        'eDZ' AccountNumber,
                        i.Payment_Type_Name InvoiceType,  
                        i.Payment_Status_Name InvoiceStatus, 
                        i.Payment_Method_Name PaymentType, 
                        i.Payment_Amount AS Amount,  
                        i.DATE_OF_PAYMENT AS DatePaid,
                        CASE WHEN i.POS_OPERATOR_USERNAME IS NOT NULL THEN i.POS_OPERATOR_USERNAME ELSE i.WORKSTATION_TYPE_NAME END AS UserLoginCreate,
                        i.POS_NUMBER,
                        p.RETAIL_PARTNER_ABBR AS Location,
                        fp.VS,
                        to_char(fp.DATE_TIME,'yyyy.mm.dd hh24:mi:ss') AS DateTime,
                        to_char(fp.CARD_NUMBER) AS CARD_NUMBER,
                        fp.Amount AS FP_Amount,
                        to_char(fp.AUTH_CODE) AS AuthCode,
                        fp.Result AS Result,
                        fp.Terminal AS Terminal,
                        fp.Terminal_PABK AS Terminal_PABK
                        FROM EDZ_MAA.Payments i   
                        LEFT JOIN REP_CRM.T_FP_CARD_TRANSACTIONS fp ON i.VARIABLE_SYMBOL = fp.VS
                        LEFT JOIN EDZ_MAA.POSES p ON i.POS_NUMBER = p.POS_NUMBER
                        WHERE fp.VS IS NULL AND (i.DATE_OF_PAYMENT >= Trunc(to_date(&DATUM, 'dd.mm.yyyy')) And i.DATE_OF_PAYMENT < Trunc(to_date(&DATUM, 'dd.mm.yyyy')+1)) 

                        AND i.Payment_Method_Code = 20
                        and (i.VARIABLE_SYMBOL like to_char(sysdate,'YY%') or i.VARIABLE_SYMBOL like to_char( add_months( sysdate, -12 )-1, 'YY%'))

        """

        sql_query = sql_query.replace("&DATUM", f"'{report_date}'")

        cur.execute(sql_query)
        rows = cur.fetchall()
        columns = [i[0] for i in cur.description]

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        ws.append(columns)
        for row in rows:
            ws.append(row)

        report_date_obj = datetime.strptime(report_date, '%d.%m.%Y')
        filename = f"Invoices_card_payments_{report_date_obj.strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join(UPLOAD_DIR, filename)

        wb.save(filepath)

        msg = f'📊 Report pre dátum {report_date} bol uložený ako {filename}.'
        message += msg + '<br>'
        log(msg)

    except ValueError:
        msg = '❌ Nesprávny formát dátumu. Použi DD.MM.YYYY'
        message += msg + '<br>'
        log(msg)
    except Exception as e:
        msg = f'❌ Chyba pri generovaní reportu: {str(e)}'
        message += msg + '<br>'
        log(msg)

    # F. Nahrávanie reportu na SharePoint
    msg = "------------ START nahrávanie reportu na SharePoint ------------"
    log(msg)

    try:
        # Nastavenie cesty k upload priečinku
        UPLOAD_DIR = os.path.join(settings.PAYWELL_ROOT, 'uploads')
        os.makedirs(UPLOAD_DIR, exist_ok=True)

        # Nájdi najnovší .xlsx súbor
        xlsx_files = [f for f in os.listdir(UPLOAD_DIR) if f.endswith('.xlsx')]
        if not xlsx_files:
            msg = '❌ Žiadny .xlsx súbor nebol nájdený v priečinku uploads.'
            message += msg + '<br>'
            log(msg)
            raise Exception("XLSX not found")

        xlsx_files.sort()
        filename = xlsx_files[-1]
        filepath = os.path.join(UPLOAD_DIR, filename)

        # Extrahuj dátum zo súboru
        match = re.search(r'Invoices_card_payments_(\d{8})\.xlsx', filename)
        if not match:
            msg = '❌ Názov súboru nie je v správnom formáte: Invoices_card_payments_YYYYMMDD.xlsx'
            message += msg + '<br>'
            log(msg)
            raise Exception("Invalid filename format")

        date_str = match.group(1)
        report_date_obj = datetime.strptime(date_str, '%Y%m%d')

        # Načítaj SharePoint prihlasovacie údaje zo JSON
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        json_path = os.path.join(BASE_DIR, 'edz', 'static', 'json', 'jobs.json')

        def load_cred(field):
            with open(json_path, 'r', encoding='utf-8') as file:
                data = json.load(file)
                return [entry['fields'][field] for entry in data if entry['model'] == 'EDZ_Podania'][0]

        username = load_cred('username')
        password = load_cred('password')

        # Nastav prostredie bez proxy
        os.environ['HTTP_PROXY'] = ''
        os.environ['HTTPS_PROXY'] = ''
        os.environ['NO_PROXY'] = 'login.microsoftonline.com,sharepoint.com'

        # Vytvor relatívnu cestu podľa dátumu
        current_year = report_date_obj.strftime('%Y')
        current_month = report_date_obj.strftime('%Y_%m')
        folder_date = report_date_obj.strftime('%Y_%m_%d')

        # site_url = "https://example.sharepoint.com/teams/coo-pos"
        # sharepoint_folder_path = f'/5410/EtwacacteUVFlSAMSxS0eNABkwLFLIzd6czg0cGrqq_mIA/'

        site_url = 'https://example.sharepoint.com/teams/coo-infra/5230/'
        sharepoint_folder_path = f'/teams/coo-infra/5230//Reporty/denné/{current_year}/{current_month}/{folder_date}'


        # Autentifikácia a nahratie súboru
        ctx_auth = AuthenticationContext(site_url)
        if ctx_auth.acquire_token_for_user(username, password):
            ctx = ClientContext(site_url, ctx_auth)
            with open(filepath, 'rb') as file_obj:
                file_content = file_obj.read()
                folder = ctx.web.get_folder_by_server_relative_url(sharepoint_folder_path)
                upload_result = folder.upload_file(filename, file_content).execute_query()
                msg = f'☁️ Súbor <b>{filename}</b> bol nahraný na SharePoint: <code>{upload_result.serverRelativeUrl}</code>'
                message += msg + '<br>'
                log(msg)
        else:
            msg = '❌ Nepodarilo sa autentifikovať na SharePoint pomocou poskytnutých údajov.'
            message += msg + '<br>'
            log(msg)

    except Exception as e:
        msg = f'❌ Chyba pri nahrávaní na SharePoint: {str(e)}'
        message += msg + '<br>'
        log(msg)

    # G. Vymazanie všetkých súborov z uploads
    msg = "------------ START vymazanie VŠETKÝCH súborov ------------"
    log(msg)

    deleted = 0
    for file in os.listdir(UPLOAD_DIR):
        try:
            os.remove(os.path.join(UPLOAD_DIR, file))
            deleted += 1
        except Exception as e:
            err_msg = f"⚠️ Nepodarilo sa zmazať {file}: {e}"
            log(err_msg)

    msg = f'🧹 Vymazaných {deleted} súborov.'
    message += msg + '<br>'
    log(msg)
    ###### THE END PAYWELL SCHEDULER ######





def download_log(request):
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    log_file_path = os.path.join(BASE_DIR, 'Scripts', 'Tyzdenne', 'auto.log')

    if os.path.exists(log_file_path):
        return FileResponse(open(log_file_path, 'rb'), as_attachment=True, filename='auto.log')
    else:
        return HttpResponse("Log file not found.", status=404)
    
def download_log_daily(request):
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    log_file_path = os.path.join(BASE_DIR, 'Scripts', 'Denne', 'auto.log')

    if os.path.exists(log_file_path):
        return FileResponse(open(log_file_path, 'rb'), as_attachment=True, filename='auto.log')
    else:
        return HttpResponse("Log file not found.", status=404)




import os
import json
from typing import List, Dict, Union
from django.shortcuts import render

def build_file_tree(root: str) -> List[Dict[str, Union[str, List]]]:
    """
    Rekurzívne vytvorí strom priečinkov a súborov zo zadanej cesty.

    Args:
        root (str): Koreňový adresár.

    Returns:
        List[Dict]: Zoznam priečinkov a súborov so štruktúrou stromu.
    """
    strom = []
    for entry in os.scandir(root):
        if entry.is_dir():
            strom.append({
                "type": "directory",
                "name": entry.name,
                "children": build_file_tree(entry.path)
            })
        else:
            strom.append({
                "type": "file",
                "name": entry.name
            })
    return strom

from django.http import FileResponse, Http404
from urllib.parse import unquote

def download_daily_file(request, relative_path: str):
    """
    Umožní stiahnutie konkrétneho súboru z Reports/Daily.
    
    Args:
        request (HttpRequest): HTTP požiadavka.
        relative_path (str): Relatívna cesta k súboru.

    Returns:
        FileResponse: Súbor ako odpoveď.
    """
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    export_root = os.path.join(BASE_DIR, 'Reports', 'Daily')
    
    file_path = os.path.abspath(os.path.join(export_root, unquote(relative_path)))

    # Bezpečnostná kontrola: súbor musí byť v export_root
    if not file_path.startswith(export_root):
        raise Http404("Neoprávnený prístup k súboru")

    if not os.path.isfile(file_path):
        raise Http404("Súbor neexistuje")

    return FileResponse(open(file_path, 'rb'), as_attachment=True)

def download_inactivity_file(request, relative_path: str):
    """
    Umožní stiahnutie konkrétneho súboru z Reports/Daily.
    
    Args:
        request (HttpRequest): HTTP požiadavka.
        relative_path (str): Relatívna cesta k súboru.

    Returns:
        FileResponse: Súbor ako odpoveď.
    """
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    export_root = os.path.join(BASE_DIR, 'Reports', 'Weekly_4m_6m')
    
    file_path = os.path.abspath(os.path.join(export_root, unquote(relative_path)))

    # Bezpečnostná kontrola: súbor musí byť v export_root
    if not file_path.startswith(export_root):
        raise Http404("Neoprávnený prístup k súboru")

    if not os.path.isfile(file_path):
        raise Http404("Súbor neexistuje")

    return FileResponse(open(file_path, 'rb'), as_attachment=True)

# def inactivity(request):
#     BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
#     export_root = os.path.join(BASE_DIR, 'Reports', 'Weekly_4m_6m')
#     log_file_path = os.path.join(BASE_DIR, 'Scripts', 'Tyzdenne', 'auto.log')
#     with open(log_file_path, 'r', encoding='utf-8') as f:
#                 lines = f.readlines()
#                 log_lines = lines[-30:]  # posledných N riadkov

#     # === Cesta k jobs.json ===
#     JOBS_JSON_PATH = os.path.join(BASE_DIR, 'work', 'static', 'json', 'jobs.json')

#     # === Cesta k jobs.json ===
#     JOBS_JSON_PATH = os.path.join(BASE_DIR, 'work', 'static', 'json', 'jobs.json')
#     message = ""
#     current_run = {
#         "hour": "*",
#         "minute": "*",
#         "day_of_week": "*"
#     }

#     # === POST: aktualizuj plánovanie ===
#     # === POST: aktualizuj plánovanie a emaily ===
#     if request.method == "POST" and request.POST.get("update_scheduler"):
#         updated_run = {
#             "hour": request.POST.get("hour", "*"),
#             "minute": request.POST.get("minute", "*"),
#             "day_of_week": request.POST.get("day_of_week", "*")
#         }
#         # Načítanie emailov z formulára
#         raw_emails = request.POST.get("emails", "")
#         # Prevod na list: rozdelenie podľa čiarky a odstránenie prázdnych znakov
#         email_list = [e.strip() for e in raw_emails.split(',') if e.strip()]

#         try:
#             with open(JOBS_JSON_PATH, 'r', encoding='utf-8') as f:
#                 data = json.load(f)

#             for entry in data:
#                 if entry.get("model") == "inactivity":
#                     entry["fields"]["run"] = json.dumps(updated_run)
#                     entry["fields"]["emails"] = email_list  # Uloženie listu emailov
#                     break

#             with open(JOBS_JSON_PATH, 'w', encoding='utf-8') as f:
#                 json.dump(data, f, indent=4, ensure_ascii=False)

#             message = "✅ Nastavenia a emaily boli úspešne aktualizované."
#             current_run = updated_run
            
#             from scheduler import reload_job_if_inactivity
#             reload_job_if_inactivity('inactivity', 'work/static/json/jobs.json')
            
#         except Exception as e:
#             message = f"❌ Chyba pri ukladaní: {e}"

#     # === GET: načítaj aktuálne hodnoty ===
#     else:
#         try:
#             with open(JOBS_JSON_PATH, 'r', encoding='utf-8') as f:
#                 data = json.load(f)
#                 for entry in data:
#                     if entry.get('model') == 'inactivity':
#                         run_raw = entry.get('fields', {}).get('run', '{}')
#                         current_run = json.loads(run_raw)
#                         break
#         except Exception as e:
#             message = f"❌ Chyba pri načítaní jobs.json: {e}"
    
#         # === Načítanie stromu súborov ===
#     try:
#         file_tree = build_file_tree(export_root)
#     except Exception as e:
#         file_tree = [{"type": "error", "name": f"❌ Chyba pri načítaní priečinkov: {e}"}]


#     return render(request, 'ems/inactivity.html', {
#         'log_lines': log_lines,
#         'current_run': current_run,
#         'message': message,
#         'file_tree':file_tree
#     })

def inactivity(request):
    import json
    import os
    from django.shortcuts import render

    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    export_root = os.path.join(BASE_DIR, 'Reports', 'Weekly_4m_6m')
    log_file_path = os.path.join(BASE_DIR, 'Scripts', 'Tyzdenne', 'auto.log')
    JOBS_JSON_PATH = os.path.join(BASE_DIR, 'work', 'static', 'json', 'jobs.json')

    # Načítanie logov
    try:
        with open(log_file_path, 'r', encoding='utf-8') as f:
            log_lines = f.readlines()[-30:]
    except:
        log_lines = ["Log súbor nebol nájdený."]

    message = ""
    current_run = {"hour": "*", "minute": "*", "day_of_week": "*"}
    current_emails = "" # <--- Inicializácia premennej pre HTML

    # === POST: aktualizuj plánovanie a emaily ===
    if request.method == "POST" and request.POST.get("update_scheduler"):
        updated_run = {
            "hour": request.POST.get("hour", "*"),
            "minute": request.POST.get("minute", "*"),
            "day_of_week": request.POST.get("day_of_week", "*")
        }
        raw_emails = request.POST.get("emails", "")
        email_list = [e.strip() for e in raw_emails.split(',') if e.strip()]

        try:
            with open(JOBS_JSON_PATH, 'r', encoding='utf-8') as f:
                data = json.load(f)

            for entry in data:
                if entry.get("model") == "inactivity":
                    entry["fields"]["run"] = json.dumps(updated_run)
                    entry["fields"]["emails"] = email_list
                    break

            with open(JOBS_JSON_PATH, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=4, ensure_ascii=False)

            message = "✅ Nastavenia a emaily boli úspešne aktualizované."
            current_run = updated_run
            current_emails = ", ".join(email_list) # <--- Aktualizujeme pre zobrazenie po POSTe
            
            from scheduler import reload_job_if_inactivity
            reload_job_if_inactivity('inactivity', 'work/static/json/jobs.json')
            
        except Exception as e:
            message = f"❌ Chyba pri ukladaní: {e}"

    # === GET: načítaj aktuálne hodnoty ===
    else:
        try:
            with open(JOBS_JSON_PATH, 'r', encoding='utf-8') as f:
                data = json.load(f)
                for entry in data:
                    if entry.get('model') == 'inactivity':
                        # Načítanie času
                        run_raw = entry.get('fields', {}).get('run', '{}')
                        current_run = json.loads(run_raw)
                        # Načítanie emailov a ich spojenie do stringu pre HTML input
                        emails_list = entry.get('fields', {}).get('emails', [])
                        current_emails = ", ".join(emails_list) # <--- TOTO TI CHÝBALO
                        break
        except Exception as e:
            message = f"❌ Chyba pri načítaní jobs.json: {e}"
    
    # Načítanie stromu súborov
    try:
        file_tree = build_file_tree(export_root)
    # try:
    #     from .utils import build_file_tree # Predpokladám, že máš build_file_tree niekde definované
        file_tree = build_file_tree(export_root)
    except Exception as e:
        file_tree = [{"type": "error", "name": f"❌ Chyba pri načítaní priečinkov: {e}"}]

    return render(request, 'ems/inactivity.html', {
        'log_lines': log_lines,
        'current_run': current_run,
        'current_emails': current_emails, # <--- Musíme poslať do template
        'message': message,
        'file_tree': file_tree
    })

def download_daily_file_weekly(request, relative_path: str):
    """
    Umožní stiahnutie konkrétneho súboru z Reports/Weekly_4m_6m.

    Args:
        request (HttpRequest): HTTP požiadavka.
        relative_path (str): Relatívna cesta k súboru.

    Returns:
        FileResponse: Súbor ako odpoveď.
    """
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    export_root = os.path.join(BASE_DIR, 'Reports', 'Weekly_4m_6m')

    file_path = os.path.abspath(os.path.join(export_root, unquote(relative_path)))

    # Bezpečnostná kontrola: súbor musí byť v export_root
    if not file_path.startswith(export_root):
        raise Http404("Neoprávnený prístup k súboru")

    if not os.path.isfile(file_path):
        raise Http404("Súbor neexistuje")

    return FileResponse(open(file_path, 'rb'), as_attachment=True)


def weekly_reports_view(request):
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    export_root = os.path.join(base_dir, 'Reports', 'Weekly_4m_6m')

    def build_tree(path):
        tree = []
        for name in os.listdir(path):
            full_path = os.path.join(path, name)
            if os.path.isdir(full_path):
                tree.append({
                    'type': 'directory',
                    'name': name,
                    'children': build_tree(full_path)
                })
            else:
                tree.append({
                    'type': 'file',
                    'name': name
                })
        return tree

    file_tree = build_tree(export_root)
    return render(request, 'ems/weekly_reports.html', {'file_tree': file_tree})


# def inactivity_report():
#     from openpyxl import Workbook
#     from datetime import datetime
#     import os
#     import cx_Oracle
#     import json
#     from office365.runtime.auth.authentication_context import AuthenticationContext
#     from office365.sharepoint.client_context import ClientContext
#     from work.models import EMS_DWHDB

#     BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
#     sql_storage = 'Scripts/Tyzdenne'
#     query_6m = 'REP_INACTIVITY_6M_PRP.sql'
#     query_4m = 'REP_INACTIVITY_4M.sql'

#     sql_6m_prp_path = os.path.join(BASE_DIR, sql_storage, query_6m)
#     sql_4m_path = os.path.join(BASE_DIR, sql_storage, query_4m)
#     export_dir = os.path.join(BASE_DIR, 'Reports', 'Weekly_4m_6m')
#     log_file_path = os.path.join(BASE_DIR, 'Scripts', 'Tyzdenne', 'auto.log')
#     os.makedirs(export_dir, exist_ok=True)

#     def log(msg):
#         timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#         with open(log_file_path, 'a', encoding='utf-8') as f:
#             f.write(f"[{timestamp}] {msg}\n")

#     def load_cred(field):
#         json_path = os.path.join(BASE_DIR, 'edz', 'static', 'json', 'jobs.json')
#         with open(json_path, 'r', encoding='utf-8') as file:
#             data = json.load(file)
#             return [entry['fields'][field] for entry in data if entry['model'] == 'EDZ_Podania'][0]

#     username = load_cred('username')
#     password = load_cred('password')

#     date_str = datetime.now().strftime('%Y%m%d')

#     def generate_report(query_path, report_name):
#         try:
#             with open(query_path, 'r', encoding='utf-8') as f:
#                 sql_query = f.read()
#             log(f"📥 Načítaný SQL súbor: {os.path.basename(query_path)}")
#         except Exception as e:
#             log(f"❌ Chyba pri čítaní SQL súboru: {e}")
#             return

#         try:
#             EMS_DWH = EMS_DWHDB.objects.get(user=1)
#             con = cx_Oracle.connect(
#                 f"{EMS_DWH.EMS_DWHDB_username}/{EMS_DWH.EMS_DWHDB_password}@{EMS_DWH.EMS_DWHDB_hostname}:{EMS_DWH.EMS_DWHDB_port}/{EMS_DWH.EMS_DWHDB_servicename}"
#             )
#             cur = con.cursor()
#             log("✅ Pripojenie k databáze úspešné.")

#             cur.execute(sql_query)
#             rows = cur.fetchall()
#             columns = [desc[0] for desc in cur.description]
#             log(f"📊 {report_name}: Načítaných {len(rows)} riadkov z databázy.")

#             cur.close()
#             con.close()
#         except Exception as e:
#             log(f"❌ {report_name}: Chyba počas pripojenia alebo načítania dát: {e}")
#             return

#         wb = Workbook()
#         ws = wb.active
#         ws.title = report_name
#         ws.append(columns)
#         for row in rows:
#             ws.append(row)

#         file_name = f"{report_name}_{date_str}.xlsx"
#         file_path = os.path.join(export_dir, file_name)
#         wb.save(file_path)
#         log(f"✅ {report_name}: Report bol uložený ako {file_name}")

#         # Nastav prostredie bez proxy
#         os.environ['HTTP_PROXY'] = ''
#         os.environ['HTTPS_PROXY'] = ''
#         os.environ['NO_PROXY'] = 'login.microsoftonline.com,sharepoint.com'

#         site_url = 'https://example.sharepoint.com/teams/coo-pos/5410/'
#         sharepoint_folder_path = '/teams/coo-pos/5410/Zdielane dokumenty/Inaktivita_PJ/zdrojove_reporty'

#         ctx_auth = AuthenticationContext(site_url)
#         if ctx_auth.acquire_token_for_user(username, password):
#             ctx = ClientContext(site_url, ctx_auth)
#             try:
#                 log(f"🔍 Pokúšam sa načítať priečinok na SharePointe: {sharepoint_folder_path}")
#                 folder = ctx.web.get_folder_by_server_relative_url(sharepoint_folder_path)
#                 ctx.load(folder)
#                 ctx.execute_query()
#                 log(f"📂 SharePoint priečinok {sharepoint_folder_path} úspešne načítaný.")
#             except Exception as e:
#                 msg = f"❌ Chyba pri načítaní folderu na SharePointe: {str(e)}"
#                 log(msg)
#                 return  # ukonči ak folder nie je prístupný

#             # === Nahraj všetky .xlsx súbory z export_dir ===
#             for filename in os.listdir(export_dir):
#                 if filename.lower().endswith('.xlsx'):
#                     file_path = os.path.join(export_dir, filename)
#                     try:
#                         with open(file_path, 'rb') as file_obj:
#                             file_content = file_obj.read()
#                             upload_result = folder.upload_file(filename, file_content).execute_query()
#                             msg = f"☁️ Súbor {filename} bol nahraný na SharePoint: {upload_result.serverRelativeUrl}"
#                             log(msg)
#                     except Exception as e:
#                         log(f"❌ Chyba pri nahrávaní súboru {filename}: {str(e)}")
#         else:
#             msg = '❌ Nepodarilo sa autentifikovať na SharePoint pomocou poskytnutých údajov.'
#             log(msg)

#     log("🔄 Spúšťam generovanie reportov REP_INACTIVITY_6M_PRP a REP_INACTIVITY_4M...")

#     for file_name in os.listdir(export_dir):
#         file_path = os.path.join(export_dir, file_name)
#         if os.path.isfile(file_path):
#             os.remove(file_path)
#     log("🧹 Staré súbory boli vymazané.")

#     generate_report(sql_6m_prp_path, "REP_INACTIVITY_6M_PRP")
#     generate_report(sql_4m_path, "REP_INACTIVITY_4M")
#     log("🔚 Generovanie reportov dokončené.")

def inactivity_report():
    from openpyxl import Workbook
    from datetime import datetime
    import os
    import cx_Oracle
    import json
    import zipfile  # <--- Pridaný import pre zipovanie
    from office365.runtime.auth.authentication_context import AuthenticationContext
    from office365.sharepoint.client_context import ClientContext
    from work.models import EMS_DWHDB

    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    sql_storage = 'Scripts/Tyzdenne'
    query_6m = 'REP_INACTIVITY_6M_PRP.sql'
    query_4m = 'REP_INACTIVITY_4M.sql'

    sql_6m_prp_path = os.path.join(BASE_DIR, sql_storage, query_6m)
    sql_4m_path = os.path.join(BASE_DIR, sql_storage, query_4m)
    export_dir = os.path.join(BASE_DIR, 'Reports', 'Weekly_4m_6m')
    log_file_path = os.path.join(BASE_DIR, 'Scripts', 'Tyzdenne', 'auto.log')
    os.makedirs(export_dir, exist_ok=True)

    def log(msg):
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        with open(log_file_path, 'a', encoding='utf-8') as f:
            f.write(f"[{timestamp}] {msg}\n")

    def load_cred(field):
        json_path = os.path.join(BASE_DIR, 'edz', 'static', 'json', 'jobs.json')
        with open(json_path, 'r', encoding='utf-8') as file:
            data = json.load(file)
            return [entry['fields'][field] for entry in data if entry['model'] == 'EDZ_Podania'][0]

    username = load_cred('username')
    password = load_cred('password')
    date_str = datetime.now().strftime('%Y%m%d')

    # 1. Funkcia na generovanie Excelu (bez nahrávania na SharePoint)
    def generate_excel(query_path, report_name):
        try:
            with open(query_path, 'r', encoding='utf-8') as f:
                sql_query = f.read()
            log(f"📥 Načítaný SQL súbor: {os.path.basename(query_path)}")
        except Exception as e:
            log(f"❌ Chyba pri čítaní SQL súboru: {e}")
            return None

        try:
            EMS_DWH = EMS_DWHDB.objects.get(user=1)
            con = cx_Oracle.connect(
                f"{EMS_DWH.EMS_DWHDB_username}/{EMS_DWH.EMS_DWHDB_password}@{EMS_DWH.EMS_DWHDB_hostname}:{EMS_DWH.EMS_DWHDB_port}/{EMS_DWH.EMS_DWHDB_servicename}"
            )
            cur = con.cursor()
            cur.execute(sql_query)
            rows = cur.fetchall()
            columns = [desc[0] for desc in cur.description]
            log(f"📊 {report_name}: Načítaných {len(rows)} riadkov.")
            cur.close()
            con.close()
        except Exception as e:
            log(f"❌ {report_name}: Chyba DB: {e}")
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = report_name
        ws.append(columns)
        for row in rows:
            ws.append(row)

        file_name = f"{report_name}_{date_str}.xlsx"
        file_path = os.path.join(export_dir, file_name)
        wb.save(file_path)
        log(f"✅ {report_name}: Uložené do {file_name}")
        return file_path

    # --- ZAČIATOK PROCESU ---
    log("🔄 Spúšťam proces generovania reportov a ZIP archívu...")

    # Vyčistenie starých súborov
    for f in os.listdir(export_dir):
        os.remove(os.path.join(export_dir, f))
    log("🧹 Priečinok vyčistený.")

    # Generovanie oboch Excelov
    file1 = generate_excel(sql_6m_prp_path, "REP_INACTIVITY_6M_PRP")
    file2 = generate_excel(sql_4m_path, "REP_INACTIVITY_4M")

    # 2. Vytvorenie ZIP súboru
    zip_name = f"Inactivity_Reports_{date_str}.zip"
    zip_path = os.path.join(export_dir, zip_name)
    
    try:
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            if file1: zipf.write(file1, os.path.basename(file1))
            if file2: zipf.write(file2, os.path.basename(file2))
        log(f"📦 ZIP archív vytvorený: {zip_name}")
    except Exception as e:
        log(f"❌ Chyba pri vytváraní ZIP: {e}")

    # 3. Nahrávanie na SharePoint (nahrá všetko v export_dir, vrátane ZIPu)
    # os.environ['HTTP_PROXY'] = ''
    # os.environ['HTTPS_PROXY'] = ''
    
    # site_url = 'https://example.sharepoint.com/teams/coo-pos/5410/'
    # sharepoint_folder_path = '/teams/coo-pos/5410/Zdielane dokumenty/Inaktivita_PJ/zdrojove_reporty'

    # ctx_auth = AuthenticationContext(site_url)
    # if ctx_auth.acquire_token_for_user(username, password):
    #     ctx = ClientContext(site_url, ctx_auth)
    #     try:
    #         folder = ctx.web.get_folder_by_server_relative_url(sharepoint_folder_path)
    #         for filename in os.listdir(export_dir):
    #             file_full_path = os.path.join(export_dir, filename)
    #             with open(file_full_path, 'rb') as f_obj:
    #                 folder.upload_file(filename, f_obj.read()).execute_query()
    #             log(f"☁️ Nahrané na SharePoint: {filename}")
    #     except Exception as e:
    #         log(f"❌ SharePoint error: {e}")
    # else:
    #     log("❌ SharePoint autentifikácia zlyhala.")

    # 1. Definícia cesty a načítanie dát (podľa tvojho zadania)
    JOBS_JSON_PATH = os.path.join(BASE_DIR, 'work', 'static', 'json', 'jobs.json')

    def load_inactivity_config():
        try:
            with open(JOBS_JSON_PATH, 'r', encoding='utf-8') as file:
                data = json.load(file)
                for entry in data:
                    if entry.get('model') == 'inactivity':
                        return entry.get('fields', {})
        except Exception as e:
            log(f"❌ Chyba pri načítaní JSON: {e}")
        return {}

    # Načítanie polí z JSON
    config = load_inactivity_config()
    
    # 2. Príprava e-mailových premenných
    email_addresses = config.get('emails', [])  # V JSON máš list [], takže netreba split
    subject_head = config.get('head', 'Inactivity Report')
    message_body = config.get('body', 'Dobrý deň, v prílohe posielame report.')

    # --- Sekcia odoslania emailu (vložiť po vytvorení ZIPu) ---
    if email_addresses:
        try:
            from django.core.mail import EmailMessage
            
            message = EmailMessage(
                subject=subject_head,
                body=message_body,
                from_email='reports@example.com',
                to=email_addresses  # Django priamo akceptuje list z JSONu
            )

            # Priloženie vytvoreného ZIP súboru
            if os.path.exists(zip_path):
                message.attach_file(zip_path)
            
            message.send()
            log(f"📧 Report úspešne odoslaný na: {', '.join(email_addresses)}")
            print('report poslany')
            
        except Exception as e:
            log(f"❌ Chyba pri odosielaní emailu: {str(e)}")
    else:
        log("⚠️ Žiadne emailové adresy neboli nájdené v JSON.")

    log("🔚 Hotovo.")

def daily_rep(request):
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    log_file_path = os.path.join(BASE_DIR, 'Scripts', 'Denne', 'auto.log')
    export_root = os.path.join(BASE_DIR, 'Reports', 'Daily')
    JOBS_JSON_PATH = os.path.join(BASE_DIR, 'work', 'static', 'json', 'jobs.json')

    message = ""
    current_run = {"hour": "*", "minute": "*", "day": "*", "month": "*"}

    # === Načítanie logu ===
    try:
        with open(log_file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            log_lines = lines[-50:]
    except Exception as e:
        log_lines = [f"❌ Chyba pri čítaní log súboru: {e}"]

    # === POST: aktualizácia plánovača ===
    if request.method == "POST" and request.POST.get("update_scheduler"):
        updated_run = {
            "hour": request.POST.get("hour", "*"),
            "minute": request.POST.get("minute", "*"),
            "day": request.POST.get("day", "*"),
            "month": request.POST.get("month", "*"),
        }

        try:
            with open(JOBS_JSON_PATH, 'r', encoding='utf-8') as f:
                data = json.load(f)

            for entry in data:
                if entry.get("model") == "daily_reports":
                    entry["fields"]["run"] = json.dumps(updated_run)
                    break

            with open(JOBS_JSON_PATH, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=4, ensure_ascii=False)

            message = "✅ Scheduler bol úspešne aktualizovaný."
            current_run = updated_run

            from scheduler import reload_job_if_daily_reports
            reload_job_if_daily_reports('daily_reports', 'work/static/json/jobs.json')

        except Exception as e:
            message = f"❌ Chyba pri ukladaní: {e}"

    else:
        try:
            with open(JOBS_JSON_PATH, 'r', encoding='utf-8') as f:
                data = json.load(f)
                for entry in data:
                    if entry.get('model') == 'daily_reports':
                        run_raw = entry.get('fields', {}).get('run', '{}')
                        current_run = json.loads(run_raw)
                        break
        except Exception as e:
            message = f"❌ Chyba pri načítaní jobs.json: {e}"

    # === Načítanie stromu súborov ===
    try:
        file_tree = build_file_tree(export_root)
    except Exception as e:
        file_tree = [{"type": "error", "name": f"❌ Chyba pri načítaní priečinkov: {e}"}]

    return render(request, 'ems/daily_rep.html', {
        'log_lines': log_lines,
        'current_run': current_run,
        'message': message,
        'file_tree': file_tree
    })


def daily_reports(
    user_id: int = 1,
    sql_filename: str = 'TOIS_Invoices_v17.sql',
    report_name: str = 'Invoices'
    ):
    from openpyxl import Workbook
    from datetime import datetime, timedelta
    import os
    import shutil
    import cx_Oracle
    import json
    from office365.runtime.auth.authentication_context import AuthenticationContext
    from office365.sharepoint.client_context import ClientContext
    from work.models import BILLDB

    """
    1. Vytvorí priečinok DDMMYYYY v Reports/Daily
    2. Zmaže priečinky staršie ako 3 dni
    3. Spustí SQL skript a uloží výsledok do XLSX
    4. Loguje každú akciu do auto.log
    5. Nahrá vygenerovaný report na SharePoint
    """


    BASE_DIR      = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    sql_path      = os.path.join(BASE_DIR, 'Scripts', 'Denne', sql_filename)
    export_root   = os.path.join(BASE_DIR, 'Reports', 'Daily')
    log_file_path = os.path.join(BASE_DIR, 'Scripts', 'Denne', 'auto.log')
    os.makedirs(export_root, exist_ok=True)

    def log(msg: str):
        ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        with open(log_file_path, 'a', encoding='utf-8') as f:
            f.write(f"[{ts}] {msg}\n")

    log("🟢 Spúšťam denný reporting...")

    # ===== Priečinok pre dnešok + čistenie ===============================
    today        = datetime.now()
    today_str    = today.strftime("%d%m%Y")
    yesterday_str = (today - timedelta(days=1)).strftime("%d%m%Y")
    today_dir    = os.path.join(export_root, today_str)
    threshold    = today - timedelta(days=3)

    if not os.path.exists(today_dir):
        os.makedirs(today_dir)
        log(f"📂 Vytvorený priečinok pre dnešný deň: {today_str}")
    else:
        log(f"📂 Priečinok pre dnešný deň už existuje: {today_str}")

    for name in os.listdir(export_root):
        path = os.path.join(export_root, name)
        if os.path.isdir(path) and name.isdigit() and len(name) == 8:
            try:
                folder_date = datetime.strptime(name, "%d%m%Y")
            except ValueError:
                continue
            if folder_date <= threshold:
                try:
                    shutil.rmtree(path)
                    log(f"🗑️  Odstránený starý priečinok: {name}")
                except Exception as e:
                    log(f"❌ Chyba pri odstraňovaní {name}: {e}")

    # ===== Načítaj SQL ===================================================
    log(f"📄 Načítavam SQL súbor: {sql_filename}")
    try:
        with open(sql_path, encoding='utf-8') as qf:
            sql_query = qf.read()
        log("✅ SQL skript úspešne načítaný.")
    except FileNotFoundError:
        log(f"❌ SQL súbor {sql_filename} sa nenašiel.")
        log("🔴 Ukončenie procesu kvôli chybe pri načítaní SQL.")
        return

    # ===== Pripojenie pomocou BILLDB ====================================
    log(f"🔐 Načítavam prihlasovacie údaje z BILLDB pre user_id={user_id}")
    try:
        bill = BILLDB.objects.get(user__id=user_id)
        log("✅ Údaje z BILLDB úspešne načítané.")
    except BILLDB.DoesNotExist:
        log(f"❌ Záznam BILLDB pre user={user_id} neexistuje.")
        log("🔴 Ukončenie procesu kvôli chýbajúcim údajom.")
        return

    dsn = (
        f"{bill.BILLDB_username}/"
        f"{bill.BILLDB_password}@"
        f"{bill.BILLDB_hostname}:"
        f"{bill.BILLDB_port}/"
        f"{bill.BILLDB_servicename}"
    )

    log("🌐 Pripájam sa na Oracle databázu...")
    try:
        con = cx_Oracle.connect(dsn, encoding="UTF-8")
        cur = con.cursor()
        log("✅ Pripojenie k databáze BILLDB úspešné.")
    except Exception as e:
        log(f"❌ Nepodarilo sa pripojiť: {e}")
        log("🔴 Ukončenie procesu kvôli chybe pri pripojení.")
        return

    # ===== Spusti SQL a exportuj ========================================
    try:
        log("▶ Spúšťam SQL dotaz...")
        cur.execute(sql_query)
        rows     = cur.fetchall()
        columns  = [desc[0] for desc in cur.description]
        log(f"📊 {report_name}: Načítaných {len(rows)} riadkov z databázy.")

        # Export XLSX
        wb  = Workbook()
        ws  = wb.active
        ws.title = report_name
        ws.append(columns)
        for r in rows:
            ws.append(r)

        report_path = os.path.join(today_dir, f"{report_name}_{yesterday_str}.xlsx")
        wb.save(report_path)
        log(f"💾 Report uložený: {report_path}")

    except Exception as e:
        log(f"❌ Chyba pri generovaní reportu: {e}")

    finally:
        try:
            cur.close()
            con.close()
            log("🔒 Spojenie s databázou uzavreté.")
        except Exception as e:
            log(f"⚠️  Chyba pri uzatváraní spojenia: {e}")




    # ===== Načítaj SharePoint prihlasovacie údaje zo JSON =================
    json_path = os.path.join(BASE_DIR, 'edz', 'static', 'json', 'jobs.json')

    def load_cred(field):
        with open(json_path, 'r', encoding='utf-8') as file:
            data = json.load(file)
            return [entry['fields'][field] for entry in data if entry['model'] == 'EDZ_Podania'][0]

    username = load_cred('username')
    password = load_cred('password')

    # Nastav prostredie bez proxy
    os.environ['HTTP_PROXY'] = ''
    os.environ['HTTPS_PROXY'] = ''
    os.environ['NO_PROXY'] = 'login.microsoftonline.com,sharepoint.com'

    # Vytvor relatívnu cestu podľa včerajšieho dátumu
    yesterday = today - timedelta(days=1)
    current_year = yesterday.strftime('%Y')
    current_month = yesterday.strftime('%Y_%m')
    folder_date = yesterday.strftime('%Y_%m_%d')
    site_url = 'https://example.sharepoint.com/teams/coo-infra/5230/'
    sharepoint_folder_path = f'/teams/coo-infra/5230/Reporty/denné/{current_year}/{current_month}/{folder_date}'
    
    # Autentifikácia a nahratie súbor/
    try:
        ctx_auth = AuthenticationContext(site_url)
        if ctx_auth.acquire_token_for_user(username, password):
            ctx = ClientContext(site_url, ctx_auth)
            report_filename = f"{report_name}_{yesterday_str}.xlsx"
            report_filepath = os.path.join(today_dir, report_filename)
            with open(report_filepath, 'rb') as file_obj:
                file_content = file_obj.read()
            folder = ctx.web.get_folder_by_server_relative_url(sharepoint_folder_path)
            upload_result = folder.upload_file(report_filename, file_content).execute_query()
            log(f'☁️ Súbor <b>{report_filename}</b> bol nahraný na SharePoint: <code>{upload_result.serverRelativeUrl}</code>')
        else:
            log('❌ Nepodarilo sa autentifikovať na SharePoint pomocou poskytnutých údajov.')
    except Exception as e:
        log(f"⚠️ Chyba pri nahrávaní na SharePoint: {e}")


    # ====== Druhý report: Vyrubene_myto_kategorie ==============================
    try:
        log("🟡 Spúšťam druhý report: Vyrubene_myto_kategorie...")

        try:
            EMS_DWH = EMS_DWHDB.objects.get(user_id=1)

            dsn = (
                f"{EMS_DWH.EMS_DWHDB_username}/"
                f"{EMS_DWH.EMS_DWHDB_password}@"
                f"{EMS_DWH.EMS_DWHDB_hostname}:"
                f"{EMS_DWH.EMS_DWHDB_port}/"
                f"{EMS_DWH.EMS_DWHDB_servicename}"
            )

            con_emsdwh_stdby = cx_Oracle.connect(dsn, encoding="UTF-8")
            cur_emsdwh = con_emsdwh_stdby.cursor()

        except ObjectDoesNotExist:
            raise RuntimeError("❌ Záznam EMS_DWHDB pre user_id=1 sa nenašiel.")
        except cx_Oracle.Error as e:
            raise RuntimeError(f"❌ Chyba pri pripájaní na Oracle databázu: {e}")

        sql_file_path = os.path.join(BASE_DIR, 'Scripts', 'Denne', 'Vyrubene_myto_kategorie.sql')
        try:
            with open(sql_file_path, 'r', encoding='utf-8') as f:
                vyrubene_sql = f.read()
        except UnicodeDecodeError:
            log("⚠️ UTF-8 dekódovanie zlyhalo. Skúšam cp1250 (Windows-1250)...")
            with open(sql_file_path, 'r', encoding='cp1250') as f:
                vyrubene_sql = f.read()

        cur_emsdwh.execute(vyrubene_sql)
        rows = cur_emsdwh.fetchall()
        columns = [desc[0] for desc in cur_emsdwh.description]

        import pandas as pd
        df = pd.DataFrame(rows, columns=columns)

        vyrubene_filename = f"Vyrubene_myto_kategorie_{yesterday_str}.xlsx"
        vyrubene_path = os.path.join(today_dir, vyrubene_filename)

        df.to_excel(vyrubene_path, index=False)
        log(f"💾 Druhý report uložený: {vyrubene_path}")

        cur_emsdwh.close()
        con_emsdwh_stdby.close()

        # Upload na SharePoint
        with open(vyrubene_path, 'rb') as file_obj:
            file_content = file_obj.read()
            folder = ctx.web.get_folder_by_server_relative_url(sharepoint_folder_path)
            upload_result = folder.upload_file(vyrubene_filename, file_content).execute_query()
            log(f'☁️ Súbor <b>{vyrubene_filename}</b> bol nahraný na SharePoint: <code>{upload_result.serverRelativeUrl}</code>')

    except Exception as e:
        log(f"❌ Chyba pri generovaní alebo nahrávaní Vyrubene_myto_kategorie: {e}")

    # ====== Tretí report: Vyrubene_myto_krajiny ==============================
    try:
        log("🟠 Spúšťam tretí report: Vyrubene_myto_krajiny...")

        try:
            EMS_DWH = EMS_DWHDB.objects.get(user_id=1)

            dsn = (
                f"{EMS_DWH.EMS_DWHDB_username}/"
                f"{EMS_DWH.EMS_DWHDB_password}@"
                f"{EMS_DWH.EMS_DWHDB_hostname}:"
                f"{EMS_DWH.EMS_DWHDB_port}/"
                f"{EMS_DWH.EMS_DWHDB_servicename}"
            )

            con_emsdwh_krajiny = cx_Oracle.connect(dsn, encoding="UTF-8")
            cur_emsdwh_krajiny = con_emsdwh_krajiny.cursor()

        except ObjectDoesNotExist:
            raise RuntimeError("❌ Záznam EMS_DWHDB pre user_id=1 sa nenašiel.")
        except cx_Oracle.Error as e:
            raise RuntimeError(f"❌ Chyba pri pripájaní na Oracle databázu pre krajiny: {e}")

        sql_file_path_krajiny = os.path.join(BASE_DIR, 'Scripts', 'Denne', 'Vyrubene_myto_krajiny.sql')
        try:
            with open(sql_file_path_krajiny, 'r', encoding='utf-8') as f:
                krajiny_sql = f.read()
        except UnicodeDecodeError:
            log("⚠️ UTF-8 dekódovanie zlyhalo. Skúšam cp1250 (Windows-1250)...")
            with open(sql_file_path_krajiny, 'r', encoding='cp1250') as f:
                krajiny_sql = f.read()

        cur_emsdwh_krajiny.execute(krajiny_sql)
        rows = cur_emsdwh_krajiny.fetchall()
        columns = [desc[0] for desc in cur_emsdwh_krajiny.description]

        import pandas as pd
        df = pd.DataFrame(rows, columns=columns)

        krajiny_filename = f"Vyrubene_myto_krajiny_{yesterday_str}.xlsx"
        krajiny_path = os.path.join(today_dir, krajiny_filename)

        df.to_excel(krajiny_path, index=False)
        log(f"💾 Tretí report uložený: {krajiny_path}")

        cur_emsdwh_krajiny.close()
        con_emsdwh_krajiny.close()

        # Upload na SharePoint
        with open(krajiny_path, 'rb') as file_obj:
            file_content = file_obj.read()
            folder = ctx.web.get_folder_by_server_relative_url(sharepoint_folder_path)
            upload_result = folder.upload_file(krajiny_filename, file_content).execute_query()
            log(f'☁️ Súbor <b>{krajiny_filename}</b> bol nahraný na SharePoint: <code>{upload_result.serverRelativeUrl}</code>')

    except Exception as e:
        log(f"❌ Chyba pri generovaní alebo nahrávaní Vyrubene_myto_krajiny: {e}")
  

    log(f" Denný reporting za {yesterday} úspešne dokončený.\n")


def ict_dph(request):
    # 1. IMPORTY na začiatku (rieši UnboundLocalError)
    import os
    import glob
    import pandas as pd
    import numpy as np
    import cx_Oracle
    import logging
    from django.utils.text import get_valid_filename
    from django.conf import settings

    # 2. NASTAVENIE CESTY A LOGOVANIA
    base_dir = os.path.join(settings.BASE_DIR, 'Reports', 'ICT_DPH')
    os.makedirs(base_dir, exist_ok=True)
    
    log_path = os.path.join(base_dir, 'auto.log')
    logger = logging.getLogger('ict_dph_logger')
    logger.setLevel(logging.INFO)

    if not logger.handlers:
        handler = logging.FileHandler(log_path, encoding='utf-8')
        formatter = logging.Formatter('%(asctime)s [%(levelname)s] %(message)s')
        handler.setFormatter(formatter)
        logger.addHandler(handler)

def ict_dph(request):
    # 1. Všetky potrebné importy na začiatku funkcie
    import os
    import glob
    import pandas as pd
    import numpy as np
    import cx_Oracle
    import logging
    from django.utils.text import get_valid_filename
    from django.conf import settings

    # 2. Nastavenie ciest a logovania
    base_dir = os.path.join(settings.BASE_DIR, 'Reports', 'ICT_DPH')
    os.makedirs(base_dir, exist_ok=True)
    
    log_path = os.path.join(base_dir, 'auto.log')
    logger = logging.getLogger('ict_dph_logger')
    logger.setLevel(logging.INFO)

    if not logger.handlers:
        handler = logging.FileHandler(log_path, encoding='utf-8')
        formatter = logging.Formatter('%(asctime)s [%(levelname)s] %(message)s')
        handler.setFormatter(formatter)
        logger.addHandler(handler)

    # ================= DB CONNECT + import dat do DWH DB =================

    if request.method == "POST" and "connect_db" in request.POST:
        import os
        import glob
        import pandas as pd
        import numpy as np
        import cx_Oracle
        import logging
        from django.utils.text import get_valid_filename
        from django.conf import settings

        con = None # Inicializácia pre exception block
        try:
            logger.info("=== Spúšťam import a generovanie reportov ===")
            
            mapping = {
                'Číslo PV': 'CISLO_PV',
                'Meno a priezvisko/Názov PV': 'MENO',
                'Typ MÚ': 'TYP_MU',
                'Číslo MÚ': 'CISLO_MU',
                'Daňová kategória PV': 'DANOVA_KATEGORIA',
                'Účtovný účet': 'UCTOVNY_UCET',
                'Prijímateľ platby': 'PRIJIMATEL_PLATBY',
                'Číslo faktúry': 'CISLO_FA',
                'Kategória faktúry': 'KATEGORIA_FA',
                'Dátum dodania': 'DATUM_DODANIA',
                'Dátum vystavenia': 'DATUM_VYSTAVENIA',
                'Dátum splatnosti': 'DATUM_SPLATNOSTI',
                'Základ dane v €': 'ZAKLAD_DANE',
                'Daň v €': 'DAN',
                'Fakturovaná suma v €': 'FAKTUROVANA_SUMA'
            }

            cols_order = list(mapping.values())

            EMS_DWH = EMS_DWHDB.objects.get(user=1)
            dsn = f"{EMS_DWH.EMS_DWHDB_username}/{EMS_DWH.EMS_DWHDB_password}@{EMS_DWH.EMS_DWHDB_hostname}:{EMS_DWH.EMS_DWHDB_port}/{EMS_DWH.EMS_DWHDB_servicename}"
            con = cx_Oracle.connect(dsn)
            cursor = con.cursor()

            # 1. Počiatočné vyčistenie tabuľky
            cursor.execute("DELETE FROM TMP_IC_DPH")
            con.commit()

            excel_files = glob.glob(os.path.join(base_dir, "*.xlsx"))
            excel_files = [f for f in excel_files if not os.path.basename(f).startswith('final_')]
            
            for file in excel_files:
                orig_fname = os.path.basename(file)
                logger.info(f"Spracovávam: {orig_fname}")
                
                # Načítanie: Explicitne povieme, že Číslo PV a Číslo MÚ sú TEXT (zachová nuly na začiatku)
                df = pd.read_excel(file, engine='openpyxl', dtype={'Číslo PV': str, 'Číslo MÚ': str})
                
                df = df.rename(columns=mapping)
                for col in cols_order:
                    if col not in df.columns: 
                        df[col] = None
                df = df[cols_order]

                def clean_value(val):
                    if pd.isna(val): return None
                    if hasattr(val, 'strftime'): return val.strftime('%d.%m.%Y')
                    
                    # Ak je to už string (vďaka dtype v read_excel), len otrimujeme
                    if isinstance(val, str):
                        return val.strip()
                    
                    # Ak je to číslo, ktoré prišlo inak (float/int)
                    if isinstance(val, (int, float)):
                        if float(val).is_integer(): 
                            return str(int(val))
                        return str(val)
                    
                    return str(val).strip()

                df = df.map(clean_value)

                # Príprava insertu
                sql_insert = f"""
                    INSERT INTO TMP_IC_DPH ({', '.join(cols_order)}) 
                    VALUES (:1, :2, :3, :4, :5, :6, :7, :8, :9, :10, :11, :12, :13, :14, :15)
                """
                
                cursor.executemany(sql_insert, [tuple(x) for x in df.values])
                con.commit()

                # 2. GENEROVANIE FINAL REPORTU
                logger.info(f"Generujem report pre: {orig_fname}")
                
                sql_report = """
                    SELECT
                        tmp.cislo_pv as "Číslo PV",
                        cus.ORGANIZATION_ID,
                        cus.ORGANIZATION_TAX_ID,
                        cus.ORGANIZATION_VAT_ID,
                        tmp.meno as "Meno a priezvisko/Názov PV",
                        tmp.typ_mu as "Typ MÚ",
                        tmp.cislo_mu as "Číslo MÚ",
                        tmp.danova_kategoria as "Daňová kategória PV",
                        tmp.uctovny_ucet as "Účtovný účet",
                        tmp.prijimatel_platby as "Prijímateľ platby",
                        tmp.cislo_fa as "Číslo faktúry",
                        tmp.kategoria_fa as "Kategória faktúry",
                        tmp.datum_dodania as "Dátum dodania",
                        tmp.datum_vystavenia as "Dátum vystavenia",
                        tmp.datum_splatnosti as "Dátum splatnosti",
                        tmp.zaklad_dane as "Základ dane v €",
                        tmp.dan as "Daň v €",
                        tmp.fakturovana_suma as "Fakturovaná suma v €"
                    FROM
                        tmp_ic_dph tmp
                    LEFT JOIN billien_maa.customers cus 
                        ON tmp.cislo_pv = cus.CUSTOMER_NUMBER
                """
                
                cursor.execute(sql_report)
                rows = cursor.fetchall()
                colnames = [desc[0] for desc in cursor.description]
                df_final = pd.DataFrame(rows, columns=colnames)

                if df_final.empty:
                    logger.warning(f"POZOR: SQL dopyt vrátil 0 riadkov pre {orig_fname}!")
                else:
                    logger.info(f"Report obsahuje {len(df_final)} riadkov.")

                # Uloženie do nového Excelu
                final_fname = f"final_{orig_fname}"
                final_path = os.path.join(base_dir, final_fname)
                df_final.to_excel(final_path, index=False, sheet_name='Report')
                
                logger.info(f"Report uložený: {final_fname}")
                
                # Vymazanie TMP tabuľky po každom súbore
                cursor.execute("DELETE FROM TMP_IC_DPH")
                con.commit()

            cursor.close()
            con.close()
            logger.info("=== Proces úspešne dokončený ===")

        except Exception as e:
            if con:
                con.rollback()
                con.close()
            logger.exception(f"Chyba: {e}")


    # ================= EXCEL UPLOAD =================
    if request.method == "POST" and "upload_excel" in request.POST:
        try:
            logger.info("=== Kliknutý button: Upload Excel súboru ===")

            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                raise ValueError("Excel súbor nebol priložený")

            filename = get_valid_filename(excel_file.name)
            file_path = os.path.join(base_dir, filename)

            with open(file_path, 'wb+') as destination:
                for chunk in excel_file.chunks():
                    destination.write(chunk)

            logger.info(f"Excel úspešne uložený: {filename}")

        except Exception as e:
            logger.exception(f"Upload Excel chyba: {e}")

    # ================= DELETE FILE =================
    if request.method == "POST" and "delete_file" in request.POST:
        try:
            filename = request.POST.get("filename")
            file_path = os.path.join(base_dir, filename)

            # bezpečnostná kontrola
            if not os.path.abspath(file_path).startswith(os.path.abspath(base_dir)):
                raise ValueError("Neplatná cesta k súboru")

            if os.path.exists(file_path):
                os.remove(file_path)
                logger.info(f"Súbor zmazaný: {filename}")
            else:
                logger.warning(f"Súbor neexistuje: {filename}")

        except Exception as e:
            logger.exception(f"Delete súbor chyba: {e}")

        # ================= DOWNLOAD FILE =================
    if request.method == "POST" and "download_file" in request.POST:
        try:
            filename = request.POST.get("filename")
            file_path = os.path.join(base_dir, filename)

            # bezpečnostná kontrola
            if not os.path.abspath(file_path).startswith(os.path.abspath(base_dir)):
                raise ValueError("Neplatná cesta k súboru")

            if not os.path.exists(file_path):
                raise FileNotFoundError("Súbor neexistuje")

            logger.info(f"Súbor stiahnutý: {filename}")

            return FileResponse(
                open(file_path, 'rb'),
                as_attachment=True,
                filename=filename
            )

        except Exception as e:
            logger.exception(f"Download súbor chyba: {e}")
            raise Http404("Súbor sa nepodarilo stiahnuť")

    # ================= DOWNLOAD LOG =================
    if request.method == "POST" and "download_log" in request.POST:
        try:
            if not os.path.exists(log_path):
                raise FileNotFoundError("Log súbor neexistuje")

            logger.info("Log súbor stiahnutý")

            return FileResponse(
                open(log_path, 'rb'),
                as_attachment=True,
                filename='auto.log'
            )

        except Exception as e:
            logger.exception(f"Download log chyba: {e}")
            raise Http404("Log súbor sa nepodarilo stiahnuť")

    # ================= ZOZNAM SÚBOROV =================
    try:
        files = [
            f for f in os.listdir(base_dir)
            if os.path.isfile(os.path.join(base_dir, f)) and f != 'auto.log'
        ]
    except Exception as e:
        logger.exception(f"Chyba pri čítaní súborov: {e}")

    # ================= ČÍTANIE LOGU =================
    try:
        if os.path.exists(log_path):
            with open(log_path, 'r', encoding='utf-8') as f:
                log_lines = f.readlines()[-30:]
        else:
            log_lines = ["⚠️ Súbor auto.log neexistuje"]
    except Exception as e:
        log_lines = [f"⚠️ Chyba pri čítaní logu: {e}"]

    return render(
        request,
        'ems/ict_dph.html',
        {
            'log_lines': log_lines,
            'files': files
        }
    )


